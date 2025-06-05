import os
import sys
import json
import csv
import logging
import calendar
import platform
import threading
import shutil
import webbrowser
import warnings
import zipfile
import traceback
import time
import hashlib  # New for v2.0.0
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Union, Any
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.utils import formatdate
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import ttkthemes
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')  # Must be before importing pyplot
import matplotlib.pyplot as plt
import matplotlib.font_manager as font_manager
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg,
    NavigationToolbar2Tk
)
from matplotlib.patches import Rectangle
from matplotlib import colors
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from reportlab.lib import colors as reportlab_colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image
)
import serial
import serial.tools.list_ports
import requests
from requests.exceptions import RequestException
import urllib3
from github import Github  # New for v2.0.0
from cryptography.fernet import Fernet
from base64 import b64encode, b64decode

