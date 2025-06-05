#!/usr/bin/env python3

# Standard library imports
import os
import sys
import json
import csv
import logging
import calendar
import platform
import threading
import shutil
import webbrowser
import warnings
import zipfile
import traceback
import time
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Union, Any

# Email handling imports
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.utils import formatdate

# GUI imports
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import ttkthemes

# Data processing and visualization imports
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')  # Must be before importing pyplot
import matplotlib.pyplot as plt
import matplotlib.font_manager as font_manager
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg,
    NavigationToolbar2Tk
)
from matplotlib.patches import Rectangle
from matplotlib import colors

# Excel handling imports
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter

# PDF handling imports
from reportlab.lib import colors as reportlab_colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image
)

# Serial communication imports
import serial
import serial.tools.list_ports

# Network and API imports
import requests
from requests.exceptions import RequestException
import urllib3

# Encryption imports
from cryptography.fernet import Fernet
from base64 import b64encode, b64decode

# Suppress specific warnings
warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib")
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Configure matplotlib parameters
plt.style.use('default')
matplotlib.rcParams.update({
    'font.family': 'sans-serif',
    'font.sans-serif': ['Arial', 'Helvetica', 'DejaVu Sans'],
    'font.size': 10,
    'axes.labelsize': 12,
    'axes.titlesize': 14,
    'xtick.labelsize': 10,
    'ytick.labelsize': 10,
    'figure.titlesize': 16,
    'figure.dpi': 100,
    'figure.autolayout': True,
    'figure.facecolor': '#FFFFFF',
    'axes.facecolor': 'white',
    'grid.color': '#E0E0E0',
    'grid.linestyle': '--',
    'grid.linewidth': 0.5,
    'axes.grid': True,
    'axes.edgecolor': '#666666',
    'axes.linewidth': 1.0,
    'lines.linewidth': 2.0,
    'lines.markersize': 6,
    'xtick.color': 'black',
    'ytick.color': 'black',
    'axes.labelcolor': 'black',
    'savefig.dpi': 300,
    'savefig.bbox': 'tight',
    'savefig.pad_inches': 0.1
})

# Disable font cache to prevent warnings
matplotlib.font_manager._get_font.cache_clear()

# Set matplotlib logging level
logging.getLogger('matplotlib').setLevel(logging.WARNING)

# Fallback font configuration
if platform.system() == 'Windows':
    plt.rcParams['font.family'] = 'Arial'
else:
    plt.rcParams['font.family'] = 'DejaVu Sans'


# Suppress matplotlib font warnings
warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib")

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/pool_app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Create necessary directories
for directory in ['logs', 'data', 'backup', 'temp']:
    Path(directory).mkdir(exist_ok=True)

# Version information
__version__ = "2.0.0"
__author__ = "Virtual Control LLC"
__copyright__ = "Copyright © 2025 Virtual Control LLC"

# Constants
DEFAULT_WINDOW_SIZE = (1400, 1000)
MIN_WINDOW_SIZE = (1200, 800)
WEATHER_API_KEY = "0dd8ec60f736447f92611037253005"

# Chemical ranges and thresholds
CHEMICAL_RANGES = {
    'pH': (7.2, 7.8),
    'Free Chlorine': (1.0, 3.0),
    'Total Chlorine': (1.0, 3.0),
    'Alkalinity': (80, 120),
    'Hardness': (200, 400),
    'Cyanuric Acid': (30, 80),
    'Salt': (2700, 3400)
}

DISPLAY_RANGES = {
    'pH': (6.0, 8.0),
    'Free Chlorine': (0.0, 5.0),
    'Total Chlorine': (0.0, 5.0),
    'Alkalinity': (0, 240),
    'Hardness': (0, 1000),
    'Cyanuric Acid': (0, 100),
    'Salt': (0, 5000)
}

# Color schemes
COLOR_SCHEME = {
    'primary': "#2196F3",    # Material Blue
    'secondary': "#BBDEFB",  # Light Blue
    'success': "#4CAF50",    # Green
    'warning': "#FFC107",    # Yellow
    'danger': "#F44336",     # Red
    'info': "#03A9F4",       # Light Blue
    'background': "#F5F5F5"  # Light Gray
}

# Graph colors
GRAPH_COLORS = {
    'pH': "#2196F3",
    'Chlorine': "#4CAF50",
    'Alkalinity': "#FFC107",
    'Hardness': "#9C27B0",
    'Stabilizer': "#FF5722"
}

# Chemical adjustment factors
CHEMICAL_FACTORS = {
    'Chlorine': {
        'pH': 0.5,
        'Free Chlorine': 1.0,
        'Alkalinity': 0.1
    },
    'Salt': {
        'pH': 0.5,
        'Salt': 1.0,
        'Alkalinity': 0.1
    },
    'Bromine': {
        'pH': 0.5,
        'Bromine': 1.0,
        'Alkalinity': 0.1
    }
}

# Chemical treatment options
INCREASE_CHEMICALS = {
    'pH': 'Soda Ash',
    'Free Chlorine': 'Liquid Chlorine',
    'Total Chlorine': 'Shock Treatment',
    'Alkalinity': 'Baking Soda',
    'Hardness': 'Calcium Chloride',
    'Cyanuric Acid': 'Stabilizer',
    'Salt': 'Pool Salt'
}

DECREASE_CHEMICALS = {
    'pH': 'Muriatic Acid',
    'Free Chlorine': 'Sodium Thiosulfate',
    'Total Chlorine': 'Sodium Thiosulfate',
    'Alkalinity': 'Muriatic Acid',
    'Hardness': 'Partial Water Change',
    'Cyanuric Acid': 'Partial Water Change',
    'Salt': 'Partial Water Change'
}

# Time intervals (in seconds)
TIME_INTERVALS = {
    'WEATHER_UPDATE': 1800,  # 30 minutes
    'TASK_CHECK': 60,        # 1 minute
    'ALERT_CHECK': 300,      # 5 minutes
    'DATA_BACKUP': 3600      # 1 hour
}

class WaterTester:
    def __init__(self):
        """Initialize WaterTester with industry standard ranges"""
        logger.info("Initializing WaterTester")
        try:
            # Initialize pool types
            self.pool_types = [
                "Chlorine Pool",
                "Salt Water Pool", 
                "Bromine Pool",
                "Concrete/Gunite",
                "Vinyl Liner",
                "Fiberglass",
                "Above Ground - Metal Wall",
                "Above Ground - Resin",
                "Infinity Pool",
                "Lap Pool",
                "Indoor Pool"
            ]

            # Initialize ideal chemical levels
            self.ideal_levels = {
                "Hardness": (200, 400),      # Range 0-800
                "Free Chlorine": (1.0, 3.0),  # Range 0-6.0
                "Total Chlorine": (1.0, 3.0), # Range 0-6.0
                "Bromine": (3.0, 5.0),        # Range 0-10.0
                "Alkalinity": (80, 120),      # Range 0-240
                "pH": (7.2, 7.8),             # Range 6.2-8.4
                "Cyanuric Acid": (30, 80),    # Range 0-100
                "Salt": (2700, 3400)          # Range 0-6800
            }

            # Initialize chemicals for increasing levels
            self.increase_chemicals = {
                "pH": "Soda Ash",
                "Free Chlorine": "Liquid Chlorine",
                "Total Chlorine": "Shock Treatment",
                "Alkalinity": "Baking Soda",
                "Hardness": "Calcium Chloride",
                "Cyanuric Acid": "Stabilizer",
                "Salt": "Pool Salt",
                "Bromine": "Bromine Tablets"
            }

            # Initialize chemicals for decreasing levels
            self.decrease_chemicals = {
                "pH": "Muriatic Acid",
                "Free Chlorine": "Sodium Thiosulfate",
                "Total Chlorine": "Sodium Thiosulfate",
                "Alkalinity": "Muriatic Acid",
                "Hardness": "Partial Water Change",
                "Cyanuric Acid": "Partial Water Change",
                "Salt": "Partial Water Change",
                "Bromine": "Partial Water Change"
            }

            # Initialize chemical factors
            self.chemical_factors = {
                "Chlorine Pool": {
                    "pH": 1.0,
                    "Free Chlorine": 1.0,
                    "Total Chlorine": 1.0,
                    "Alkalinity": 1.0,
                    "Hardness": 1.0,
                    "Cyanuric Acid": 1.0
                },
                "Salt Water Pool": {
                    "pH": 1.0,
                    "Free Chlorine": 0.8,
                    "Total Chlorine": 0.8,
                    "Alkalinity": 1.0,
                    "Hardness": 1.0,
                    "Cyanuric Acid": 1.0,
                    "Salt": 1.0
                },
                "Bromine Pool": {
                    "pH": 1.0,
                    "Bromine": 1.0,
                    "Alkalinity": 1.0,
                    "Hardness": 1.0
                }
            }

            # Add default factors for other pool types
            default_factors = {
                "pH": 1.0,
                "Free Chlorine": 1.0,
                "Total Chlorine": 1.0,
                "Alkalinity": 1.0,
                "Hardness": 1.0,
                "Cyanuric Acid": 1.0
            }

            # Add factors for remaining pool types
            for pool_type in [pt for pt in self.pool_types if pt not in self.chemical_factors]:
                self.chemical_factors[pool_type] = default_factors.copy()

            # Initialize chemical warnings
            self.chemical_warnings = {
                "Free Chlorine": """
SAFETY WARNING: 
- Always add chlorine to water, never water to chlorine
- Wear protective gloves and eye protection
- Add during evening hours to prevent rapid dissipation
- Keep away from children and pets
- Never mix with other chemicals
- Ensure proper ventilation when handling
- Store in a cool, dry, well-ventilated area""",
                "Total Chlorine": """
SAFETY WARNING:
- Follow same precautions as Free Chlorine
- Monitor levels carefully after shock treatment
- Wait for levels to normalize before swimming""",
                "pH": """
SAFETY WARNING:
- Handle pH adjusters with care
- Add chemicals slowly and test frequently
- Avoid splashing and inhaling fumes
- Keep chemicals separate""",
                "Alkalinity": """
SAFETY WARNING:
- Add chemicals gradually
- Test frequently during adjustment
- Allow for proper mixing
- Maintain proper pH levels""",
                "Cyanuric Acid": """
SAFETY WARNING:
- Add through skimmer or pre-dissolve
- Avoid direct contact with pool surface
- Monitor levels carefully
- Do not exceed recommended levels""",
                "Salt": """
SAFETY WARNING:
- Add salt gradually
- Allow for complete dissolution
- Verify proper cell operation
- Monitor stabilizer levels"""
            }

            logger.info("WaterTester initialized successfully")

        except Exception as e:
            logger.error(f"Error initializing WaterTester: {str(e)}")
            raise

    def _validate_ideal_levels(self):
        """Validate all ideal levels are properly configured"""
        try:
            for chemical, (min_val, max_val) in self.ideal_levels.items():
                if not isinstance(min_val, (int, float)) or not isinstance(max_val, (int, float)):
                    raise DataValidationError(f"Invalid ideal level values for {chemical}")
                if min_val >= max_val:
                    raise DataValidationError(f"Minimum value greater than or equal to maximum for {chemical}")
                if min_val < 0:
                    raise DataValidationError(f"Negative minimum value for {chemical}")
                    
            # Validate pool type ranges
            for pool_type, ranges in self.pool_type_ranges.items():
                for param, (min_val, max_val) in ranges.items():
                    if not isinstance(min_val, (int, float)) or not isinstance(max_val, (int, float)):
                        raise DataValidationError(f"Invalid range values for {pool_type} {param}")
                    if min_val >= max_val:
                        raise DataValidationError(f"Invalid range for {pool_type} {param}")
                        
            logger.info("Ideal levels validated successfully")
            
        except Exception as e:
            logger.error(f"Error validating ideal levels: {str(e)}")
            raise

    def _initialize_chemical_warnings(self):
        """Initialize chemical safety warnings"""
        return {
            "Free Chlorine (ppm)": """
    SAFETY WARNING: 
    - Always add chlorine to water, never water to chlorine
    - Wear protective gloves and eye protection
    - Add during evening hours to prevent rapid dissipation
    - Keep away from children and pets
    - Never mix with other chemicals
    - Ensure proper ventilation when handling
    - Store in a cool, dry, well-ventilated area""",

                "Total Chlorine (ppm)": """
    SAFETY WARNING:
    - Always add chlorine to water, never water to chlorine
    - Wear protective gloves and eye protection
    - Add during evening hours to prevent rapid dissipation
    - Keep away from children and pets
    - Never mix with other chemicals
    - Ensure proper ventilation when handling
    - Store in a cool, dry, well-ventilated area""",

                # ... [other chemical warnings remain the same] ...
            }

    def set_pool_volume(self, volume: float) -> None:
        """
        Set pool volume with validation
        
        Args:
            volume (float): Pool volume in gallons
            
        Raises:
            DataValidationError: If volume is invalid
        """
        try:
            if not isinstance(volume, (int, float)):
                raise DataValidationError("Pool volume must be a number", value=volume)
            if volume <= 0:
                raise DataValidationError("Pool volume must be greater than 0", value=volume)
            if volume > 1000000:  # Reasonable upper limit
                raise DataValidationError("Pool volume seems unreasonably high", value=volume)
            
            self.pool_volume = volume
            logger.info(f"Pool volume set to {volume:,} gallons")
            
        except Exception as e:
            logger.error(f"Error setting pool volume: {str(e)}")
            raise

    def get_chemical_recommendations(self, current_levels: Dict[str, float], 
                                   pool_type: str, 
                                   weather: Optional[Dict] = None) -> Dict[str, str]:
        """
        Generate chemical recommendations based on pool type and current levels
        
        Args:
            current_levels (Dict[str, float]): Current chemical levels
            pool_type (str): Type of pool
            weather (Optional[Dict]): Weather data if available
            
        Returns:
            Dict[str, str]: Recommendations for each chemical
            
        Raises:
            ValueError: If inputs are invalid
        """
        try:
            # Validate inputs
            if not current_levels:
                raise ValueError("Current levels cannot be empty")
            if not pool_type:
                raise ValueError("Pool type must be specified")
            if pool_type not in self.pool_types:
                raise ValueError(f"Invalid pool type: {pool_type}")

            recommendations = {}
            
            # Get pool-type specific factors
            factors = self.chemical_factors.get(pool_type, self.chemical_factors["Chlorine Pool"])
            
            for param, current in current_levels.items():
                if param not in self.ideal_levels:
                    continue

                ideal_min, ideal_max = self.ideal_levels[param]
                
                if current < ideal_min:
                    adjustment = self._calculate_increase(param, current, ideal_min, factors)
                    recommendations[param] = adjustment
                elif current > ideal_max:
                    adjustment = self._calculate_decrease(param, current, ideal_max, factors)
                    recommendations[param] = adjustment
                else:
                    recommendations[param] = "Level is ideal"

            # Add weather-specific recommendations
            if weather:
                self._add_weather_recommendations(recommendations, weather)

            return recommendations
            
        except Exception as e:
            logger.error(f"Error generating recommendations: {str(e)}")
            raise

    def _calculate_increase(self, param: str, current: float, target: float, 
                           factors: Dict[str, float]) -> str:
        """
        Calculate chemical increase needed
        
        Args:
            param (str): Chemical parameter
            current (float): Current level
            target (float): Target level
            factors (Dict[str, float]): Chemical factors for pool type
            
        Returns:
            str: Recommended adjustment
        """
        try:
            difference = target - current
            factor = factors.get(param, 1.0)
            
            if param not in self.increase_chemicals:
                return f"No increase method available for {param}"
                
            chemical = self.increase_chemicals[param]
            amount = difference * factor
            
            if amount <= 0:
                return "No adjustment needed"
                
            warning = self.chemical_warnings.get(param, "")
            
            return f"Add {amount:.2f} units of {chemical}\n\n{warning}"
            
        except Exception as e:
            logger.error(f"Error calculating increase: {str(e)}")
            raise

    def _calculate_decrease(self, param: str, current: float, target: float, 
                           factors: Dict[str, float]) -> str:
        """
        Calculate chemical decrease needed
        
        Args:
            param (str): Chemical parameter
            current (float): Current level
            target (float): Target level
            factors (Dict[str, float]): Chemical factors for pool type
            
        Returns:
            str: Recommended adjustment
        """
        try:
            difference = current - target
            factor = factors.get(param, 1.0)
            
            if param not in self.decrease_chemicals:
                return f"No decrease method available for {param}"
                
            chemical = self.decrease_chemicals[param]
            amount = difference * factor
            
            if amount <= 0:
                return "No adjustment needed"
                
            warning = self.chemical_warnings.get(param, "")
            
            return f"Add {amount:.2f} units of {chemical}\n\n{warning}"
            
        except Exception as e:
            logger.error(f"Error calculating decrease: {str(e)}")
            raise

    def get_safety_warning(self, chemical: str) -> str:
        """
        Get safety warning for specific chemical
        
        Args:
            chemical (str): Chemical name
            
        Returns:
            str: Safety warning text
        """
        try:
            return self.chemical_warnings.get(chemical, "No specific safety warnings available.")
        except Exception as e:
            logger.error(f"Error getting safety warning: {str(e)}")
            return "Error retrieving safety warning."




   
        """
        Set pool volume with validation
        
        Args:
            volume (float): Pool volume in gallons
            
        Raises:
            DataValidationError: If volume is invalid
        """
        try:
            if not isinstance(volume, (int, float)):
                raise DataValidationError("Pool volume must be a number", value=volume)
            if volume <= 0:
                raise DataValidationError("Pool volume must be greater than 0", value=volume)
            if volume > 1000000:  # Reasonable upper limit
                raise DataValidationError("Pool volume seems unreasonably high", value=volume)
            
            self.pool_volume = volume
            logger.info(f"Pool volume set to {volume:,} gallons")
            
        except Exception as e:
            logger.error(f"Error setting pool volume: {str(e)}")
            raise

    def _add_weather_recommendations(self, recommendations: Dict[str, str], 
                                   weather: Dict) -> None:
        """Add weather-specific recommendations"""
        try:
            temp = weather.get('temp', 0)
            uv_index = weather.get('uv_index', 0)
            
            if temp > 85:
                recommendations['Weather'] = (
                    "High temperature detected:\n"
                    "- Increase testing frequency\n"
                    "- Monitor chlorine levels more closely\n"
                    "- Consider adding extra stabilizer"
                )
            
            if uv_index > 8:
                if 'Weather' in recommendations:
                    recommendations['Weather'] += (
                        "\nHigh UV index detected:\n"
                        "- Add extra chlorine stabilizer\n"
                        "- Test chlorine levels more frequently"
                    )
                else:
                    recommendations['Weather'] = (
                        "High UV index detected:\n"
                        "- Add extra chlorine stabilizer\n"
                        "- Test chlorine levels more frequently"
                    )
                    
        except Exception as e:
            logger.error(f"Error adding weather recommendations: {str(e)}")
            raise

    def _validate_chemical_levels(self, levels: Dict[str, float]) -> Tuple[bool, List[str]]:
        """
        Validate chemical levels against acceptable ranges
        
        Args:
            levels (Dict[str, float]): Chemical levels to validate
            
        Returns:
            Tuple[bool, List[str]]: (is_valid, error_messages)
        """
        try:
            errors = []
            
            for param, value in levels.items():
                if param not in self.ideal_levels:
                    errors.append(f"Unknown parameter: {param}")
                    continue
                    
                if not isinstance(value, (int, float)):
                    errors.append(f"Invalid value type for {param}: {type(value)}")
                    continue
                    
                if value < 0:
                    errors.append(f"Negative value for {param}: {value}")
                    continue
                    
                # Check against reasonable limits
                min_val, max_val = self.ideal_levels[param]
                absolute_max = max_val * 2  # Allow up to double the max ideal value
                
                if value > absolute_max:
                    errors.append(f"Value for {param} exceeds reasonable maximum: {value}")
                    
            return len(errors) == 0, errors
            
        except Exception as e:
            logger.error(f"Error validating chemical levels: {str(e)}")
            return False, [str(e)]

    def calculate_dosage(self, chemical: str, amount_needed: float) -> str:
        """
        Calculate chemical dosage based on pool volume
        
        Args:
            chemical (str): Chemical name
            amount_needed (float): Amount needed in ppm
            
        Returns:
            str: Dosage instructions
            
        Raises:
            ValueError: If chemical is invalid or amount is negative
        """
        try:
            # Validate inputs
            if chemical not in self.ideal_levels:
                raise ValueError(f"Invalid chemical: {chemical}")
            if amount_needed <= 0:
                return "No adjustment needed"
            
            # Get the appropriate calculation method
            calculation_methods = {
                "Free Chlorine (ppm)": self._calculate_chlorine_dosage,
                "Total Chlorine (ppm)": self._calculate_chlorine_dosage,
                "Bromine (ppm)": self._calculate_bromine_dosage,
                "pH": self._calculate_ph_dosage,
                "Alkalinity (ppm)": self._calculate_alkalinity_dosage,
                "Hardness (ppm)": self._calculate_hardness_dosage,
                "Stabilizer (ppm)": self._calculate_stabilizer_dosage,
                "Salt (ppm)": self._calculate_salt_dosage
            }
            
            # Calculate dosage using appropriate method
            if chemical in calculation_methods:
                return calculation_methods[chemical](amount_needed)
            
            return "Dosage calculation not available for this chemical"
            
        except Exception as e:
            logger.error(f"Error calculating dosage: {str(e)}")
            raise

    def _calculate_chlorine_dosage(self, ppm_needed: float) -> str:
        """
        Calculate chlorine dosage with validation
        
        Args:
            ppm_needed (float): Amount of chlorine needed in ppm
            
        Returns:
            str: Formatted dosage instruction
            
        Raises:
            ValueError: If calculation results in unreasonable values
        """
        try:
            if ppm_needed <= 0:
                raise ValueError("PPM needed must be positive")
                
            gallons_needed = (ppm_needed * self.pool_volume) / (12.5 * 10000)
            oz_needed = gallons_needed * 128
            
            if oz_needed > 1000:  # Sanity check
                raise ValueError("Calculated dosage seems unreasonably high")
                
            return f"{oz_needed:.1f} oz liquid chlorine"
            
        except Exception as e:
            logger.error(f"Error calculating chlorine dosage: {str(e)}")
            raise

    def _calculate_ph_dosage(self, ppm_needed: float) -> str:
        """Calculate pH adjustment dosage with validation"""
        try:
            if ppm_needed == 0:
                return "No pH adjustment needed"
                
            if ppm_needed > 0:
                oz_needed = (ppm_needed * self.pool_volume * 6) / (0.2 * 10000)
                if oz_needed > 500:  # Sanity check
                    raise ValueError("Calculated pH increaser dosage seems unreasonably high")
                return f"{oz_needed:.1f} oz pH increaser (soda ash)"
            else:
                oz_needed = (abs(ppm_needed) * self.pool_volume * 12) / (0.2 * 10000)
                if oz_needed > 500:  # Sanity check
                    raise ValueError("Calculated pH decreaser dosage seems unreasonably high")
                return f"{oz_needed:.1f} oz pH decreaser (muriatic acid)"
                
        except Exception as e:
            logger.error(f"Error calculating pH dosage: {str(e)}")
            raise

    def get_ideal_range(self, parameter: str) -> Tuple[float, float]:
        """
        Get ideal range for a chemical parameter
        
        Args:
            parameter (str): Chemical parameter
            
        Returns:
            Tuple[float, float]: (minimum, maximum) ideal values
            
        Raises:
            ValueError: If parameter is invalid
        """
        try:
            if parameter not in self.ideal_levels:
                raise ValueError(f"Unknown parameter: {parameter}")
                
            return self.ideal_levels[parameter]
            
        except Exception as e:
            logger.error(f"Error getting ideal range: {str(e)}")
            raise

    def _calculate_alkalinity_dosage(self, ppm_needed: float) -> str:
        """Calculate alkalinity dosage with validation"""
        try:
            if ppm_needed <= 0:
                raise ValueError("PPM needed must be positive")
                
            lbs_needed = (ppm_needed * self.pool_volume * 1.5) / (10 * 10000)
            
            if lbs_needed > 100:  # Sanity check
                raise ValueError("Calculated dosage seems unreasonably high")
                
            return f"{lbs_needed:.2f} lbs sodium bicarbonate"
            
        except Exception as e:
            logger.error(f"Error calculating alkalinity dosage: {str(e)}")
            raise

    def _calculate_hardness_dosage(self, ppm_needed: float) -> str:
        """Calculate hardness dosage with validation"""
        try:
            if ppm_needed <= 0:
                raise ValueError("PPM needed must be positive")
                
            lbs_needed = (ppm_needed * self.pool_volume) / (10 * 10000)
            
            if lbs_needed > 100:  # Sanity check
                raise ValueError("Calculated dosage seems unreasonably high")
                
            return f"{lbs_needed:.2f} lbs calcium chloride"
            
        except Exception as e:
            logger.error(f"Error calculating hardness dosage: {str(e)}")
            raise

    def _calculate_cyanuric_acid_dosage(self, ppm_needed: float) -> str:
        """Calculate cyanuric acid dosage with validation"""
        try:
            if ppm_needed <= 0:
                raise ValueError("PPM needed must be positive")
                
            oz_needed = (ppm_needed * self.pool_volume * 13) / (10 * 10000)
            
            if oz_needed > 200:  # Sanity check
                raise ValueError("Calculated dosage seems unreasonably high")
                
            return f"{oz_needed:.1f} oz cyanuric acid"
            
        except Exception as e:
            logger.error(f"Error calculating cyanuric acid dosage: {str(e)}")
            raise

    def _calculate_salt_dosage(self, ppm_needed: float) -> str:
        """Calculate salt dosage with validation"""
        try:
            if ppm_needed <= 0:
                raise ValueError("PPM needed must be positive")
                
            lbs_needed = (ppm_needed * self.pool_volume * 80) / (1000 * 10000)
            
            if lbs_needed > 1000:  # Sanity check
                raise ValueError("Calculated dosage seems unreasonably high")
                
            return f"{lbs_needed:.1f} lbs salt"
            
        except Exception as e:
            logger.error(f"Error calculating salt dosage: {str(e)}")
            raise

    def _calculate_bromine_dosage(self, ppm_needed: float) -> str:
        """Calculate bromine dosage with validation"""
        try:
            if ppm_needed <= 0:
                raise ValueError("PPM needed must be positive")
                
            tablets_needed = (ppm_needed * self.pool_volume) / (1.5 * 10000)
            
            if tablets_needed > 100:  # Sanity check
                raise ValueError("Calculated dosage seems unreasonably high")
                
            return f"{tablets_needed:.1f} bromine tablets"
            
        except Exception as e:
            logger.error(f"Error calculating bromine dosage: {str(e)}")
            raise

    def get_chemical_status(self, param: str, value: float) -> Tuple[str, str]:
        """
        Get status and color code for chemical level
        
        Args:
            param (str): Chemical parameter
            value (float): Current value
            
        Returns:
            Tuple[str, str]: (status, color_code)
        """
        try:
            if param not in self.ideal_levels:
                raise ValueError(f"Unknown parameter: {param}")
                
            ideal_min, ideal_max = self.ideal_levels[param]
            
            if value < ideal_min:
                percent_low = ((ideal_min - value) / ideal_min) * 100
                if percent_low > 25:
                    return "VERY LOW", "#FF0000"  # Red
                return "LOW", "#FFA500"  # Orange
            elif value > ideal_max:
                percent_high = ((value - ideal_max) / ideal_max) * 100
                if percent_high > 25:
                    return "VERY HIGH", "#FF0000"  # Red
                return "HIGH", "#FFA500"  # Orange
            else:
                return "IDEAL", "#008000"  # Green
                
        except Exception as e:
            logger.error(f"Error getting chemical status: {str(e)}")
            raise

    def format_chemical_value(self, param: str, value: float) -> str:
        """
        Format chemical value with appropriate units
        
        Args:
            param (str): Chemical parameter
            value (float): Value to format
            
        Returns:
            str: Formatted value with units
        """
        try:
            if param == "pH":
                return f"{value:.1f}"
            elif param in ["Free Chlorine", "Total Chlorine", "Bromine"]:
                return f"{value:.1f} ppm"
            elif param == "Salt":
                return f"{int(value)} ppm"
            else:
                return f"{int(value)} ppm"
                
        except Exception as e:
            logger.error(f"Error formatting chemical value: {str(e)}")
            raise

    def get_adjustment_priority(self, param: str, value: float) -> int:
        """
        Get priority level for chemical adjustment
        
        Args:
            param (str): Chemical parameter
            value (float): Current value
            
        Returns:
            int: Priority level (1-5, 1 being highest)
        """
        try:
            if param not in self.ideal_levels:
                return 5  # Lowest priority for unknown parameters
                
            ideal_min, ideal_max = self.ideal_levels[param]
            
            # Priority order: pH > Chlorine > Alkalinity > Others
            base_priority = {
                "pH": 1,
                "Free Chlorine": 2,
                "Total Chlorine": 2,
                "Alkalinity": 3,
                "Cyanuric Acid": 4,
                "Others": 5
            }.get(param, 5)
            
            # Adjust priority based on deviation from ideal range
            if value < ideal_min:
                deviation = (ideal_min - value) / ideal_min
            elif value > ideal_max:
                deviation = (value - ideal_max) / ideal_max
            else:
                return 5  # Lowest priority if in ideal range
                
            # Increase priority for severe deviations
            if deviation > 0.5:  # More than 50% off
                return max(1, base_priority - 1)
                
            return base_priority
            
        except Exception as e:
            logger.error(f"Error calculating adjustment priority: {str(e)}")
            return 5  # Return lowest priority on error


        """
        Get safety warning for specific chemical
        
        Args:
            chemical (str): Chemical name
            
        Returns:
            str: Safety warning text
        """
        try:
            return self.chemical_warnings.get(chemical, "No specific safety warnings available.")
        except Exception as e:
            logger.error(f"Error getting safety warning: {str(e)}")
            return "Error retrieving safety warning."

class ToolTip:
    """Create tooltips for widgets"""
    
    def __init__(self, widget: tk.Widget, text: str):
        """
        Initialize tooltip
        
        Args:
            widget (tk.Widget): Widget to attach tooltip to
            text (str): Tooltip text
        """
        self.widget = widget
        self.text = text
        self.tooltip = None
        self.id = None
        self.x = 0
        self.y = 0

        # Bind events
        self.widget.bind('<Enter>', self.enter)
        self.widget.bind('<Leave>', self.leave)
        self.widget.bind('<Motion>', self.motion)

    def enter(self, event=None):
        """Handle mouse enter event"""
        self.schedule()

    def leave(self, event=None):
        """Handle mouse leave event"""
        self.unschedule()
        self.hide()

    def motion(self, event=None):
        """Handle mouse motion event"""
        self.x = event.x
        self.y = event.y

    def schedule(self):
        """Schedule tooltip display"""
        self.unschedule()
        self.id = self.widget.after(500, self.show)

    def unschedule(self):
        """Unschedule tooltip display"""
        if self.id:
            self.widget.after_cancel(self.id)
            self.id = None

    def show(self):
        """Show tooltip"""
        if self.tooltip:
            return

        # Get screen coordinates
        x = self.widget.winfo_rootx() + self.x + 20
        y = self.widget.winfo_rooty() + self.y + 20

        # Create tooltip window
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)

        # Create tooltip label
        label = ttk.Label(
            self.tooltip,
            text=self.text,
            justify=tk.LEFT,
            background="#ffffe0",
            relief=tk.SOLID,
            borderwidth=1,
            font=("Helvetica", "9", "normal")
        )
        label.pack()

        # Position tooltip
        self.tooltip.wm_geometry(f"+{x}+{y}")

    def hide(self):
        """Hide tooltip"""
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None

    def __del__(self):
        """Clean up on deletion"""
        self.hide()
        if self.widget:
            self.widget.unbind('<Enter>')
            self.widget.unbind('<Leave>')
            self.widget.unbind('<Motion>')

class PoolApp(ttkthemes.ThemedTk):
    def __init__(self, tester: WaterTester, *args, **kwargs):
        try:
            # Initialize parent class first
            super().__init__(*args, **kwargs)
            
            # Initialize basic data structures
            self.customer_info = {
                'name': '',
                'address': '',
                'phone': '',
                'pool_type': 'Select Pool Type'
            }
            
            self.chemistry_data = {
                'readings': []
            }
            
            self.last_update = datetime.now()
            
            # Initialize critical attributes
            self.tooltips = []
            self.task_history = []
            self.reminders = []
            
            # Store tester instance
            if not isinstance(tester, WaterTester):
                raise ValueError("Invalid tester instance provided")
            self.tester = tester

            # Initialize core components
            self._initialize_variables()
            
            # Configure window properties
            self.title("Deep Blue Pool Chemistry Monitor")
            self.geometry("1400x1000")
            self.minsize(1200, 800)
            
            # Configure appearance
            self._configure_colors()
            self.set_theme("arc")
            self.style = ttk.Style()
            self._configure_styles()
            
            # Create UI components
            self.main_container = ttk.Frame(self, style="Main.TFrame")
            self.main_container.pack(fill='both', expand=True, padx=20, pady=20)
            self._create_header()
            self._create_main_content()
            self._create_menu()
            
            # Initialize functionality
            self._setup_arduino_communication()
            self._load_data()
            self._start_periodic_updates()
            
            # Bind events
            self.bind('<Configure>', self._on_window_configure)
            self.protocol("WM_DELETE_WINDOW", self._on_closing)
            
            logger.info("PoolApp initialization completed successfully")
            
        except Exception as e:
            logger.error(f"Error initializing PoolApp: {str(e)}")
            messagebox.showerror("Initialization Error", 
                               "Failed to initialize application. Check logs for details.")
            raise
 
    def _initialize_variables(self):
        """Initialize all application variables"""
        try:
            # API Keys and External Services
            self.weather_api_key = "b5a290b128a44885bb7112326251305"
            
            # Data Storage
            self.weather_data = None
            self.entries = {}
            self.customer_info = {}
            self.parameter_vars = {}
            self.selected_parameters = set()
            self.last_update = None

            # Graph Settings and Parameters
            self.selected_param = tk.StringVar(value="pH")
            self.time_range = tk.StringVar(value="1W")
            self.graph_settings = {
                "show_ideal_range": True,
                "show_grid": True
            }

            # Initialize Matplotlib Components
            self.fig = Figure(figsize=(8, 6))
            self.ax = self.fig.add_subplot(111)
            self.canvas = FigureCanvasTkAgg(self.fig, master=self)
            self.canvas.draw()

            # Settings and Configurations
            self.settings = {
                'email_reminders': True,
                'sms_reminders': False,
                'data_retention': '6M',
                'auto_backup': True,
                'backup_path': './backup',
                'theme': 'default',
                'window_size': DEFAULT_WINDOW_SIZE
            }
            self.alert_settings = {}
            self.email_config = {}
            self.sms_config = {}
            
            # File Paths
            self.data_file = Path("data/pool_data.json")
            self.settings_file = Path("data/settings.json")

            # Graph Colors
            self.graph_colors = {
                "pH": "#2196F3",
                "Chlorine": "#4CAF50",
                "Alkalinity": "#FFC107",
                "Hardness": "#9C27B0",
                "Stabilizer": "#FF5722"
            }

            # Ensure data directory exists
            self.data_file.parent.mkdir(parents=True, exist_ok=True)

            # Initialize chemistry data
            self.chemistry_data = {
                'timestamp': datetime.now().isoformat(),
                'readings': []
            }

            logger.info("Variables initialized successfully")
            
        except Exception as e:
            logger.error(f"Error initializing variables: {str(e)}")
            raise

    def _configure_colors(self):
        """Configure application color scheme"""
        try:
            # Base colors
            self.primary_color = "#2196F3"    # Material Blue
            self.secondary_color = "#BBDEFB"  # Light Blue
            self.background_color = "#F5F5F5" # Light Gray

            # Add these two lines exactly here
            self.success_color = "#4CAF50"    # Material Green
            self.danger_color = "#F44336"     # Material Red

            logger.info("Colors configured successfully")
            
        except Exception as e:
            logger.error(f"Error configuring colors: {str(e)}")
            raise

    def _configure_styles(self):
        """Configure ttk styles"""
        try:
            self.style.configure("Main.TFrame",
                               background=self.background_color)
            
            # Header styles with blue color
            self.style.configure("Header.TLabel",
                               foreground="#2196F3",  # Material Blue
                               background=self.background_color,
                               font=('Helvetica', 24, 'bold'))
            
            # DateTime label style
            self.style.configure("DateTime.TLabel",
                               foreground="black",
                               background=self.background_color,
                               font=('Helvetica', 12))
            
            # Range label style
            self.style.configure("Range.TLabel",
                               foreground="black",
                               background=self.background_color,
                               font=('Helvetica', 10, 'italic'))
            
            # Button styles
            self.style.configure("Primary.TButton",
                               foreground="black",
                               background=self.primary_color,
                               font=('Helvetica', 10),
                               padding=(10, 5))
            
            self.style.configure("Success.TButton",
                    foreground="black",
                    background=self.success_color,  # Used here
                    font=('Helvetica', 10),
                    padding=(10, 5))

            self.style.configure("Danger.TButton",
                    foreground="black",
                    background=self.danger_color,  # Used here
                    font=('Helvetica', 10),
                    padding=(10, 5))
            
            # Custom button styles
            self.style.configure("Black.TButton",
                               foreground="black",
                               font=('Helvetica', 10),
                               padding=(10, 5))
            
            # Entry styles
            self.style.configure("Custom.TEntry",
                               fieldbackground="white",
                               foreground="black",
                               font=('Helvetica', 10),
                               padding=5)
            
            # Frame styles
            self.style.configure("Card.TFrame",
                               background=self.background_color)
            
            self.style.configure("Custom.TLabelframe",
                               background=self.background_color)
            
            self.style.configure("Custom.TLabelframe.Label",
                               background=self.background_color,
                               foreground="#2196F3",  # Material Blue
                               font=('Helvetica', 12, 'bold'))
            
            # Alert styles
            self.style.configure("Alert.TLabel",
                               foreground="red",
                               background=self.background_color,
                               font=('Helvetica', 11, 'bold'))
            
            # Weather styles
            self.style.configure("Weather.TLabel",
                               foreground="black",
                               background=self.background_color,
                               font=('Helvetica', 10))
            
            # Tooltip style
            self.style.configure("Tooltip.TLabel",
                               background="#ffffe0",  # Light yellow
                               foreground="black",
                               font=('Helvetica', 9),
                               relief="solid",
                               borderwidth=1)

            # Graph styles
            plt.style.use('default')  # Reset to default style first
            matplotlib.rcParams.update({
                'font.size': 10,
                'axes.labelsize': 12,
                'axes.titlesize': 14,
                'xtick.labelsize': 10,
                'ytick.labelsize': 10,
                'figure.titlesize': 16,
                'figure.facecolor': self.background_color,
                'axes.facecolor': 'white',
                'grid.color': '#E0E0E0',
                'grid.linestyle': '--',
                'grid.linewidth': 0.5,
                'axes.grid': True,
                'axes.edgecolor': '#666666',
                'axes.linewidth': 1.0,
                'lines.linewidth': 2.0,
                'lines.markersize': 6,
                'xtick.color': 'black',
                'ytick.color': 'black',
                'axes.labelcolor': 'black',
                'axes.titlecolor': "#2196F3"  # Blue titles in graphs
            })
            
            logger.info("Styles configured successfully")
            
        except Exception as e:
                logger.error(f"Error configuring styles: {str(e)}")
                raise       

    def _create_header(self):
        """Create the application header section"""
        try:
            # Create header frame
            self.header_frame = ttk.Frame(self.main_container)
            self.header_frame.pack(fill='x', pady=(0, 20))
            
            # Create title frame
            title_frame = ttk.Frame(self.header_frame)
            title_frame.pack(fill='x')
            
            # Add logo/title
            title_label = ttk.Label(title_frame,
                                  text="Deep Blue Pool Chemistry Monitor",
                                  style="Header.TLabel",
                                  font=('Helvetica', 24, 'bold'))
            title_label.pack(side='left', padx=10)
            
            # Add date/time display
            self.datetime_label = ttk.Label(title_frame,
                                          text="",
                                          style="DateTime.TLabel",
                                          font=('Helvetica', 12))
            self.datetime_label.pack(side='right', padx=10)
            
            # Start datetime updates
            self._update_datetime()
            
            # Add separator
            ttk.Separator(self.header_frame, 
                         orient='horizontal').pack(fill='x', pady=10)

            logger.info("Header created successfully")
            
        except Exception as e:
            logger.error(f"Error creating header: {str(e)}")
            raise

    def _create_main_content(self):
        """Create main content section with notebook"""
        try:
            # Create notebook
            self.notebook = ttk.Notebook(self.main_container)
            self.notebook.pack(fill='both', expand=True)

            # Create input tab
            self.input_tab = ttk.Frame(self.notebook)
            self.notebook.add(self.input_tab, text="Input Parameters")
            self._create_input_tab()

            # Create recommendation tab
            self.recommendation_tab = ttk.Frame(self.notebook)
            self.notebook.add(self.recommendation_tab, text="Recommendations")
            self._create_recommendation_tab()

            # Create test strip tab
            self._create_test_strip_tab()

            # Create trends tab
            self.trends_tab = ttk.Frame(self.notebook)
            self.notebook.add(self.trends_tab, text="Trends & History")
            self._create_trends_tab(self.trends_tab)

            logger.info("Main content created successfully")
            
        except Exception as e:
            logger.error(f"Error creating main content: {str(e)}")
            raise
        
    def _create_menu(self):
            """Create application menu bar"""
            try:
                menubar = tk.Menu(self)
                self.config(menu=menubar)
                
                # File Menu
                file_menu = tk.Menu(menubar, tearoff=0)
                menubar.add_cascade(label="File", menu=file_menu)
                file_menu.add_command(label="Save", command=self._save_data)
                file_menu.add_command(label="Export", command=self._export_data)
                file_menu.add_separator()
                file_menu.add_command(label="Exit", command=self._confirm_exit)
                
                # Tools Menu
                tools_menu = tk.Menu(menubar, tearoff=0)
                menubar.add_cascade(label="Tools", menu=tools_menu)
                tools_menu.add_command(label="Settings", command=self._show_settings)
                tools_menu.add_command(label="Run Tests", command=self._run_tests)
                tools_menu.add_command(label="Calibrate Sensors", command=self._calibrate_ph_sensor)
                
                # Help Menu
                help_menu = tk.Menu(menubar, tearoff=0)
                menubar.add_cascade(label="Help", menu=help_menu)
                help_menu.add_command(label="Documentation", command=self._show_help)
                help_menu.add_command(label="Check Updates", command=self._check_updates)
                help_menu.add_separator()
                help_menu.add_command(label="About", command=self._show_about)
                
                logger.info("Menu bar created successfully")
                
            except Exception as e:
                logger.error(f"Error creating menu: {str(e)}")
                raise

    def _load_saved_data(self):
        """Load saved application data"""
        try:
            if self.data_file.exists():
                with open(self.data_file, 'r') as f:
                    data = json.load(f)

                # Load customer info
                self.customer_info = data.get('customer_info', {})
                
                # Update customer info fields if they exist
                if 'customer_name_entry' in self.__dict__:
                    self.customer_name_entry.delete(0, tk.END)
                    self.customer_name_entry.insert(
                        0, 
                        self.customer_info.get('name', '')
                    )
                
                if hasattr(self, 'address_entry'):
                    self.address_entry.delete(0, tk.END)
                    self.address_entry.insert(
                        0, 
                        self.customer_info.get('address', '')
                    )
                
                if hasattr(self, 'phone_entry'):
                    self.phone_entry.delete(0, tk.END)
                    self.phone_entry.insert(
                        0, 
                        self.customer_info.get('phone', '')
                    )

                # Load chemistry data
                self.chemistry_data = data.get('chemistry_data', {
                    'timestamp': datetime.now().isoformat(),
                    'readings': []
                })

                # Update last update time
                self.last_update = datetime.fromisoformat(
                    data.get('last_update', datetime.now().isoformat())
                )

                logger.info("Data loaded successfully")
                
            else:
                logger.info("No saved data found")
                self._initialize_empty_data()
                
        except Exception as e:
            logger.error(f"Error loading saved data: {str(e)}")
            self._initialize_empty_data()

    def _initialize_empty_data(self):
        """Initialize empty data structures"""
        try:
            self.customer_info = {
                'name': '',
                'address': '',
                'phone': '',
                'pool_type': 'Select Pool Type'
            }
            
            self.chemistry_data = {
                'timestamp': datetime.now().isoformat(),
                'readings': []
            }
            
            self.last_update = datetime.now()
            
            logger.debug("Empty data initialized")
            
        except Exception as e:
            logger.error(f"Error initializing empty data: {str(e)}")
            raise

    def _setup_arduino_communication(self):
        """Initialize Arduino serial communication"""
        try:
            # Initialize serial connection variables
            self.arduino_connected = False
            self.arduino_port = None
            self.serial_conn = None
            
            # Start connection thread
            self._stop_threads = threading.Event()
            self.arduino_thread = threading.Thread(
                target=self._arduino_connection_loop,
                daemon=True
            )
            self.arduino_thread.start()
            
            logger.info("Arduino communication setup initialized")
            
        except Exception as e:
            logger.error(f"Error setting up Arduino communication: {str(e)}")
            self._handle_error("Failed to setup Arduino communication", False)

    def _load_schedule_from_file(self) -> dict:
        """
        Load schedule from file
        
        Returns:
            dict: Task schedule
        """
        try:
            schedule_file = Path("data/schedule/schedule.json")
            if schedule_file.exists():
                with open(schedule_file, 'r') as f:
                    schedule = json.load(f)
                logger.debug("Schedule loaded from file")
                return schedule
            else:
                logger.debug("No schedule file found")
                return {
                    'daily': [],
                    'weekly': [],
                    'monthly': []
                }

        except Exception as e:
            logger.error(f"Error loading schedule from file: {str(e)}")
            return {
                'daily': [],
                'weekly': [],
                'monthly': []
            }

    def _check_tasks(self):
        """Check scheduled tasks and reschedule next check"""
        try:
            # Load schedule
            schedule = self._load_schedule_from_file()
            current_time = datetime.now()

            # Process tasks
            if schedule:
                self._process_daily_tasks(schedule.get('daily', []), current_time)
                self._process_weekly_tasks(schedule.get('weekly', []), current_time)
                self._process_monthly_tasks(schedule.get('monthly', []), current_time)

            # Schedule next check
            self.task_check_id = self.after(60000, self._check_tasks)
            
            logger.debug("Task check completed successfully")

        except Exception as e:
            logger.error(f"Error in task check: {str(e)}")
            # Ensure next check is scheduled even if error occurs
            self.task_check_id = self.after(60000, self._check_tasks)

    def _save_schedule_to_file(self, schedule: dict):
        """
        Save schedule to file
        
        Args:
            schedule (dict): Task schedule to save
        """
        try:
            # Create schedule directory
            schedule_dir = Path("data/schedule")
            schedule_dir.mkdir(parents=True, exist_ok=True)

            # Save to file
            schedule_file = schedule_dir / "schedule.json"
            with open(schedule_file, 'w') as f:
                json.dump(schedule, f, indent=4)

            logger.debug("Schedule saved to file successfully")

        except Exception as e:
            logger.error(f"Error saving schedule to file: {str(e)}")
            raise

    def _start_periodic_updates(self):
        """Start all periodic update tasks"""
        try:
            # Schedule weather updates (every 30 minutes)
            if hasattr(self, 'weather_update_id'):
                self.after_cancel(self.weather_update_id)
            self._update_weather()
            self.weather_update_id = self.after(1800000, self._update_weather)

            # Schedule datetime updates (every second)
            if hasattr(self, 'datetime_update_id'):
                self.after_cancel(self.datetime_update_id)
            self._update_datetime()
            self.datetime_update_id = self.after(1000, self._update_datetime)

            # Schedule task checks (every minute)
            if hasattr(self, 'task_check_id'):
                self.after_cancel(self.task_check_id)
            self._check_tasks()  # Initial check
            self.task_check_id = self.after(60000, self._check_tasks)

            # Schedule alert checks (every 5 minutes)
            if hasattr(self, 'alert_check_id'):
                self.after_cancel(self.alert_check_id)
            self._check_alerts()
            self.alert_check_id = self.after(300000, self._check_alerts)

            # Schedule data backup (every hour)
            if hasattr(self, 'backup_id'):
                self.after_cancel(self.backup_id)
            self._create_backup('auto')
            self.backup_id = self.after(3600000, lambda: self._create_backup('auto'))

            logger.debug("Periodic updates started successfully")

        except Exception as e:
            logger.error(f"Error starting periodic updates: {str(e)}")
            self._handle_error("Failed to start periodic updates")

    def _check_tasks(self):
        """Check scheduled tasks and reschedule next check"""
        try:
            # Load and check schedule
            schedule = self._load_schedule_from_file()
            current_time = datetime.now()

            # Process tasks
            self._process_daily_tasks(schedule.get('daily', []), current_time)
            self._process_weekly_tasks(schedule.get('weekly', []), current_time)
            self._process_monthly_tasks(schedule.get('monthly', []), current_time)

            # Schedule next check
            self.task_check_id = self.after(60000, self._check_tasks)
            
            logger.debug("Task check completed successfully")

        except Exception as e:
            logger.error(f"Error in task check: {str(e)}")
            # Ensure next check is scheduled even if error occurs
            self.task_check_id = self.after(60000, self._check_tasks)

    def _process_daily_tasks(self, tasks: List[dict], current_time: datetime):
        """Process daily tasks"""
        try:
            for task in tasks:
                if self._is_task_due(task, current_time, 'daily'):
                    self._execute_scheduled_task(task)
        except Exception as e:
            logger.error(f"Error processing daily tasks: {str(e)}")

    def _process_weekly_tasks(self, tasks: List[dict], current_time: datetime):
        """Process weekly tasks"""
        try:
            for task in tasks:
                if self._is_task_due(task, current_time, 'weekly'):
                    self._execute_scheduled_task(task)
        except Exception as e:
            logger.error(f"Error processing weekly tasks: {str(e)}")

    def _process_monthly_tasks(self, tasks: List[dict], current_time: datetime):
        """Process monthly tasks"""
        try:
            for task in tasks:
                if self._is_task_due(task, current_time, 'monthly'):
                    self._execute_scheduled_task(task)
        except Exception as e:
            logger.error(f"Error processing monthly tasks: {str(e)}")

    def _on_window_configure(self, event=None):
        """Handle window resize events"""
        try:
            if event and event.widget == self:
                # Minimum size enforcement
                if event.width < 1200:
                    self.geometry(f"1200x{event.height}")
                if event.height < 800:
                    self.geometry(f"{event.width}x800")

                # Update graph size if exists
                if hasattr(self, 'canvas'):
                    self._resize_graph()

                # Save new size to settings
                self.settings['window_size'] = {
                    'width': event.width,
                    'height': event.height
                }
                
        except Exception as e:
            logger.error(f"Error handling window configuration: {str(e)}")

    def _on_closing(self):
        """Handle application closing"""
        try:
            # Ask for confirmation
            if messagebox.askokcancel("Quit", "Do you want to quit?"):
                # Save current data
                self._save_data()
                
                # Save settings
                self._save_settings_to_file()
                
                # Clean up resources
                self._cleanup()
                
                # Destroy window
                self.quit()
                self.destroy()
                
                logger.info("Application closed successfully")
                
        except Exception as e:
            logger.error(f"Error during application closing: {str(e)}")
            self.quit()
            self.destroy()

    def _create_input_tab(self):
        try:
            # Create main input frame
            main_frame = ttk.Frame(self.input_tab, style="Card.TFrame")
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)

            # Configure grid weights for main_frame
            main_frame.grid_columnconfigure(0, weight=1)
            main_frame.grid_rowconfigure(2, weight=1)  # Give weight to the row above buttons

            # Create top section frame for customer info and weather
            top_section = ttk.Frame(main_frame)
            top_section.pack(fill='x', pady=(0, 20))

            # Create customer info and weather sections...

            # Create chemical parameters section
            chemicals_frame = ttk.LabelFrame(main_frame, 
                                           text="Chemical Parameters", 
                                           padding=5,
                                           style="Custom.TLabelframe")
            chemicals_frame.pack(fill='both', expand=True, padx=10, pady=(0, 10))

            # Create button frame with fixed position
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill='x', side='bottom', pady=10)

            # Generate Recommendations button
            generate_btn = ttk.Button(button_frame,
                                    text="Generate Recommendations",
                                    command=self._update_recommendations,
                                    style="Primary.TButton")
            generate_btn.pack(side='left', padx=5)

            # Clear Values button
            clear_btn = ttk.Button(button_frame,
                                 text="Clear Values",
                                 command=self._clear_values,
                                 style="Black.TButton")
            clear_btn.pack(side='left', padx=5)

            # Save Values button
            save_btn = ttk.Button(button_frame,
                                 text="Save Values",
                                 command=self._save_values,
                                 style="Primary.TButton")
            save_btn.pack(side='left', padx=5)

        except Exception as e:
            logger.error(f"Error creating input tab: {str(e)}")
            raise

    def _create_customer_info(self, parent_frame):
        """Create customer information section"""
        try:
            # Create customer frame
            customer_frame = ttk.LabelFrame(parent_frame, 
                                          text="Customer Information", 
                                          padding=10)
            customer_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 10))

            # Configure grid weights
            customer_frame.grid_columnconfigure(1, weight=1)

            # Customer Name
            ttk.Label(customer_frame, 
                     text="Customer Name:",
                     font=("Helvetica", 10, "bold")).grid(row=0, column=0, 
                                                         padx=5, pady=5, 
                                                         sticky='e')
            self.customer_name_entry = ttk.Entry(customer_frame, 
                                               style="Custom.TEntry", 
                                               width=30)
            self.customer_name_entry.grid(row=0, column=1, padx=5, pady=5, 
                                        sticky='w')

            # Address - Add this section
            ttk.Label(customer_frame,
                     text="Address:",
                     font=("Helvetica", 10, "bold")).grid(row=1, column=0,
                                                         padx=5, pady=5,
                                                         sticky='e')
            self.address_entry = ttk.Entry(customer_frame,
                                         style="Custom.TEntry",
                                         width=30)
            self.address_entry.grid(row=1, column=1, padx=5, pady=5,
                                  sticky='w')

            # Phone Number - Add this section
            ttk.Label(customer_frame,
                     text="Phone:",
                     font=("Helvetica", 10, "bold")).grid(row=2, column=0,
                                                         padx=5, pady=5,
                                                         sticky='e')
            self.phone_entry = ttk.Entry(customer_frame,
                                       style="Custom.TEntry",
                                       width=30)
            self.phone_entry.grid(row=2, column=1, padx=5, pady=5,
                                sticky='w')

            # Pool Type Dropdown
            ttk.Label(customer_frame, 
                     text="Pool Type:",
                     font=("Helvetica", 10, "bold")).grid(row=3, column=0, 
                                                         padx=5, pady=5, 
                                                         sticky='e')
            
            self.pool_type_var = tk.StringVar(value="Select Pool Type")
            self.pool_type_dropdown = ttk.Combobox(
                customer_frame,
                textvariable=self.pool_type_var,
                values=self.tester.pool_types,
                state='readonly',
                width=27
            )
            self.pool_type_dropdown.grid(row=3, column=1, padx=5, pady=5, 
                                       sticky='w')
            
            # Bind pool type change event
            self.pool_type_dropdown.bind('<<ComboboxSelected>>', 
                                   self._on_pool_type_change)

            # Save Button
            save_btn = ttk.Button(
                customer_frame,
                text="Save Customer Info",
                command=self._save_customer_info,
                style="Primary.TButton"
            )
            save_btn.grid(row=4, column=0, columnspan=2, pady=10)

            logger.info("Customer information section created successfully")

        except Exception as e:
            logger.error(f"Error creating customer information section: {str(e)}")
            raise

    def _create_weather_section(self, parent: ttk.Frame):
        """
        Create weather information section
        
        Args:
            parent (ttk.Frame): Parent frame
        """
        try:
            # Create weather frame
            self.weather_frame = ttk.LabelFrame(
                parent,
                text="Weather Conditions",
                padding=10
            )
            self.weather_frame.grid(row=0, column=1, sticky='nsew', padx=(10, 0))

            # Current conditions frame
            current_frame = ttk.Frame(self.weather_frame)
            current_frame.pack(fill='x', pady=(0, 10))

            # Initialize current condition labels
            self.current_labels = {}
            
            # Temperature
            temp_frame = ttk.Frame(current_frame)
            temp_frame.pack(fill='x', pady=2)
            ttk.Label(
                temp_frame,
                text="Temperature:",
                font=("Helvetica", 9, "bold")
            ).pack(side='left')
            self.current_labels['temp'] = ttk.Label(
                temp_frame,
                text="--°F",
                font=("Helvetica", 9)
            )
            self.current_labels['temp'].pack(side='left', padx=5)

            # Humidity
            humidity_frame = ttk.Frame(current_frame)
            humidity_frame.pack(fill='x', pady=2)
            ttk.Label(
                humidity_frame,
                text="Humidity:",
                font=("Helvetica", 9, "bold")
            ).pack(side='left')
            self.current_labels['humidity'] = ttk.Label(
                humidity_frame,
                text="--%",
                font=("Helvetica", 9)
            )
            self.current_labels['humidity'].pack(side='left', padx=5)

            # UV Index
            uv_frame = ttk.Frame(current_frame)
            uv_frame.pack(fill='x', pady=2)
            ttk.Label(
                uv_frame,
                text="UV Index:",
                font=("Helvetica", 9, "bold")
            ).pack(side='left')
            self.current_labels['uv_index'] = ttk.Label(
                uv_frame,
                text="--",
                font=("Helvetica", 9)
            )
            self.current_labels['uv_index'].pack(side='left', padx=5)

            # Rain Chance
            rain_frame = ttk.Frame(current_frame)
            rain_frame.pack(fill='x', pady=2)
            ttk.Label(
                rain_frame,
                text="Rain Chance:",
                font=("Helvetica", 9, "bold")
            ).pack(side='left')
            self.current_labels['rain_chance'] = ttk.Label(
                rain_frame,
                text="--%",
                font=("Helvetica", 9)
            )
            self.current_labels['rain_chance'].pack(side='left', padx=5)

            # Add separator
            ttk.Separator(self.weather_frame).pack(fill='x', pady=10)

            # Impact section
            impact_label = ttk.Label(
                self.weather_frame,
                text="Pool Impact Analysis",
                font=("Helvetica", 9, "bold")
            )
            impact_label.pack(anchor='w')

            # Impact details
            self.impact_details = {
                'temperature': ttk.Label(self.weather_frame, text=""),
                'uv': ttk.Label(self.weather_frame, text=""),
                'rain': ttk.Label(self.weather_frame, text=""),
                'wind': ttk.Label(self.weather_frame, text="")
            }

            for label in self.impact_details.values():
                label.pack(fill='x', pady=2)

            # Add update button
            update_frame = ttk.Frame(self.weather_frame)
            update_frame.pack(fill='x', pady=(10, 0))

            # ZIP code entry
            self.zip_entry = ttk.Entry(
                update_frame,
                width=10
            )
            self.zip_entry.pack(side='left', padx=(0, 5))
            
            # Add tooltip
            self._create_tooltip(
                self.zip_entry,
                "Enter ZIP code for weather data"
            )

            # Update button
            update_btn = ttk.Button(
                update_frame,
                text="Update Weather",
                command=self._update_weather,
                style="Small.TButton"
            )
            update_btn.pack(side='left')

            logger.debug("Weather section created successfully")

        except Exception as e:
            logger.error(f"Error creating weather section: {str(e)}")
            raise

    def _create_chemistry_inputs(self, parent: ttk.Frame):
        """
        Create chemical parameter input fields
        
        Args:
            parent (ttk.Frame): Parent frame
        """
        try:
            # Create grid for chemical inputs
            for i, (param, (min_val, max_val)) in enumerate(self.tester.ideal_levels.items()):
                # Create label
                ttk.Label(
                    parent,
                    text=param,
                    font=("Helvetica", 10)
                ).grid(row=i, column=0, padx=5, pady=5, sticky='e')

                # Create entry
                entry = ttk.Entry(parent, width=10)
                entry.grid(row=i, column=1, padx=5, pady=5)
                
                # Store entry widget reference
                self.entries[param] = entry

                # Add validation
                vcmd = (self.register(self._validate_numeric_input), '%P')
                entry.configure(validate='key', validatecommand=vcmd)

                # Create ideal range label
                range_text = f"Ideal: {min_val} - {max_val}"
                range_label = ttk.Label(
                    parent,
                    text=range_text,
                    font=("Helvetica", 9),
                    foreground="gray"
                )
                range_label.grid(row=i, column=2, padx=5, pady=5, sticky='w')

                # Add tooltips
                tooltip_text = f"Enter current {param}\nIdeal range: {min_val} - {max_val}"
                self._create_tooltip(entry, tooltip_text)
                self._create_tooltip(range_label, f"Recommended range for {param}")

            # Configure grid weights
            parent.grid_columnconfigure(2, weight=1)

            logger.debug("Chemistry inputs created successfully")

        except Exception as e:
            logger.error(f"Error creating chemistry inputs: {str(e)}")
            raise

    def _create_recommendation_tab(self):
        """Create recommendations tab"""
        try:
            # Create main frame
            main_frame = ttk.Frame(self.recommendation_tab, padding=20)
            main_frame.pack(fill='both', expand=True)

            # Add header
            header_frame = ttk.Frame(main_frame)
            header_frame.pack(fill='x', pady=(0, 20))

            ttk.Label(
                header_frame,
                text="Chemical Recommendations",
                font=("Helvetica", 16, "bold"),
                foreground=self.primary_color
            ).pack(side='left')

            # Create scrolled text widget for recommendations
            self.recommendation_text = scrolledtext.ScrolledText(
                main_frame,
                wrap=tk.WORD,
                width=70,
                height=30,
                font=("Helvetica", 10)
            )
            self.recommendation_text.pack(fill='both', expand=True)

            # Configure tags for formatting
            self.recommendation_text.tag_configure(
                "header",
                font=("Helvetica", 12, "bold"),
                foreground=self.primary_color
            )
            
            self.recommendation_text.tag_configure(
                "section",
                font=("Helvetica", 11, "bold"),
                foreground="#333333"
            )
            
            self.recommendation_text.tag_configure(
                "normal",
                font=("Helvetica", 10),
                foreground="#333333"
            )
            
            self.recommendation_text.tag_configure(
                "warning",
                font=("Helvetica", 10, "bold"),
                foreground="#F44336"
            )
            
            self.recommendation_text.tag_configure(
                "success",
                font=("Helvetica", 10),
                foreground="#4CAF50"
            )
            
            self.recommendation_text.tag_configure(
                "alert",
                font=("Helvetica", 10),
                foreground="#FF9800"
            )

            # Make text widget read-only
            self.recommendation_text.configure(state='disabled')

            # Add initial message
            self._show_initial_recommendation_message()

            logger.info("Recommendation tab created successfully")

        except Exception as e:
            logger.error(f"Error creating recommendation tab: {str(e)}")
            raise

    def _save_values(self):
        """Save current chemical readings"""
        try:
            # Collect current values
            readings = {}
            for param, entry in self.entries.items():
                try:
                    value = float(entry.get()) if entry.get() else None
                    if value is not None:
                        readings[param] = value
                except ValueError:
                    continue

            if not readings:
                messagebox.showwarning(
                    "No Data",
                    "No values to save. Please enter chemical readings."
                )
                return

            # Create reading entry
            reading = {
                'timestamp': datetime.now().isoformat(),
                'values': readings,
                'weather': self.weather_data if hasattr(self, 'weather_data') else None
            }

            # Add to chemistry data
            if not hasattr(self, 'chemistry_data'):
                self.chemistry_data = {'readings': []}
            
            self.chemistry_data['readings'].append(reading)

            # Save to file
            self._save_data()

            # Show confirmation
            messagebox.showinfo(
                "Success",
                "Chemical readings saved successfully!"
            )

            # Update graphs if they exist
            if hasattr(self, 'canvas'):
                self._update_combined_graph()

            logger.debug("Values saved successfully")

        except Exception as e:
            logger.error(f"Error saving values: {str(e)}")
            self._handle_error("Failed to save values")

    def _clear_values(self):
            """Clear all chemical input values"""
            try:
                # Clear all entry fields
                for entry in self.entries.values():
                    entry.delete(0, tk.END)

                # Reset any status indicators or displays
                if 'recommendation_text' in self.__dict__:
                    self.recommendation_text.delete(1.0, tk.END)

                # Show confirmation
                messagebox.showinfo(
                    "Success",
                    "All values cleared successfully!"
                )

                logger.debug("Values cleared successfully")

            except Exception as e:
                logger.error(f"Error clearing values: {str(e)}")
                self._handle_error("Failed to clear values")

    def _show_settings_dialog(self):
        """Show settings configuration dialog"""
        try:
            # Create settings dialog
            dialog = tk.Toplevel(self)
            dialog.title("Settings")
            dialog.geometry("500x600")
            dialog.transient(self)
            dialog.grab_set()

            # Center dialog on screen
            dialog.geometry(f"+{self.winfo_x() + 50}+{self.winfo_y() + 50}")

            # Create notebook for settings tabs
            notebook = ttk.Notebook(dialog)
            notebook.pack(fill='both', expand=True, padx=10, pady=10)

            # General Settings Tab
            general_frame = ttk.Frame(notebook, padding=10)
            notebook.add(general_frame, text="General")
            self._create_general_settings(general_frame)

            # Email Settings Tab
            email_frame = ttk.Frame(notebook, padding=10)
            notebook.add(email_frame, text="Email")
            self._create_email_settings(email_frame)

            # SMS Settings Tab
            sms_frame = ttk.Frame(notebook, padding=10)
            notebook.add(sms_frame, text="SMS")
            self._create_sms_settings(sms_frame)

            # Alert Settings Tab
            alert_frame = ttk.Frame(notebook, padding=10)
            notebook.add(alert_frame, text="Alerts")
            self._create_alert_settings(alert_frame)

            # Add Save/Cancel buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill='x', padx=10, pady=10)

            ttk.Button(
                button_frame,
                text="Save",
                command=lambda: self._save_settings(dialog),
                style="Primary.TButton"
            ).pack(side='right', padx=5)

            ttk.Button(
                button_frame,
                text="Cancel",
                command=dialog.destroy
            ).pack(side='right', padx=5)

            logger.debug("Settings dialog created successfully")

        except Exception as e:
            logger.error(f"Error showing settings dialog: {str(e)}")
            self._handle_error("Failed to show settings")

    def _create_general_settings(self, parent):
        """Create general settings section"""
        try:
            # Theme selection
            theme_frame = ttk.LabelFrame(parent, text="Theme", padding=10)
            theme_frame.pack(fill='x', pady=(0, 10))

            self.theme_var = tk.StringVar(value=self.style.theme_use())
            themes = self.style.theme_names()
            
            ttk.Label(theme_frame, text="Application Theme:").pack(anchor='w')
            theme_combo = ttk.Combobox(
                theme_frame,
                textvariable=self.theme_var,
                values=themes,
                state='readonly'
            )
            theme_combo.pack(fill='x', pady=5)

            # Graph settings
            graph_frame = ttk.LabelFrame(parent, text="Graph Settings", padding=10)
            graph_frame.pack(fill='x', pady=(0, 10))

            self.show_grid_var = tk.BooleanVar(
                value=self.graph_settings.get('show_grid', True)
            )
            ttk.Checkbutton(
                graph_frame,
                text="Show Grid Lines",
                variable=self.show_grid_var
            ).pack(anchor='w')

            self.show_points_var = tk.BooleanVar(
                value=self.graph_settings.get('show_points', True)
            )
            ttk.Checkbutton(
                graph_frame,
                text="Show Data Points",
                variable=self.show_points_var
            ).pack(anchor='w')

            # Data retention settings
            retention_frame = ttk.LabelFrame(
                parent,
                text="Data Retention",
                padding=10
            )
            retention_frame.pack(fill='x', pady=(0, 10))

            ttk.Label(
                retention_frame,
                text="Keep history for:"
            ).pack(anchor='w')

            self.retention_var = tk.StringVar(
                value=self.settings.get('data_retention', '6M')
            )
            retention_combo = ttk.Combobox(
                retention_frame,
                textvariable=self.retention_var,
                values=['1M', '3M', '6M', '1Y', 'All'],
                state='readonly'
            )
            retention_combo.pack(fill='x', pady=5)

            # Auto-backup settings
            backup_frame = ttk.LabelFrame(parent, text="Auto-Backup", padding=10)
            backup_frame.pack(fill='x')

            self.backup_enabled_var = tk.BooleanVar(
                value=self.settings.get('auto_backup', True)
            )
            ttk.Checkbutton(
                backup_frame,
                text="Enable Auto-Backup",
                variable=self.backup_enabled_var
            ).pack(anchor='w')

            ttk.Label(
                backup_frame,
                text="Backup Location:"
            ).pack(anchor='w', pady=(5, 0))

            self.backup_path_var = tk.StringVar(
                value=self.settings.get('backup_path', './backup')
            )
            path_frame = ttk.Frame(backup_frame)
            path_frame.pack(fill='x', pady=5)

            ttk.Entry(
                path_frame,
                textvariable=self.backup_path_var,
                state='readonly'
            ).pack(side='left', fill='x', expand=True)

            ttk.Button(
                path_frame,
                text="Browse",
                command=self._browse_backup_path
            ).pack(side='right', padx=(5, 0))

            logger.debug("General settings created successfully")

        except Exception as e:
            logger.error(f"Error creating general settings: {str(e)}")
            raise
  
    def _show_settings(self):
            """Show settings configuration dialog"""
            try:
                # Create settings dialog
                dialog = tk.Toplevel(self)
                dialog.title("Settings")
                dialog.geometry("500x600")
                dialog.transient(self)
                dialog.grab_set()

                # Center dialog on screen
                dialog.geometry(f"+{self.winfo_x() + 50}+{self.winfo_y() + 50}")

                # Create notebook for settings tabs
                notebook = ttk.Notebook(dialog)
                notebook.pack(fill='both', expand=True, padx=10, pady=10)

                # General Settings Tab
                general_frame = ttk.Frame(notebook, padding=10)
                notebook.add(general_frame, text="General")
                self._create_general_settings(general_frame)

                # Email Settings Tab
                email_frame = ttk.Frame(notebook, padding=10)
                notebook.add(email_frame, text="Email")
                self._create_email_settings(email_frame)

                # SMS Settings Tab
                sms_frame = ttk.Frame(notebook, padding=10)
                notebook.add(sms_frame, text="SMS")
                self._create_sms_settings(sms_frame)

                # Alert Settings Tab
                alert_frame = ttk.Frame(notebook, padding=10)
                notebook.add(alert_frame, text="Alerts")
                self._create_alert_settings(alert_frame)

                # Add Save/Cancel buttons
                button_frame = ttk.Frame(dialog)
                button_frame.pack(fill='x', padx=10, pady=10)

                ttk.Button(
                    button_frame,
                    text="Save",
                    command=lambda: self._save_settings(dialog),
                    style="Primary.TButton"
                ).pack(side='right', padx=5)

                ttk.Button(
                    button_frame,
                    text="Cancel",
                    command=dialog.destroy
                ).pack(side='right', padx=5)

                logger.debug("Settings dialog created successfully")

            except Exception as e:
                logger.error(f"Error showing settings dialog: {str(e)}")
                self._handle_error("Failed to show settings")

    def _create_general_settings(self, parent):
        """Create general settings section"""
        try:
            # Theme selection
            theme_frame = ttk.LabelFrame(parent, text="Theme", padding=10)
            theme_frame.pack(fill='x', pady=(0, 10))

            self.theme_var = tk.StringVar(value=self.style.theme_use())
            themes = self.style.theme_names()
            
            ttk.Label(theme_frame, text="Application Theme:").pack(anchor='w')
            theme_combo = ttk.Combobox(
                theme_frame,
                textvariable=self.theme_var,
                values=themes,
                state='readonly'
            )
            theme_combo.pack(fill='x', pady=5)

            # Graph settings
            graph_frame = ttk.LabelFrame(parent, text="Graph Settings", padding=10)
            graph_frame.pack(fill='x', pady=(0, 10))

            self.show_grid_var = tk.BooleanVar(
                value=self.graph_settings.get('show_grid', True)
            )
            ttk.Checkbutton(
                graph_frame,
                text="Show Grid Lines",
                variable=self.show_grid_var
            ).pack(anchor='w')

            self.show_points_var = tk.BooleanVar(
                value=self.graph_settings.get('show_points', True)
            )
            ttk.Checkbutton(
                graph_frame,
                text="Show Data Points",
                variable=self.show_points_var
            ).pack(anchor='w')

            # Data retention settings
            retention_frame = ttk.LabelFrame(
                parent,
                text="Data Retention",
                padding=10
            )
            retention_frame.pack(fill='x', pady=(0, 10))

            ttk.Label(
                retention_frame,
                text="Keep history for:"
            ).pack(anchor='w')

            self.retention_var = tk.StringVar(
                value=self.settings.get('data_retention', '6M')
            )
            retention_combo = ttk.Combobox(
                retention_frame,
                textvariable=self.retention_var,
                values=['1M', '3M', '6M', '1Y', 'All'],
                state='readonly'
            )
            retention_combo.pack(fill='x', pady=5)

            # Auto-backup settings
            backup_frame = ttk.LabelFrame(parent, text="Auto-Backup", padding=10)
            backup_frame.pack(fill='x')

            self.backup_enabled_var = tk.BooleanVar(
                value=self.settings.get('auto_backup', True)
            )
            ttk.Checkbutton(
                backup_frame,
                text="Enable Auto-Backup",
                variable=self.backup_enabled_var
            ).pack(anchor='w')

            ttk.Label(
                backup_frame,
                text="Backup Location:"
            ).pack(anchor='w', pady=(5, 0))

            self.backup_path_var = tk.StringVar(
                value=self.settings.get('backup_path', './backup')
            )
            path_frame = ttk.Frame(backup_frame)
            path_frame.pack(fill='x', pady=5)

            ttk.Entry(
                path_frame,
                textvariable=self.backup_path_var,
                state='readonly'
            ).pack(side='left', fill='x', expand=True)

            ttk.Button(
                path_frame,
                text="Browse",
                command=self._browse_backup_path
            ).pack(side='right', padx=(5, 0))

            logger.debug("General settings created successfully")

        except Exception as e:
            logger.error(f"Error creating general settings: {str(e)}")
            raise

    def _show_initial_recommendation_message(self):
        """Show initial message in recommendation tab"""
        try:
            self.recommendation_text.configure(state='normal')
            self.recommendation_text.delete(1.0, tk.END)
            
            self.recommendation_text.insert(tk.END, 
                "Chemical Recommendations\n", "header")
            self.recommendation_text.insert(tk.END, "\n")
            
            self.recommendation_text.insert(tk.END,
                "Enter chemical readings in the Input Parameters tab and click "
                "'Generate Recommendations' to receive detailed chemical adjustment "
                "recommendations.\n\n", "normal")
            
            self.recommendation_text.insert(tk.END,
                "Recommendations will include:\n", "section")
            
            recommendations = [
                "• Required chemical adjustments",
                "• Safety precautions",
                "• Weather impact analysis",
                "• Maintenance schedule",
                "• Pool type specific guidance"
            ]
            
            for rec in recommendations:
                self.recommendation_text.insert(tk.END, f"{rec}\n", "normal")
                
            self.recommendation_text.configure(state='disabled')

        except Exception as e:
            logger.error(f"Error showing initial recommendation message: {str(e)}")
            raise
   
    def _update_datetime(self):
        """Update the datetime display"""
        try:
            # Update datetime label with current time
            current_time = datetime.now().strftime("%B %d, %Y %I:%M:%S %p")
            self.datetime_label.config(text=current_time)
            
            # Schedule next update in 1 second
            self.after(1000, self._update_datetime)
            
            logger.debug("DateTime updated successfully")
            
        except Exception as e:
            logger.error(f"Error updating datetime: {str(e)}")

    def _validate_numeric_input(self, P):
        """Validate entry input to allow only numbers and decimal point"""
        try:
            # Allow empty string (for deletion)
            if P == "":
                return True
                
            # Allow single decimal point
            if P == ".":
                return True
                
            # Allow negative sign at start
            if P == "-":
                return True
                
            # Check if string contains only one decimal point
            if P.count('.') > 1:
                return False
                
            # Check if negative sign is only at start
            if '-' in P and (P.count('-') > 1 or P[0] != '-'):
                return False
                
            # Check if string is numeric
            try:
                if P[0] == '-':
                    float(P[1:])  # Check rest of string after minus
                else:
                    float(P)
                return True
            except ValueError:
                return False
                
        except Exception as e:
            logger.error(f"Error validating numeric input: {str(e)}")
            return False

    def _validate_input_data(self, data: dict) -> Tuple[bool, List[str]]:
        """
        Validate input data
        
        Args:
            data (dict): Data to validate
            
        Returns:
            Tuple[bool, List[str]]: (is_valid, error_messages)
        """
        try:
            errors = []
            
            # Validate chemical readings
            for param, value in data.items():
                if param not in self.tester.ideal_levels:
                    errors.append(f"Invalid parameter: {param}")
                    continue
                    
                try:
                    value = float(value)
                    if value < 0:
                        errors.append(f"Negative value for {param}: {value}")
                        
                    ideal_min, ideal_max = self.tester.ideal_levels[param]
                    display_min, display_max = self.display_ranges[param]
                    
                    if not display_min <= value <= display_max:
                        errors.append(f"Value out of range for {param}: {value}")
                        
                except ValueError:
                    errors.append(f"Invalid numeric value for {param}: {value}")
                    
            return len(errors) == 0, errors
            
        except Exception as e:
            logger.error(f"Error validating input data: {str(e)}")
            return False, ["Validation error occurred"]

    def _create_tooltip(self, widget: tk.Widget, text: str) -> ToolTip:
        """
        Create and attach a tooltip to a widget.
        
        Args:
            widget (tk.Widget): The widget to attach the tooltip to
            text (str): The text to display in the tooltip
            
        Returns:
            ToolTip: The created tooltip object
            
        Note:
            Tooltips are stored in self.tooltips to prevent garbage collection
        """
        if not hasattr(self, 'tooltips'):
            self.tooltips = []
        tooltip = ToolTip(widget, text)
        self.tooltips.append(tooltip)  
        return tooltip

    def _configure_text_tags(self):
        """Configure text widget tags for formatting"""
        try:
            # Header tag for titles
            self.text.tag_configure("header", 
                font=("Helvetica", 12, "bold"),
                foreground="#2196F3",
                spacing1=10,
                spacing3=10
            )

            # Customer info tag
            self.text.tag_configure("customer_info",
                font=("Helvetica", 10),
                lmargin1=20,
                spacing1=5
            )

            # Warning tag for alerts
            self.text.tag_configure("warning",
                font=("Helvetica", 10, "bold"),
                foreground="red",
                spacing1=5,
                spacing3=5
            )

            # Success tag for good values
            self.text.tag_configure("success",
                font=("Helvetica", 10),
                foreground="green",
                spacing1=5
            )

            # Alert tag for concerning values
            self.text.tag_configure("alert",
                font=("Helvetica", 10),
                foreground="orange",
                spacing1=5
            )

            # Section tag for separating content
            self.text.tag_configure("section",
                font=("Helvetica", 11, "bold"),
                spacing1=10,
                spacing3=5
            )

            # Normal text tag
            self.text.tag_configure("normal",
                font=("Helvetica", 10),
                spacing1=2
            )

            logger.info("Text tags configured successfully")

        except Exception as e:
            logger.error(f"Error configuring text tags: {str(e)}")
            raise

    def _on_pool_type_change(self, event=None):
        """Handle pool type selection change"""
        try:
            pool_type = self.pool_type_var.get()
            if pool_type != "Select Pool Type":
                # Update customer info
                if hasattr(self, 'customer_info'):
                    self.customer_info['pool_type'] = pool_type
                
                # Update recommendations if they exist
                if hasattr(self, 'text'):
                    self._update_recommendations()
                    
                # Save the changes
                self._save_data()
                
                logger.debug(f"Pool type changed to: {pool_type}")
                
        except Exception as e:
            logger.error(f"Error handling pool type change: {str(e)}")
            self._handle_error("Failed to update pool type")

    def _update_weather(self):
        """Update weather data from API"""
        try:
            # Validate widget existence
            if not hasattr(self, 'zip_entry'):
                logger.error("ZIP entry widget not initialized")
                return

            # Get and validate ZIP code
            zip_code = self.zip_entry.get().strip()
            if not zip_code:
                logger.debug("No ZIP code provided")
                return

            # Validate ZIP code format
            if not zip_code.isdigit() or len(zip_code) != 5:
                logger.warning(f"Invalid ZIP code format: {zip_code}")
                messagebox.showwarning(
                    "Invalid ZIP Code",
                    "Please enter a valid 5-digit ZIP code."
                )
                return

            # Fetch weather data
            self.weather_data = self._fetch_weather_data(zip_code)
            
            # Debug print weather data
            logger.debug(f"Weather Data: {self.weather_data}")
            
            # Update display if weather data was fetched successfully
            if self.weather_data:
                # Update current conditions
                current = self.weather_data['current']
                for key, value in current.items():
                    if key in self.current_labels:
                        if key == 'temp':
                            self.current_labels[key].config(
                                text=f"{value}°F"
                            )
                        elif key in ['humidity', 'rain_chance']:
                            self.current_labels[key].config(
                                text=f"{value}%"
                            )
                        else:
                            self.current_labels[key].config(
                                text=str(value)
                            )
                            
                # Update forecast display
                self._update_forecast_display()
                    
                # Update impact analysis
                self._update_impact_details(current)

                # Update last update time
                self.last_update = datetime.now()
                
                logger.info("Weather data updated successfully")
                
            else:
                logger.warning("No weather data received")
                self._show_weather_error()

        except requests.RequestException as e:
            logger.error(f"Weather API request failed: {str(e)}")
            self._handle_error(
                "Failed to connect to weather service. Please check your internet connection."
            )
            self._show_weather_error()
            
        except Exception as e:
            logger.error(f"Error updating weather: {str(e)}")
            self._handle_error("Failed to update weather data")
            self._show_weather_error()
            
        finally:
            # Schedule next update regardless of success/failure
            try:
                if hasattr(self, 'weather_update_id'):
                    self.after_cancel(self.weather_update_id)
                self.weather_update_id = self.after(1800000, self._update_weather)  # 30 minutes
            except Exception as e:
                logger.error(f"Error scheduling next weather update: {str(e)}")

    def _show_weather_error(self):
        """Display weather error message"""
        try:
            if hasattr(self, 'weather_frame'):
                # Clear existing content
                for widget in self.weather_frame.winfo_children():
                    widget.destroy()
                    
                # Add error message
                ttk.Label(
                    self.weather_frame,
                    text="Weather data unavailable",
                    foreground="red",
                    font=("Helvetica", 10, "bold")
                ).pack(pady=10)
                
                # Add retry button
                ttk.Button(
                    self.weather_frame,
                    text="Retry",
                    command=self._update_weather,
                    style="Small.TButton"
                ).pack(pady=5)
                
        except Exception as e:
            logger.error(f"Error showing weather error: {str(e)}")

    def _fetch_weather_data(self, zip_code: str) -> dict:
        """
        Fetch weather data from API
        
        Args:
            zip_code (str): ZIP code for weather lookup
            
        Returns:
            dict: Weather data or None if request fails
        """
        try:
            # API endpoint (using weatherapi.com)
            base_url = "http://api.weatherapi.com/v1"
            api_key = self.weather_api_key
            
            # Get forecast data for 5 days
            forecast_url = (
                f"{base_url}/forecast.json"
                f"?key={api_key}"
                f"&q={zip_code}"
                f"&days=5"
                f"&aqi=no"
            )
            
            # Make request with timeout
            response = requests.get(forecast_url, timeout=10)
            response.raise_for_status()
            
            data = response.json()
            
            # Format weather data - Changed temp_f to temp
            weather_data = {
                'current': {
                    'temp': data['current']['temp_f'],  # Changed from temp_f to temp
                    'humidity': data['current']['humidity'],
                    'uv_index': data['current']['uv'],
                    'wind_mph': data['current']['wind_mph'],
                    'rain_chance': data['current']['precip_in'] * 100  # Convert to percentage
                },
                'forecast': []
            }
            
            # Process forecast data
            for day in data['forecast']['forecastday']:
                forecast_day = {
                    'date': day['date'],
                    'temp': day['day']['avgtemp_f'],  # Changed to match current data format
                    'condition': day['day']['condition']['text'],
                    'rain_chance': day['day']['daily_chance_of_rain'],
                    'uv_index': day['day']['uv'],
                    'wind_mph': day['day']['maxwind_mph']
                }
                weather_data['forecast'].append(forecast_day)
                
            logger.info(f"Weather data fetched for ZIP: {zip_code}")
            return weather_data
            
        except requests.RequestException as e:
            logger.error(f"Weather API request failed: {str(e)}")
            raise
            
        except Exception as e:
            logger.error(f"Error fetching weather data: {str(e)}")
            raise

    def _update_forecast_display(self):
        """Update the 5-day forecast display"""
        try:
            if not hasattr(self, 'weather_data') or not self.weather_data:
                return

            # Create forecast frames if they don't exist
            if not hasattr(self, 'forecast_frames'):
                self.forecast_frames = []
                forecast_container = ttk.Frame(self.weather_frame)
                forecast_container.pack(fill='x', pady=10)

                for _ in range(5):
                    day_frame = ttk.Frame(forecast_container)
                    day_frame.pack(side='left', expand=True, fill='x', padx=5)
                    
                    frames = {
                        'date': ttk.Label(day_frame),
                        'temp': ttk.Label(day_frame),
                        'condition': ttk.Label(day_frame),
                        'precip': ttk.Label(day_frame),
                        'uv_index': ttk.Label(day_frame),
                        'wind': ttk.Label(day_frame),
                        'pool_impact': ttk.Label(day_frame)
                    }
                    
                    for label in frames.values():
                        label.pack(pady=2)
                    
                    self.forecast_frames.append(frames)

            # Update forecast data
            for i, day in enumerate(self.weather_data.get('forecast', [])[:5]):
                try:
                    # Parse date
                    date_obj = datetime.strptime(day['date'], '%Y-%m-%d')
                    
                    # Update labels
                    frames = self.forecast_frames[i]
                    frames['date'].config(
                        text=date_obj.strftime('%A, %b %d'),
                        foreground=self.primary_color
                    )
                    frames['temp'].config(
                        text=f"Temperature: {day['temp']}°F"
                    )
                    frames['condition'].config(
                        text=f"Condition: {day['condition']}"
                    )
                    frames['precip'].config(
                        text=f"Rain Chance: {day['rain_chance']}%"
                    )
                    frames['uv_index'].config(
                        text=f"UV Index: {day['uv_index']}"
                    )
                    frames['wind'].config(
                        text=f"Wind: {day['wind_mph']} mph"
                    )

                    # Calculate and display pool impact
                    impact_level = self._calculate_forecast_impact(day)
                    impact_color = self._get_impact_color(impact_level)
                    frames['pool_impact'].config(
                        text=f"Pool Impact: {impact_level}",
                        foreground=impact_color
                    )

                except Exception as e:
                    logger.error(f"Error updating forecast day {i}: {str(e)}")

            logger.debug("Forecast display updated successfully")

        except Exception as e:
            logger.error(f"Error updating forecast display: {str(e)}")
            self._show_weather_error()

    def _update_impact_details(self, current):
        """Update detailed weather impact analysis"""
        try:
            if not hasattr(self, 'impact_details'):
                logger.error("Impact details not initialized")
                return

            # Temperature impact - Changed from temp_f to temp
            if current['temp'] > 90:
                self.impact_details['temperature'].config(
                    text=(
                        "• SEVERE Temperature Impact:\n"
                        "  - Very high chlorine depletion rate\n"
                        "  - Test chemistry twice daily\n"
                        "  - Add extra chlorine as needed\n"
                        "  - Consider using chlorine stabilizer\n"
                        "  - Monitor pH more frequently"
                    ),
                    foreground="#F44336"  # Red
                )
                self.impact_details['temperature'].pack(fill='x', padx=5)
            elif current['temp'] > 85:
                self.impact_details['temperature'].config(
                    text=(
                        "• HIGH Temperature Impact:\n"
                        "  - Increased chlorine depletion\n"
                        "  - Test chemistry daily\n"
                        "  - Monitor chlorine levels closely"
                    ),
                    foreground="#FF9800"  # Orange
                )
                self.impact_details['temperature'].pack(fill='x', padx=5)
            elif current['temp'] > 80:
                self.impact_details['temperature'].config(
                    text=(
                        "• Moderate Temperature Impact:\n"
                        "  - Normal chlorine depletion\n"
                        "  - Maintain regular testing schedule"
                    ),
                    foreground="#2196F3"  # Blue
                )
                self.impact_details['temperature'].pack(fill='x', padx=5)
            else:
                self.impact_details['temperature'].pack_forget()

            # UV impact
            if current['uv_index'] > 10:
                self.impact_details['uv'].config(
                    text=(
                        "• SEVERE UV Impact:\n"
                        "  - Extreme chlorine loss rate\n"
                        "  - Add extra chlorine stabilizer\n"
                        "  - Test frequently throughout day\n"
                        "  - Consider pool cover during peak hours"
                    ),
                    foreground="#F44336"  # Red
                )
                self.impact_details['uv'].pack(fill='x', padx=5)
            elif current['uv_index'] > 8:
                self.impact_details['uv'].config(
                    text=(
                        "• HIGH UV Impact:\n"
                        "  - Rapid chlorine loss\n"
                        "  - Add chlorine stabilizer\n"
                        "  - Test more frequently"
                    ),
                    foreground="#FF9800"  # Orange
                )
                self.impact_details['uv'].pack(fill='x', padx=5)
            elif current['uv_index'] > 5:
                self.impact_details['uv'].config(
                    text=(
                        "• Moderate UV Impact:\n"
                        "  - Normal chlorine loss\n"
                        "  - Monitor stabilizer levels"
                    ),
                    foreground="#2196F3"  # Blue
                )
                self.impact_details['uv'].pack(fill='x', padx=5)
            else:
                self.impact_details['uv'].pack_forget()

            # Rain impact
            if current['rain_chance'] > 80:
                self.impact_details['rain'].config(
                    text=(
                        "• HIGH Rain Impact:\n"
                        "  - Heavy rain likely\n"
                        "  - Check chemistry after rain\n"
                        "  - Monitor water level\n"
                        "  - Be prepared to adjust chemicals"
                    ),
                    foreground="#F44336"  # Red
                )
                self.impact_details['rain'].pack(fill='x', padx=5)
            elif current['rain_chance'] > 60:
                self.impact_details['rain'].config(
                    text=(
                        "• Moderate Rain Impact:\n"
                        "  - Rain likely\n"
                        "  - Check levels after rainfall\n"
                        "  - Monitor water balance"
                    ),
                    foreground="#FF9800"  # Orange
                )
                self.impact_details['rain'].pack(fill='x', padx=5)
            elif current['rain_chance'] > 40:
                self.impact_details['rain'].config(
                    text=(
                        "• Light Rain Impact:\n"
                        "  - Rain possible\n"
                        "  - Be prepared to test after"
                    ),
                    foreground="#2196F3"  # Blue
                )
                self.impact_details['rain'].pack(fill='x', padx=5)
            else:
                self.impact_details['rain'].pack_forget()

            logger.debug("Impact details updated successfully")

        except Exception as e:
            logger.error(f"Error updating impact details: {str(e)}")
            raise

    def _save_customer_info(self):
        """Save customer information"""
        try:
            # Collect customer info
            self.customer_info = {
                'name': self.customer_name_entry.get().strip(),
                'address': self.address_entry.get().strip(),
                'phone': self.phone_entry.get().strip(),
                'pool_type': self.pool_type_var.get()
            }
            
            # Save to file
            self._save_data()
            
            messagebox.showinfo("Success", "Customer information saved successfully!")
            logger.info("Customer information saved")
            
        except Exception as e:
            logger.error(f"Error saving customer info: {str(e)}")
            self._handle_error("Failed to save customer information")

    def _update_forecast_display(self):
        """Update the 5-day forecast display"""
        try:
            if not hasattr(self, 'weather_data') or not self.weather_data:
                return

            forecast_days = self.weather_data.get('forecast', [])
            
            for i, (frame, day) in enumerate(zip(self.forecast_frames, forecast_days)):
                try:
                    # Update date
                    date_obj = datetime.strptime(day['date'], '%Y-%m-%d')
                    frame['date'].config(
                        text=date_obj.strftime('%A, %b %d'),
                        foreground=self.primary_color
                    )
                    
                    # Update conditions with better formatting
                    frame['temp'].config(
                        text=f"Temperature: {day['temp']}°F",
                        foreground="#333333"
                    )
                    frame['condition'].config(
                        text=f"Condition: {day['condition']}",
                        foreground="#333333"
                    )
                    frame['precip'].config(
                        text=f"Rain Chance: {day['rain_chance']}%",
                        foreground="#333333"
                    )
                    frame['uv_index'].config(
                        text=f"UV Index: {day['uv_index']}",
                        foreground="#333333"
                    )
                    frame['wind'].config(
                        text=f"Wind: {day['wind_mph']} mph",
                        foreground="#333333"
                    )

                    # Analyze pool impact
                    impact_text = []
                    impact_color = "#4CAF50"  # Default green

                    # Temperature impact
                    if day['temp'] > 90:
                        impact_text.append("• HIGH Temperature Impact:")
                        impact_text.append("  - Rapid chlorine depletion")
                        impact_text.append("  - Test chemistry twice daily")
                        impact_text.append("  - Add extra chlorine")
                        impact_color = "#F44336"  # Red
                    elif day['temp'] > 85:
                        impact_text.append("• Moderate Temperature Impact:")
                        impact_text.append("  - Increased chlorine loss")
                        impact_text.append("  - Test chemistry daily")
                        impact_color = "#FF9800"  # Orange

                    # UV impact
                    if day['uv_index'] > 8:
                        impact_text.append("• HIGH UV Impact:")
                        impact_text.append("  - Rapid stabilizer depletion")
                        impact_text.append("  - Add extra stabilizer")
                        impact_text.append("  - Consider pool cover")
                        impact_color = "#F44336"  # Red
                    elif day['uv_index'] > 6:
                        impact_text.append("• Moderate UV Impact:")
                        impact_text.append("  - Monitor stabilizer levels")
                        if impact_color != "#F44336":
                            impact_color = "#FF9800"  # Orange

                    # Rain impact
                    if day['rain_chance'] > 70:
                        impact_text.append("• HIGH Rain Impact:")
                        impact_text.append("  - Check chemistry after rain")
                        impact_text.append("  - Monitor water level")
                        impact_text.append("  - Prepare to adjust chemicals")
                        if impact_color != "#F44336":
                            impact_color = "#FF9800"  # Orange
                    elif day['rain_chance'] > 40:
                        impact_text.append("• Moderate Rain Impact:")
                        impact_text.append("  - Be prepared to test after rain")

                    # Update impact text
                    if impact_text:
                        frame['pool_impact'].config(
                            text="\n".join(impact_text),
                            foreground=impact_color,
                            justify='left'
                        )
                    else:
                        frame['pool_impact'].config(
                            text="• Minimal Impact on Pool Chemistry",
                            foreground="#4CAF50",  # Green
                            justify='left'
                        )

                except Exception as e:
                    logger.error(f"Error updating forecast day {i}: {str(e)}")
                    frame['pool_impact'].config(
                        text="Error updating forecast",
                        foreground="#F44336"
                    )

            logger.debug("Forecast display updated successfully")
            
        except Exception as e:
            logger.error(f"Error updating forecast display: {str(e)}")
            self._show_weather_error()

    def _create_test_strip_tab(self):
        """Create test strip visualization tab"""
        try:
            # Create test strip tab
            self.test_strip_tab = ttk.Frame(self.notebook)
            self.notebook.add(self.test_strip_tab, text="Test Strip Analysis")

            # Create main container frame
            main_frame = ttk.Frame(self.test_strip_tab)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)

            # Create canvas for test strip visualization
            canvas_frame = ttk.Frame(main_frame)
            canvas_frame.pack(fill='both', expand=True)

            # Create canvas with scrollbar
            self.strip_canvas = tk.Canvas(
                canvas_frame,
                bg='white',
                width=1000,
                height=800,
                scrollregion=(0, 0, 1000, 1200)
            )
            
            # Create scrollbar
            scrollbar = ttk.Scrollbar(
                canvas_frame,
                orient="vertical",
                command=self.strip_canvas.yview
            )
            
            # Configure canvas scrolling
            self.strip_canvas.configure(yscrollcommand=scrollbar.set)
            
            # Pack scrollbar and canvas
            scrollbar.pack(side="right", fill="y")
            self.strip_canvas.pack(side="left", fill="both", expand=True)

            # Bind mousewheel for scrolling
            self._bind_mousewheel(self.strip_canvas)

            logger.info("Test strip tab created successfully")
            
        except Exception as e:
            logger.error(f"Error creating test strip tab: {str(e)}")
            raise

    def _bind_mousewheel(self, widget):
        """Bind mousewheel to widget for scrolling"""
        try:
            # Windows and macOS
            widget.bind("<MouseWheel>", self._on_mousewheel)
            # Linux
            widget.bind("<Button-4>", self._on_mousewheel)
            widget.bind("<Button-5>", self._on_mousewheel)
        except Exception as e:
            logger.error(f"Error binding mousewheel: {str(e)}")

    def _on_mousewheel(self, event):
        """Handle mousewheel scrolling"""
        try:
            widget = event.widget
            
            # Get the parent scrollable widget
            while widget and not isinstance(widget, (tk.Canvas, tk.Text, ttk.Treeview)):
                widget = widget.master
            
            if not widget:
                return
                
            # Determine scroll direction based on platform
            if event.num == 5 or event.delta < 0:
                widget.yview_scroll(1, "units")
            elif event.num == 4 or event.delta > 0:
                widget.yview_scroll(-1, "units")
            
        except Exception as e:
            logger.error(f"Error handling mousewheel: {str(e)}")

    def _update_test_strip(self):
        """Update the test strip visualization with current chemical values"""
        try:
            # Clear existing canvas
            self.strip_canvas.delete("all")

            # Define layout parameters with improved spacing and sizing
            layout = {
                'fonts': {
                    'parameter': ('Helvetica', 14, 'bold'),
                    'value': ('Helvetica', 12),
                    'range': ('Helvetica', 10),
                    'status': ('Helvetica', 12, 'bold')
                },
                'colors': {
                    'text': '#1976D2',  # Material Blue for headers
                    'bar_bg': '#F5F5F5',  # Light gray background
                    'ideal_range': '#E8F5E9',  # Light green for ideal range
                    'ideal_border': '#66BB6A',  # Darker green border
                    'value_good': '#4CAF50',  # Green for good values
                    'value_bad': '#F44336',  # Red for warning values
                    'below_range': '#FF9800',  # Orange for below range
                    'above_range': '#F44336'  # Red for above range
                },
                'bar_width': 450,  # Increased bar width
                'bar_height': 40,  # Increased bar height
                'spacing': 100,  # Increased spacing between parameters
                'padding': 20
            }

            # Draw title with better styling
            self.strip_canvas.create_text(
                50, 30,
                text="Chemical Test Strip Analysis",
                font=('Helvetica', 20, 'bold'),
                fill=self.primary_color,
                anchor='w'
            )

            # Parameter order and ranges
            all_params = [
                "Hardness (ppm)",
                "Free Chlorine (ppm)",
                "Alkalinity (ppm)",
                "Bromine (ppm)",
                "Cyanuric Acid (ppm)",
                "pH",
                "Total Chlorine (ppm)",
                "Salt (ppm)"
            ]

            # Split parameters into two columns
            mid_point = len(all_params) // 2
            left_column = all_params[:mid_point]
            right_column = all_params[mid_point:]

            # Draw columns with improved spacing
            y_offset = 80
            x_start_left = 50
            x_start_right = x_start_left + layout['bar_width'] + 200  # Increased column spacing

            # Draw left column
            for param in left_column:
                self._draw_parameter_bar(param, x_start_left, y_offset, layout)
                y_offset += layout['spacing']

            # Draw right column
            y_offset = 80
            for param in right_column:
                self._draw_parameter_bar(param, x_start_right, y_offset, layout)
                y_offset += layout['spacing']

            # Draw enhanced legend
            legend_y = max(len(left_column), len(right_column)) * layout['spacing'] + 100
            self._draw_enhanced_legend(50, legend_y, layout)

            logger.debug("Test strip visualization updated with enhanced layout")

        except Exception as e:
            logger.error(f"Error updating test strip: {str(e)}")
            self._handle_error("Failed to update test strip visualization")

    def _draw_parameter_bar(self, param, x_start, y, layout):
        """Draw an individual parameter bar with enhanced visuals"""
        try:
            # Get ranges and current value
            ideal_min, ideal_max = self.tester.ideal_levels[param]
            display_min, display_max = self.display_ranges[param]
            
            try:
                value = float(self.entries[param].get()) if self.entries[param].get() else None
            except ValueError:
                value = None

            # Draw parameter name with background
            name_bg = self.strip_canvas.create_rectangle(
                x_start - 5,
                y - 5,
                x_start + layout['bar_width'] + 5,
                y + 20,
                fill='#E3F2FD',  # Light blue background
                outline='#BBDEFB'
            )
            
            self.strip_canvas.create_text(
                x_start,
                y + 5,
                text=param,
                font=layout['fonts']['parameter'],
                fill=layout['colors']['text'],
                anchor='w'
            )

            # Draw bar background with rounded corners
            bar_y = y + 35
            self._draw_rounded_rectangle(
                x_start,
                bar_y,
                x_start + layout['bar_width'],
                bar_y + layout['bar_height'],
                fill=layout['colors']['bar_bg'],
                outline="#E0E0E0",
                radius=5
            )

            # Draw ideal range zone
            ideal_x1 = x_start + (ideal_min - display_min) / (display_max - display_min) * layout['bar_width']
            ideal_x2 = x_start + (ideal_max - display_min) / (display_max - display_min) * layout['bar_width']
            
            self._draw_rounded_rectangle(
                ideal_x1,
                bar_y,
                ideal_x2,
                bar_y + layout['bar_height'],
                fill=layout['colors']['ideal_range'],
                outline=layout['colors']['ideal_border'],
                radius=5
            )

            # Draw range labels with enhanced visibility
            self._draw_range_labels(x_start, bar_y, layout, param, display_min, display_max, ideal_min, ideal_max)

            # Draw current value if available
            if value is not None:
                self._draw_value_marker(x_start, bar_y, layout, value, ideal_min, ideal_max, display_min, display_max)

        except Exception as e:
            logger.error(f"Error drawing parameter bar: {str(e)}")
            raise

    def _draw_rounded_rectangle(self, x1, y1, x2, y2, radius=5, **kwargs):
        """Draw a rectangle with rounded corners"""
        points = [
            x1 + radius, y1,
            x2 - radius, y1,
            x2, y1,
            x2, y1 + radius,
            x2, y2 - radius,
            x2, y2,
            x2 - radius, y2,
            x1 + radius, y2,
            x1, y2,
            x1, y2 - radius,
            x1, y1 + radius,
            x1, y1
        ]
        return self.strip_canvas.create_polygon(points, smooth=True, **kwargs)

    def _draw_gradient(self, x, y, width, height, color1, color2, steps, vertical=False):
        """Draw a smooth gradient between two colors"""
        try:
            step_size = width if not vertical else height
            step_size = step_size / steps
            
            for i in range(steps):
                factor = i / (steps - 1)
                color = self._interpolate_color(color1, color2, factor)
                
                if vertical:
                    y1 = y + (i * step_size)
                    y2 = y + ((i + 1) * step_size)
                    self.strip_canvas.create_rectangle(
                        x, y1, x + width, y2,
                        fill=color, outline=color
                    )
                else:
                    x1 = x + (i * step_size)
                    x2 = x + ((i + 1) * step_size)
                    self.strip_canvas.create_rectangle(
                        x1, y, x2, y + height,
                        fill=color, outline=color
                    )
                    
        except Exception as e:
            logger.error(f"Error drawing gradient: {str(e)}")
            raise

    def _update_test_strip(self):
        """Update the test strip visualization with current chemical values"""
        try:
            # Clear existing canvas
            self.strip_canvas.delete("all")

            # Define layout parameters with improved spacing and sizing
            layout = {
                'fonts': {
                    'parameter': ('Helvetica', 14, 'bold'),
                    'value': ('Helvetica', 12),
                    'range': ('Helvetica', 10),
                    'status': ('Helvetica', 12, 'bold')
                },
                'colors': {
                    'text': '#1976D2',  # Material Blue for headers
                    'bar_bg': '#F5F5F5',  # Light gray background
                    'ideal_range': '#E8F5E9',  # Light green for ideal range
                    'ideal_border': '#66BB6A',  # Darker green border
                    'value_good': '#4CAF50',  # Green for good values
                    'value_bad': '#F44336',  # Red for warning values
                    'below_range': '#FF9800',  # Orange for below range
                    'above_range': '#F44336'  # Red for above range
                },
                'bar_width': 450,  # Increased bar width
                'bar_height': 40,  # Increased bar height
                'spacing': 100,  # Increased spacing between parameters
                'padding': 20
            }

            # Draw title with better styling
            self.strip_canvas.create_text(
                50, 30,
                text="Chemical Test Strip Analysis",
                font=('Helvetica', 20, 'bold'),
                fill=self.primary_color,
                anchor='w'
            )

            # Parameter order and ranges
            all_params = [
                "Hardness (ppm)",
                "Free Chlorine (ppm)",
                "Alkalinity (ppm)",
                "Bromine (ppm)",
                "Cyanuric Acid (ppm)",
                "pH",
                "Total Chlorine (ppm)",
                "Salt (ppm)"
            ]

            # Split parameters into two columns
            mid_point = len(all_params) // 2
            left_column = all_params[:mid_point]
            right_column = all_params[mid_point:]

            # Draw columns with improved spacing
            y_offset = 80
            x_start_left = 50
            x_start_right = x_start_left + layout['bar_width'] + 200  # Increased column spacing

            # Draw left column
            for param in left_column:
                self._draw_parameter_bar(param, x_start_left, y_offset, layout)
                y_offset += layout['spacing']

            # Draw right column
            y_offset = 80
            for param in right_column:
                self._draw_parameter_bar(param, x_start_right, y_offset, layout)
                y_offset += layout['spacing']

            # Draw enhanced legend
            legend_y = max(len(left_column), len(right_column)) * layout['spacing'] + 100
            self._draw_enhanced_legend(50, legend_y, layout)

            logger.debug("Test strip visualization updated with enhanced layout")

        except Exception as e:
            logger.error(f"Error updating test strip: {str(e)}")
            self._handle_error("Failed to update test strip visualization")

    def _draw_parameter_bar(self, param, x_start, y, layout):
        """Draw an individual parameter bar with enhanced visuals"""
        try:
            # Get ranges and current value
            ideal_min, ideal_max = self.tester.ideal_levels[param]
            display_min, display_max = self.display_ranges[param]
            
            try:
                value = float(self.entries[param].get()) if self.entries[param].get() else None
            except ValueError:
                value = None

            # Draw parameter name with background
            name_bg = self.strip_canvas.create_rectangle(
                x_start - 5,
                y - 5,
                x_start + layout['bar_width'] + 5,
                y + 20,
                fill='#E3F2FD',  # Light blue background
                outline='#BBDEFB'
            )
            
            self.strip_canvas.create_text(
                x_start,
                y + 5,
                text=param,
                font=layout['fonts']['parameter'],
                fill=layout['colors']['text'],
                anchor='w'
            )

            # Draw bar background with rounded corners
            bar_y = y + 35
            self._draw_rounded_rectangle(
                x_start,
                bar_y,
                x_start + layout['bar_width'],
                bar_y + layout['bar_height'],
                fill=layout['colors']['bar_bg'],
                outline="#E0E0E0",
                radius=5
            )

            # Draw ideal range zone
            ideal_x1 = x_start + (ideal_min - display_min) / (display_max - display_min) * layout['bar_width']
            ideal_x2 = x_start + (ideal_max - display_min) / (display_max - display_min) * layout['bar_width']
            
            self._draw_rounded_rectangle(
                ideal_x1,
                bar_y,
                ideal_x2,
                bar_y + layout['bar_height'],
                fill=layout['colors']['ideal_range'],
                outline=layout['colors']['ideal_border'],
                radius=5
            )

            # Draw range labels with enhanced visibility
            self._draw_range_labels(x_start, bar_y, layout, param, display_min, display_max, ideal_min, ideal_max)

            # Draw current value if available
            if value is not None:
                self._draw_value_marker(x_start, bar_y, layout, value, ideal_min, ideal_max, display_min, display_max)

        except Exception as e:
            logger.error(f"Error drawing parameter bar: {str(e)}")
            raise

    def _draw_rounded_rectangle(self, x1, y1, x2, y2, radius=5, **kwargs):
        """Draw a rectangle with rounded corners"""
        points = [
            x1 + radius, y1,
            x2 - radius, y1,
            x2, y1,
            x2, y1 + radius,
            x2, y2 - radius,
            x2, y2,
            x2 - radius, y2,
            x1 + radius, y2,
            x1, y2,
            x1, y2 - radius,
            x1, y1 + radius,
            x1, y1
        ]
        return self.strip_canvas.create_polygon(points, smooth=True, **kwargs)

    def _draw_range_labels(self, x_start, bar_y, layout, param, display_min, display_max, ideal_min, ideal_max):
        """Draw range labels with enhanced visibility"""
        try:
            # Format numbers based on parameter type
            def format_value(value, param):
                if "pH" in param:
                    return f"{value:.1f}"
                elif any(x in param for x in ["Chlorine", "Bromine"]):
                    return f"{value:.1f}"
                else:
                    return f"{int(value)}"

            # Draw min/max labels
            self.strip_canvas.create_text(
                x_start - 5,
                bar_y + layout['bar_height']/2,
                text=format_value(display_min, param),
                font=layout['fonts']['range'],
                fill=layout['colors']['text'],
                anchor='e'
            )

            self.strip_canvas.create_text(
                x_start + layout['bar_width'] + 5,
                bar_y + layout['bar_height']/2,
                text=format_value(display_max, param),
                font=layout['fonts']['range'],
                fill=layout['colors']['text'],
                anchor='w'
            )

            # Draw ideal range labels
            self.strip_canvas.create_text(
                x_start + (ideal_min - display_min) / (display_max - display_min) * layout['bar_width'],
                bar_y + layout['bar_height'] + 15,
                text=format_value(ideal_min, param),
                font=layout['fonts']['range'],
                fill=layout['colors']['text'],
                anchor='n'
            )

            self.strip_canvas.create_text(
                x_start + (ideal_max - display_min) / (display_max - display_min) * layout['bar_width'],
                bar_y + layout['bar_height'] + 15,
                text=format_value(ideal_max, param),
                font=layout['fonts']['range'],
                fill=layout['colors']['text'],
                anchor='n'
            )

        except Exception as e:
            logger.error(f"Error drawing range labels: {str(e)}")
            raise

    def _draw_value_marker(self, x_start, bar_y, layout, value, ideal_min, ideal_max, display_min, display_max):
        """Draw current value marker with enhanced visibility"""
        try:
            # Calculate marker position
            marker_x = x_start + (value - display_min) / (display_max - display_min) * layout['bar_width']
            
            # Determine status and color
            if value < ideal_min:
                color = layout['colors']['below_range']
            elif value > ideal_max:
                color = layout['colors']['above_range']
            else:
                color = layout['colors']['value_good']

            # Draw glow effect
            for i in range(3):
                size = 8 - (i * 2)
                opacity = 0.3 - (i * 0.1)
                glow_color = self._adjust_color(color, opacity)
                
                self.strip_canvas.create_oval(
                    marker_x - size,
                    bar_y + layout['bar_height']/2 - size,
                    marker_x + size,
                    bar_y + layout['bar_height']/2 + size,
                    fill=glow_color,
                    outline=''
                )

            # Draw marker
            self.strip_canvas.create_oval(
                marker_x - 6,
                bar_y + layout['bar_height']/2 - 6,
                marker_x + 6,
                bar_y + layout['bar_height']/2 + 6,
                fill=color,
                outline='white',
                width=2
            )

            # Format and draw value label
            if isinstance(value, float):
                formatted_value = f"{value:.1f}"
            else:
                formatted_value = str(value)

            # Draw value background
            text_width = len(formatted_value) * 7
            self.strip_canvas.create_rectangle(
                marker_x - text_width/2 - 5,
                bar_y - 25,
                marker_x + text_width/2 + 5,
                bar_y - 5,
                fill='white',
                outline=color
            )

            # Draw value text
            self.strip_canvas.create_text(
                marker_x,
                bar_y - 15,
                text=formatted_value,
                font=layout['fonts']['value'],
                fill=color,
                anchor='c'
            )

        except Exception as e:
            logger.error(f"Error drawing value marker: {str(e)}")
            raise

    def _adjust_color(self, color: str, opacity: float) -> str:
        """Adjust color opacity"""
        try:
            # Remove '#' if present
            color = color.lstrip('#')
            
            # Convert hex to RGB
            r = int(color[0:2], 16)
            g = int(color[2:4], 16)
            b = int(color[4:6], 16)
            
            # Calculate alpha value (0-255)
            a = int(opacity * 255)
            
            # Return RGB color with adjusted opacity
            factor = opacity
            r = int(r + (255 - r) * (1 - factor))
            g = int(g + (255 - g) * (1 - factor))
            b = int(b + (255 - b) * (1 - factor))
            
            return f'#{r:02x}{g:02x}{b:02x}'
            
        except Exception as e:
            logger.error(f"Error adjusting color opacity: {str(e)}")
            return color

    def _draw_enhanced_legend(self, x, y, layout):
        """Draw enhanced legend with modern styling"""
        try:
            # Create legend container with rounded corners
            container_width = 900
            container_height = 80
            
            # Draw legend background
            self._draw_rounded_rectangle(
                x, y,
                x + container_width,
                y + container_height,
                radius=10,
                fill='white',
                outline='#E0E0E0'
            )

            # Legend title
            self.strip_canvas.create_text(
                x + 20, y + 20,
                text="Reading Guide:",
                font=('Helvetica', 12, 'bold'),
                fill=layout['colors']['text'],
                anchor='w'
            )

            # Legend items with enhanced visual indicators
            items = [
                ("Ideal Range", layout['colors']['ideal_range'], layout['colors']['ideal_border'], "rectangle"),
                ("Below Range", layout['colors']['below_range'], layout['colors']['below_range'], "circle"),
                ("Above Range", layout['colors']['above_range'], layout['colors']['above_range'], "circle"),
                ("Current Value", layout['colors']['value_good'], layout['colors']['value_good'], "circle"),
                ("Warning Value", layout['colors']['value_bad'], layout['colors']['value_bad'], "circle")
            ]

            # Calculate spacing for legend items
            item_width = (container_width - 40) // len(items)
            
            for i, (text, fill, outline, shape) in enumerate(items):
                x_pos = x + 30 + (i * item_width)
                y_center = y + 50

                if shape == "rectangle":
                    # Draw rectangle indicator
                    self._draw_rounded_rectangle(
                        x_pos, y_center - 10,
                        x_pos + 30, y_center + 10,
                        radius=3,
                        fill=fill,
                        outline=outline
                    )
                else:
                    # Draw circle indicator
                    self.strip_canvas.create_oval(
                        x_pos + 5, y_center - 8,
                        x_pos + 21, y_center + 8,
                        fill=fill,
                        outline=outline,
                        width=2
                    )

                # Draw legend text
                text_item = self.strip_canvas.create_text(
                    x_pos + 40, y_center,
                    text=text,
                    font=('Helvetica', 11),
                    fill='#424242',
                    anchor='w'
                )

                # Create invisible rectangle for tooltip area
                text_bbox = self.strip_canvas.bbox(text_item)
                tooltip_area = self.strip_canvas.create_rectangle(
                    text_bbox[0] - 2,
                    text_bbox[1] - 2,
                    text_bbox[2] + 2,
                    text_bbox[3] + 2,
                    fill='',
                    outline='',
                    tags=f'tooltip_{text_item}'
                )

                # Add tooltip with canvas tag binding
                tooltip_text = self._get_legend_tooltip(text)
                self.strip_canvas.tag_bind(
                    f'tooltip_{text_item}',
                    '<Enter>',
                    lambda e, txt=tooltip_text: self._show_canvas_tooltip(e, txt)
                )
                self.strip_canvas.tag_bind(
                    f'tooltip_{text_item}',
                    '<Leave>',
                    self._hide_canvas_tooltip
                )

            logger.debug("Legend drawn successfully")

        except Exception as e:
            logger.error(f"Error drawing legend: {str(e)}")
            raise

    def _get_legend_tooltip(self, legend_item: str) -> str:
        """Get detailed tooltip text for legend items"""
        tooltips = {
            "Ideal Range": "The recommended range for optimal pool chemistry",
            "Below Range": "Chemical level is too low and needs adjustment",
            "Above Range": "Chemical level is too high and needs adjustment",
            "Current Value": "The current measured value is within ideal range",
            "Warning Value": "The current measured value needs immediate attention"
        }
        return tooltips.get(legend_item, "")

    def _show_canvas_tooltip(self, event, text):
        """Show tooltip for canvas items"""
        try:
            x = self.strip_canvas.winfo_rootx() + event.x + 20
            y = self.strip_canvas.winfo_rooty() + event.y + 20
            
            # Create tooltip window
            self.canvas_tooltip = tk.Toplevel(self)
            self.canvas_tooltip.wm_overrideredirect(True)
            self.canvas_tooltip.wm_geometry(f"+{x}+{y}")
            
            # Create tooltip label
            label = ttk.Label(
                self.canvas_tooltip,
                text=text,
                justify='left',
                background="#ffffe0",
                relief='solid',
                borderwidth=1,
                font=("Helvetica", "9", "normal")
            )
            label.pack()
            
        except Exception as e:
            logger.error(f"Error showing canvas tooltip: {str(e)}")

    def _hide_canvas_tooltip(self, event=None):
        """Hide canvas tooltip"""
        try:
            if hasattr(self, 'canvas_tooltip') and self.canvas_tooltip:
                self.canvas_tooltip.destroy()
                self.canvas_tooltip = None
        except Exception as e:
            logger.error(f"Error hiding canvas tooltip: {str(e)}")

    def _interpolate_color(self, color1: str, color2: str, factor: float) -> str:
        """
        Interpolate between two colors
        
        Args:
            color1 (str): Starting color in hex format
            color2 (str): Ending color in hex format
            factor (float): Interpolation factor (0-1)
            
        Returns:
            str: Interpolated color in hex format
        """
        try:
            # Handle different hex formats
            color1 = color1.lstrip('#')
            color2 = color2.lstrip('#')
            
            # Convert hex to RGB
            r1 = int(color1[0:2], 16)
            g1 = int(color1[2:4], 16)
            b1 = int(color1[4:6], 16)
            
            r2 = int(color2[0:2], 16)
            g2 = int(color2[2:4], 16)
            b2 = int(color2[4:6], 16)
            
            # Interpolate
            r = int(r1 + (r2 - r1) * factor)
            g = int(g1 + (g2 - g1) * factor)
            b = int(b1 + (b2 - b1) * factor)
            
            # Convert back to hex
            return f"#{r:02x}{g:02x}{b:02x}"
            
        except Exception as e:
            logger.error(f"Error interpolating colors: {str(e)}")
            return "#000000"

    def _create_trends_tab(self, tab_frame):
        """Create trends and history tab"""
        try:
            # Create main frame for trends
            trends_frame = ttk.Frame(tab_frame)
            trends_frame.pack(fill='both', expand=True, padx=40, pady=40)

            # Configure grid weights
            trends_frame.grid_columnconfigure(0, weight=3)  # Graph gets more space
            trends_frame.grid_columnconfigure(1, weight=1)  # Controls get less space
            trends_frame.grid_rowconfigure(1, weight=1)

            # Add title section
            title_frame = ttk.Frame(trends_frame)
            title_frame.grid(row=0, column=0, columnspan=2, sticky='ew', pady=(0, 20))

            title_label = ttk.Label(
                title_frame,
                text="Chemical History & Trends",
                font=("Helvetica", 24, "bold"),
                foreground=self.primary_color
            )
            title_label.pack(side='left')

            # Create left side (graph section)
            graph_frame = ttk.LabelFrame(
                trends_frame,
                text="Chemical Level History",
                padding=20,
                style="Custom.TLabelframe"
            )
            graph_frame.grid(row=1, column=0, sticky='nsew', padx=(0, 20))

            # Create matplotlib figure
            self.fig = Figure(figsize=(10, 6), dpi=100)
            self.ax = self.fig.add_subplot(111)
            
            # Create canvas
            self.canvas = FigureCanvasTkAgg(self.fig, master=graph_frame)
            self.canvas.draw()
            self.canvas.get_tk_widget().pack(fill='both', expand=True)

            # Add toolbar
            toolbar_frame = ttk.Frame(graph_frame)
            toolbar_frame.pack(fill='x')
            toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
            toolbar.update()

            # Create right side (controls)
            controls_frame = ttk.Frame(trends_frame)
            controls_frame.grid(row=1, column=1, sticky='nsew')

            # Parameter selection section
            param_frame = ttk.LabelFrame(
                controls_frame,
                text="Parameters",
                padding=20,
                style="Custom.TLabelframe"
            )
            param_frame.pack(fill='x', pady=(0, 20))

            # Create parameter checkboxes
            self.parameter_vars = {}
            for param in self.tester.ideal_levels.keys():
                var = tk.BooleanVar(value=False)
                self.parameter_vars[param] = var
                
                cb = ttk.Checkbutton(
                    param_frame,
                    text=param,
                    variable=var,
                    command=self._update_combined_graph
                )
                cb.pack(anchor='w', pady=2)
                
                # Add tooltip
                self._create_tooltip(cb, f"Show/hide {param} in graph")

            # Time range selection
            time_frame = ttk.LabelFrame(
                controls_frame,
                text="Time Range",
                padding=20,
                style="Custom.TLabelframe"
            )
            time_frame.pack(fill='x', pady=(0, 20))

            # Time range options
            self.time_range = tk.StringVar(value="1W")
            time_ranges = [
                ("24 Hours", "1D"),
                ("1 Week", "1W"),
                ("1 Month", "1M"),
                ("3 Months", "3M"),
                ("6 Months", "6M"),
                ("1 Year", "1Y")
            ]

            for text, value in time_ranges:
                rb = ttk.Radiobutton(
                    time_frame,
                    text=text,
                    value=value,
                    variable=self.time_range,
                    command=self._update_combined_graph
                )
                rb.pack(anchor='w', pady=2)
                
                # Add tooltip
                self._create_tooltip(rb, f"Show data for the last {text}")

            # Graph settings
            settings_frame = ttk.LabelFrame(
                controls_frame,
                text="Graph Settings",
                padding=20,
                style="Custom.TLabelframe"
            )
            settings_frame.pack(fill='x')

            # Show ideal range option
            self.show_ideal_range = tk.BooleanVar(value=True)
            ideal_cb = ttk.Checkbutton(
                settings_frame,
                text="Show Ideal Range",
                variable=self.show_ideal_range,
                command=self._update_combined_graph
            )
            ideal_cb.pack(anchor='w', pady=2)
            
            # Add tooltip
            self._create_tooltip(ideal_cb, "Show/hide ideal range bands on graph")

            # Show grid option
            self.show_grid = tk.BooleanVar(value=True)
            grid_cb = ttk.Checkbutton(
                settings_frame,
                text="Show Grid",
                variable=self.show_grid,
                command=self._update_combined_graph
            )
            grid_cb.pack(anchor='w', pady=2)
            
            # Add tooltip
            self._create_tooltip(grid_cb, "Show/hide grid lines on graph")

            # Export section
            export_frame = ttk.LabelFrame(
                controls_frame,
                text="Export Options",
                padding=20,
                style="Custom.TLabelframe"
            )
            export_frame.pack(fill='x', pady=(20, 0))

            # Export buttons
            export_excel_btn = ttk.Button(
                export_frame,
                text="Export to Excel",
                command=self._export_to_excel,
                style="Primary.TButton"
            )
            export_excel_btn.pack(fill='x', pady=(0, 5))
            
            export_pdf_btn = ttk.Button(
                export_frame,
                text="Export to PDF",
                command=self._export_to_pdf,
                style="Primary.TButton"
            )
            export_pdf_btn.pack(fill='x', pady=5)
            
            export_csv_btn = ttk.Button(
                export_frame,
                text="Export to CSV",
                command=self._export_to_csv,
                style="Primary.TButton"
            )
            export_csv_btn.pack(fill='x', pady=(5, 0))

            # Add tooltips for export buttons
            self._create_tooltip(export_excel_btn, "Export data to Excel spreadsheet")
            self._create_tooltip(export_pdf_btn, "Export data to PDF report")
            self._create_tooltip(export_csv_btn, "Export data to CSV file")

            # Initial graph update
            self._update_combined_graph()

            logger.info("Trends tab created successfully")
            
        except Exception as e:
            logger.error(f"Error creating trends tab: {str(e)}")
            raise

    def _update_graph(self):
        """Update the chemical trend graph"""
        try:
            # Clear existing plot
            self.ax.clear()

            # Get selected time range
            cutoff_date = self._get_cutoff_date(self.time_range.get())

            # Filter data based on time range
            if not hasattr(self, 'chemistry_data') or not self.chemistry_data.get('readings'):
                self._show_no_data_message()
                return

            readings = [
                reading for reading in self.chemistry_data['readings']
                if datetime.fromisoformat(reading['timestamp']) > cutoff_date
            ]

            if not readings:
                self._show_no_data_message()
                return

            # Plot data for selected parameters
            for param in self.selected_parameters:
                dates = [datetime.fromisoformat(r['timestamp']) for r in readings 
                        if param in r['values']]
                values = [r['values'][param] for r in readings if param in r['values']]

                if dates and values:
                    self.ax.plot(dates, values, 
                               label=param,
                               color=self.graph_colors.get(param, '#333333'),
                               marker='o')

                    # Add ideal range if enabled
                    if self.graph_settings['show_ideal_range']:
                        self._add_ideal_range(param)

            # Configure graph appearance
            self.ax.set_xlabel('Date')
            self.ax.set_ylabel('Value')
            self.ax.set_title('Chemical Trends')
            self.ax.grid(self.graph_settings['show_grid'])
            self.ax.legend()

            # Rotate x-axis labels for better readability
            plt.setp(self.ax.get_xticklabels(), rotation=45, ha='right')

            # Adjust layout to prevent label cutoff
            self.fig.tight_layout()

            # Update canvas
            self.canvas.draw()

            logger.debug("Graph updated successfully")

        except Exception as e:
            logger.error(f"Error updating graph: {str(e)}")
            self._show_graph_error()

    def _update_combined_graph(self):
        """Update the combined parameters graph"""
        try:
            self.ax.clear()
            
            # Get data range
            cutoff_date = self._get_cutoff_date(self.time_range.get())
            
            # Filter readings
            readings = [
                r for r in self.chemistry_data.get('readings', [])
                if datetime.fromisoformat(r['timestamp']) > cutoff_date
            ]
            
            if not readings:
                self._show_no_data_message()
                return
                
            # Plot each parameter
            for param in self.selected_parameters:
                dates = []
                values = []
                for reading in readings:
                    if param in reading['values']:
                        dates.append(datetime.fromisoformat(reading['timestamp']))
                        values.append(reading['values'][param])
                        
                if dates and values:
                    self.ax.plot(dates, values,
                               label=param,
                               color=self.graph_colors.get(param, '#333333'),
                               marker='o')
                               
            # Configure graph
            self.ax.set_xlabel('Date')
            self.ax.set_ylabel('Value')
            self.ax.set_title('Chemical Trends')
            self.ax.grid(True)
            self.ax.legend()
            
            # Update display
            self.canvas.draw()
            
            logger.debug("Combined graph updated successfully")
            
        except Exception as e:
            logger.error(f"Error updating combined graph: {str(e)}")
            self._show_graph_error()

    def _get_cutoff_date(self, time_range: str) -> datetime:
        """
        Get cutoff date based on selected time range
        
        Args:
            time_range (str): Time range code ('1D', '1W', '1M', '3M', '6M', '1Y')
            
        Returns:
            datetime: Cutoff date for data filtering
        """
        try:
            current_time = datetime.now()
            
            # Define time ranges
            ranges = {
                '1D': timedelta(days=1),
                '1W': timedelta(weeks=1),
                '1M': timedelta(days=30),
                '3M': timedelta(days=90),
                '6M': timedelta(days=180),
                '1Y': timedelta(days=365)
            }
            
            # Get appropriate timedelta
            delta = ranges.get(time_range, timedelta(weeks=1))  # Default to 1 week
            
            # Calculate and return cutoff date
            cutoff_date = current_time - delta
            
            logger.debug(f"Cutoff date calculated: {cutoff_date} for range: {time_range}")
            return cutoff_date
            
        except Exception as e:
            logger.error(f"Error calculating cutoff date: {str(e)}")
            # Return default of 1 week ago if error occurs
            return datetime.now() - timedelta(weeks=1)

    def _show_no_data_message(self):
        """Show message when no data is available for graph"""
        try:
            self.ax.text(0.5, 0.5, 
                        'No data available for selected time range',
                        horizontalalignment='center',
                        verticalalignment='center',
                        transform=self.ax.transAxes,
                        fontsize=12,
                        color='gray')
            self.canvas.draw()
            logger.debug("No data message displayed")
        except Exception as e:
            logger.error(f"Error showing no data message: {str(e)}")

    def _show_graph_error(self):
        """Show error message on graph"""
        try:
            self.ax.clear()
            self.ax.text(0.5, 0.5,
                        'Error updating graph.\nPlease check logs for details.',
                        horizontalalignment='center',
                        verticalalignment='center',
                        transform=self.ax.transAxes,
                        fontsize=12,
                        color='red')
            self.canvas.draw()
            logger.debug("Graph error message displayed")
        except Exception as e:
            logger.error(f"Error showing graph error message: {str(e)}")

    def _export_to_excel(self):
        """Export data to Excel file"""
        try:
            # Get file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Export to Excel"
            )
            
            if not file_path:
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Chemical Analysis"

            # Add header
            headers = ["Date", "Time"] + list(self.tester.ideal_levels.keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Add data
            if hasattr(self, 'chemistry_data') and self.chemistry_data.get('readings'):
                for row, reading in enumerate(self.chemistry_data['readings'], 2):
                    # Date and time
                    dt = datetime.fromisoformat(reading['timestamp'])
                    ws.cell(row=row, column=1, value=dt.strftime('%Y-%m-%d'))
                    ws.cell(row=row, column=2, value=dt.strftime('%H:%M:%S'))
                    
                    # Chemical values
                    for col, param in enumerate(self.tester.ideal_levels.keys(), 3):
                        value = reading['values'].get(param)
                        if value is not None:
                            cell = ws.cell(row=row, column=col, value=value)
                            cell.number_format = '0.00'

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Save workbook
            wb.save(file_path)
            messagebox.showinfo("Success", "Data exported to Excel successfully!")
            logger.info(f"Data exported to Excel: {file_path}")
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            self._handle_error("Failed to export to Excel")

    def _export_to_pdf(self):
        """Export data to PDF file"""
        try:
            # Get file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Export to PDF"
            )
            
            if not file_path:
                return

            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )

            # Create story
            story = []
            styles = getSampleStyleSheet()
            
            # Add title
            title = Paragraph(
                "Chemical Analysis Report",
                styles['Title']
            )
            story.append(title)
            story.append(Spacer(1, 12))

            # Add date
            date_text = Paragraph(
                f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles['Normal']
            )
            story.append(date_text)
            story.append(Spacer(1, 12))

            # Create data table
            if hasattr(self, 'chemistry_data') and self.chemistry_data.get('readings'):
                # Prepare headers
                headers = ["Date", "Time"] + list(self.tester.ideal_levels.keys())
                
                # Prepare data
                data = [headers]
                for reading in self.chemistry_data['readings']:
                    dt = datetime.fromisoformat(reading['timestamp'])
                    row = [
                        dt.strftime('%Y-%m-%d'),
                        dt.strftime('%H:%M:%S')
                    ]
                    for param in self.tester.ideal_levels.keys():
                        value = reading['values'].get(param, '')
                        row.append(f"{value:.2f}" if value else '')
                    data.append(row)

                # Create table
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BOX', (0, 0), (-1, -1), 2, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                
                story.append(table)

            # Build PDF
            doc.build(story)
            messagebox.showinfo("Success", "Data exported to PDF successfully!")
            logger.info(f"Data exported to PDF: {file_path}")
        
        except Exception as e:
            logger.error(f"Error exporting to PDF: {str(e)}")
            self._handle_error("Failed to export to PDF")

    def _export_to_csv(self):
        """Export data to CSV file"""
        try:
            # Get file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")],
                title="Export to CSV"
            )
            
            if not file_path:
                return

            # Write CSV file
            with open(file_path, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers
                headers = ["Date", "Time"] + list(self.tester.ideal_levels.keys())
                writer.writerow(headers)
                
                # Write data
                if hasattr(self, 'chemistry_data') and self.chemistry_data.get('readings'):
                    for reading in self.chemistry_data['readings']:
                        dt = datetime.fromisoformat(reading['timestamp'])
                        row = [
                            dt.strftime('%Y-%m-%d'),
                            dt.strftime('%H:%M:%S')
                        ]
                        for param in self.tester.ideal_levels.keys():
                            value = reading['values'].get(param, '')
                            row.append(f"{value:.2f}" if value else '')
                        writer.writerow(row)

            messagebox.showinfo("Success", "Data exported to CSV successfully!")
            logger.info(f"Data exported to CSV: {file_path}")
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            self._handle_error("Failed to export to CSV")

    def _send_report_email(self):
        """Send report via email using Google App Password"""
        try:
            # Check if email configuration exists
            if not hasattr(self, 'email_config') or not self.email_config:
                self._show_email_setup_dialog()
                return

            # Get recipient email
            recipient = simpledialog.askstring(
                "Send Report", 
                "Enter recipient email address:",
                parent=self
            )
            
            if not recipient:
                return

            # Validate email address
            if not self._validate_email(recipient):
                messagebox.showerror(
                    "Invalid Email",
                    "Please enter a valid email address"
                )
                return

            # Create message
            msg = MIMEMultipart()
            msg['From'] = self.email_config.get('sender_email')
            msg['To'] = recipient
            msg['Subject'] = "Pool Chemistry Report"
            
            # Add timestamp
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Create HTML body
            html_body = f"""
            <html>
                <head>
                    <style>
                        table {{
                            border-collapse: collapse;
                            width: 100%;
                            margin: 20px 0;
                        }}
                        th, td {{
                            border: 1px solid #ddd;
                            padding: 8px;
                            text-align: left;
                        }}
                        th {{
                            background-color: #4CAF50;
                            color: white;
                        }}
                        tr:nth-child(even) {{
                            background-color: #f2f2f2;
                        }}
                        .warning {{
                            color: red;
                            font-weight: bold;
                        }}
                        .good {{
                            color: green;
                        }}
                    </style>
                </head>
                <body>
                    <h2>Pool Chemistry Report</h2>
                    <p>Generated on: {current_time}</p>
                    <h3>Current Readings:</h3>
                    <table>
                        <tr>
                            <th>Parameter</th>
                            <th>Current Value</th>
                            <th>Ideal Range</th>
                            <th>Status</th>
                        </tr>
            """

            # Add chemical readings
            for param, entry in self.entries.items():
                try:
                    value = float(entry.get()) if entry.get() else None
                    if value is not None:
                        ideal_min, ideal_max = self.tester.ideal_levels[param]
                        status = "GOOD" if ideal_min <= value <= ideal_max else "ACTION REQUIRED"
                        status_class = "good" if status == "GOOD" else "warning"
                        
                        html_body += f"""
                        <tr>
                            <td>{param}</td>
                            <td>{value:.2f}</td>
                            <td>{ideal_min:.1f} - {ideal_max:.1f}</td>
                            <td class="{status_class}">{status}</td>
                        </tr>
                        """
                except ValueError:
                    continue

            html_body += """
                    </table>
                </body>
            </html>
            """

            msg.attach(MIMEText(html_body, 'html'))

            # Connect to Gmail SMTP
            try:
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                
                # Use Google App Password instead of regular password
                server.login(
                    self.email_config.get('sender_email'),
                    self.email_config.get('app_password')
                )
                
                # Send email
                server.send_message(msg)
                server.quit()
                
                messagebox.showinfo(
                    "Success",
                    "Report sent successfully!"
                )
                
                logger.info(f"Report sent to {recipient}")
                
            except Exception as e:
                logger.error(f"SMTP Error: {str(e)}")
                raise
                
        except Exception as e:
            logger.error(f"Error sending email: {str(e)}")
            self._handle_error("Failed to send email")

    def _send_report_sms(self):
        """Send report via SMS using Twilio"""
        try:
            from twilio.rest import Client

            # Create SMS dialog
            dialog = tk.Toplevel(self)
            dialog.title("Send Report SMS")
            dialog.geometry("300x200")
            dialog.transient(self)

            # Center dialog
            dialog.geometry("+%d+%d" % (
                self.winfo_rootx() + 50,
                self.winfo_rooty() + 50))

            # Phone number entry
            form_frame = ttk.LabelFrame(dialog, text="SMS Details", padding=10)
            form_frame.pack(fill='both', expand=True, padx=10, pady=5)

            ttk.Label(form_frame, text="Phone Number:").pack()
            phone_entry = ttk.Entry(form_frame, width=20)
            phone_entry.pack(pady=5)
            if hasattr(self, 'customer_info') and self.customer_info.get('phone'):
                phone_entry.insert(0, self.customer_info['phone'])

            def send_sms():
                try:
                    phone_number = phone_entry.get().strip()
                    if not phone_number:
                        messagebox.showerror("Error", "Please enter a phone number")
                        return

                    # Initialize Twilio client
                    client = Client(self.sms_config['account_sid'], 
                                 self.sms_config['auth_token'])

                    # Create message text
                    message_text = self._generate_sms_text()

                    # Send message
                    message = client.messages.create(
                        body=message_text,
                        from_=self.sms_config['twilio_number'],
                        to=phone_number
                    )

                    messagebox.showinfo("Success", f"Report sent to {phone_number}")
                    dialog.destroy()

                except Exception as e:
                    logger.error(f"Error sending SMS: {str(e)}")
                    messagebox.showerror("SMS Error", 
                                       "Failed to send text message. Check logs for details.")

            # Add buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill='x', padx=10, pady=10)
            ttk.Button(button_frame, text="Send", command=send_sms).pack(side='right', padx=5)
            ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side='right', padx=5)

        except Exception as e:
            logger.error(f"Error creating SMS dialog: {str(e)}")
            messagebox.showerror("Error", "Failed to create SMS dialog")

    def _generate_sms_text(self) -> str:
        """Generate SMS report text"""
        try:
            # Get current readings
            readings = []
            for param, entry in self.entries.items():
                try:
                    value = float(entry.get())
                    readings.append(f"{param}: {self._format_value(value, param)}")
                except ValueError:
                    continue
                    
            # Create message
            message = [
                "Pool Chemistry Report",
                f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                "",
                "Current Readings:",
                *readings,
                "",
                "For full report, check email."
            ]
            
            return "\n".join(message)
            
        except Exception as e:
            logger.error(f"Error generating SMS text: {str(e)}")
            return "Error generating report. Please check email for details."

    def _send_report_email(self):
        """Send report via email using Google App Password"""
        try:
            # Check if email configuration exists
            if not hasattr(self, 'email_config') or not self.email_config:
                self._show_email_setup_dialog()
                return

            # Get recipient email
            recipient = simpledialog.askstring(
                "Send Report", 
                "Enter recipient email address:",
                parent=self
            )
            
            if not recipient:
                return

            # Validate email address
            if not self._validate_email(recipient):
                messagebox.showerror(
                    "Invalid Email",
                    "Please enter a valid email address"
                )
                return

            # Create message
            msg = MIMEMultipart()
            msg['From'] = self.email_config.get('sender_email')
            msg['To'] = recipient
            msg['Subject'] = "Pool Chemistry Report"
            
            # Add timestamp
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Create HTML body
            html_body = f"""
            <html>
                <head>
                    <style>
                        table {{
                            border-collapse: collapse;
                            width: 100%;
                            margin: 20px 0;
                        }}
                        th, td {{
                            border: 1px solid #ddd;
                            padding: 8px;
                            text-align: left;
                        }}
                        th {{
                            background-color: #4CAF50;
                            color: white;
                        }}
                        tr:nth-child(even) {{
                            background-color: #f2f2f2;
                        }}
                        .warning {{
                            color: red;
                            font-weight: bold;
                        }}
                        .good {{
                            color: green;
                        }}
                    </style>
                </head>
                <body>
                    <h2>Pool Chemistry Report</h2>
                    <p>Generated on: {current_time}</p>
                    <h3>Current Readings:</h3>
                    <table>
                        <tr>
                            <th>Parameter</th>
                            <th>Current Value</th>
                            <th>Ideal Range</th>
                            <th>Status</th>
                        </tr>
            """

            # Add chemical readings
            for param, entry in self.entries.items():
                try:
                    value = float(entry.get()) if entry.get() else None
                    if value is not None:
                        ideal_min, ideal_max = self.tester.ideal_levels[param]
                        status = "GOOD" if ideal_min <= value <= ideal_max else "ACTION REQUIRED"
                        status_class = "good" if status == "GOOD" else "warning"
                        
                        html_body += f"""
                        <tr>
                            <td>{param}</td>
                            <td>{value:.2f}</td>
                            <td>{ideal_min:.1f} - {ideal_max:.1f}</td>
                            <td class="{status_class}">{status}</td>
                        </tr>
                        """
                except ValueError:
                    continue

            html_body += """
                    </table>
                </body>
            </html>
            """

            msg.attach(MIMEText(html_body, 'html'))

            # Connect to Gmail SMTP
            try:
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                
                # Use Google App Password instead of regular password
                server.login(
                    self.email_config.get('sender_email'),
                    self.email_config.get('app_password')
                )
                
                # Send email
                server.send_message(msg)
                server.quit()
                
                messagebox.showinfo(
                    "Success",
                    "Report sent successfully!"
                )
                
                logger.info(f"Report sent to {recipient}")
                
            except Exception as e:
                logger.error(f"SMTP Error: {str(e)}")
                raise
                
        except Exception as e:
            logger.error(f"Error sending email: {str(e)}")
            self._handle_error("Failed to send email")

    def _send_report_sms(self):
        """Send report via SMS using Twilio"""
        try:
            from twilio.rest import Client

            # Create SMS dialog
            dialog = tk.Toplevel(self)
            dialog.title("Send Report SMS")
            dialog.geometry("300x200")
            dialog.transient(self)

            # Center dialog
            dialog.geometry("+%d+%d" % (
                self.winfo_rootx() + 50,
                self.winfo_rooty() + 50))

            # Phone number entry
            form_frame = ttk.LabelFrame(dialog, text="SMS Details", padding=10)
            form_frame.pack(fill='both', expand=True, padx=10, pady=5)

            ttk.Label(form_frame, text="Phone Number:").pack()
            phone_entry = ttk.Entry(form_frame, width=20)
            phone_entry.pack(pady=5)
            if hasattr(self, 'customer_info') and self.customer_info.get('phone'):
                phone_entry.insert(0, self.customer_info['phone'])

            def send_sms():
                try:
                    phone_number = phone_entry.get().strip()
                    if not phone_number:
                        messagebox.showerror("Error", "Please enter a phone number")
                        return

                    # Initialize Twilio client
                    client = Client(self.sms_config['account_sid'], 
                                 self.sms_config['auth_token'])

                    # Create message text
                    message_text = self._generate_sms_text()

                    # Send message
                    message = client.messages.create(
                        body=message_text,
                        from_=self.sms_config['twilio_number'],
                        to=phone_number
                    )

                    messagebox.showinfo("Success", f"Report sent to {phone_number}")
                    dialog.destroy()

                except Exception as e:
                    logger.error(f"Error sending SMS: {str(e)}")
                    messagebox.showerror("SMS Error", 
                                       "Failed to send text message. Check logs for details.")

            # Add buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill='x', padx=10, pady=10)
            ttk.Button(button_frame, text="Send", command=send_sms).pack(side='right', padx=5)
            ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side='right', padx=5)

        except Exception as e:
            logger.error(f"Error creating SMS dialog: {str(e)}")
            messagebox.showerror("Error", "Failed to create SMS dialog")

    def _generate_sms_text(self) -> str:
        """Generate SMS report text"""
        try:
            # Get current readings
            readings = []
            for param, entry in self.entries.items():
                try:
                    value = float(entry.get())
                    readings.append(f"{param}: {self._format_value(value, param)}")
                except ValueError:
                    continue
                    
            # Create message
            message = [
                "Pool Chemistry Report",
                f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                "",
                "Current Readings:",
                *readings,
                "",
                "For full report, check email."
            ]
            
            return "\n".join(message)
            
        except Exception as e:
            logger.error(f"Error generating SMS text: {str(e)}")
            return "Error generating report. Please check email for details."

    def _create_email_settings(self, parent):
        """Create email settings section"""
        try:
            # Email configuration frame
            email_config_frame = ttk.LabelFrame(
                parent,
                text="Email Configuration",
                padding=10
            )
            email_config_frame.pack(fill='x', pady=(0, 10))

            # Sender email
            ttk.Label(
                email_config_frame,
                text="Sender Email:"
            ).pack(anchor='w')
            
            self.sender_email_var = tk.StringVar(
                value=self.email_config.get('sender_email', '')
            )
            ttk.Entry(
                email_config_frame,
                textvariable=self.sender_email_var
            ).pack(fill='x', pady=(0, 10))

            # App password entry
            ttk.Label(
                email_config_frame,
                text="App Password:"
            ).pack(anchor='w')
            
            self.app_password_var = tk.StringVar(
                value=self.email_config.get('app_password', '')
            )
            password_entry = ttk.Entry(
                email_config_frame,
                textvariable=self.app_password_var,
                show='*'
            )
            password_entry.pack(fill='x', pady=(0, 5))

            # Show/hide password
            self.show_password_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(
                email_config_frame,
                text="Show Password",
                variable=self.show_password_var,
                command=lambda: self._toggle_password_visibility(
                    password_entry,
                    self.show_password_var
                )
            ).pack(anchor='w')

            # Test email button
            ttk.Button(
                email_config_frame,
                text="Test Email Settings",
                command=self._test_email_settings,
                style="Primary.TButton"
            ).pack(pady=10)

            logger.debug("Email settings created successfully")

        except Exception as e:
            logger.error(f"Error creating email settings: {str(e)}")
            raise

    def _create_sms_settings(self, parent):
        """Create SMS settings section"""
        try:
            # SMS configuration frame
            sms_config_frame = ttk.LabelFrame(
                parent,
                text="Twilio SMS Configuration",
                padding=10
            )
            sms_config_frame.pack(fill='x', pady=(0, 10))

            # Account SID
            ttk.Label(
                sms_config_frame,
                text="Account SID:"
            ).pack(anchor='w')
            
            self.account_sid_var = tk.StringVar(
                value=self.sms_config.get('account_sid', '')
            )
            ttk.Entry(
                sms_config_frame,
                textvariable=self.account_sid_var
            ).pack(fill='x', pady=(0, 10))

            # Auth Token
            ttk.Label(
                sms_config_frame,
                text="Auth Token:"
            ).pack(anchor='w')
            
            self.auth_token_var = tk.StringVar(
                value=self.sms_config.get('auth_token', '')
            )
            token_entry = ttk.Entry(
                sms_config_frame,
                textvariable=self.auth_token_var,
                show='*'
            )
            token_entry.pack(fill='x', pady=(0, 5))

            # Show/hide token
            self.show_token_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(
                sms_config_frame,
                text="Show Token",
                variable=self.show_token_var,
                command=lambda: self._toggle_password_visibility(
                    token_entry,
                    self.show_token_var
                )
            ).pack(anchor='w')

            # Twilio Phone Number
            ttk.Label(
                sms_config_frame,
                text="Twilio Phone Number:"
            ).pack(anchor='w', pady=(10, 0))
            
            self.twilio_number_var = tk.StringVar(
                value=self.sms_config.get('twilio_number', '')
            )
            ttk.Entry(
                sms_config_frame,
                textvariable=self.twilio_number_var
            ).pack(fill='x', pady=(0, 10))

            # Test SMS button
            ttk.Button(
                sms_config_frame,
                text="Test SMS Settings",
                command=self._test_sms_settings,
                style="Primary.TButton"
            ).pack(pady=10)

            logger.debug("SMS settings created successfully")

        except Exception as e:
            logger.error(f"Error creating SMS settings: {str(e)}")
            raise

    def _create_alert_settings(self, parent):
        """Create alert settings section"""
        try:
            # Alert configuration frame
            alert_config_frame = ttk.LabelFrame(
                parent,
                text="Alert Configuration",
                padding=10
            )
            alert_config_frame.pack(fill='x', pady=(0, 10))

            # Enable alerts
            self.alerts_enabled_var = tk.BooleanVar(
                value=self.alert_settings.get('enabled', True)
            )
            ttk.Checkbutton(
                alert_config_frame,
                text="Enable Automatic Alerts",
                variable=self.alerts_enabled_var
            ).pack(anchor='w', pady=(0, 10))

            # Alert methods
            methods_frame = ttk.LabelFrame(
                alert_config_frame,
                text="Alert Methods",
                padding=10
            )
            methods_frame.pack(fill='x', pady=(0, 10))

            self.email_alerts_var = tk.BooleanVar(
                value=self.alert_settings.get('email_enabled', True)
            )
            ttk.Checkbutton(
                methods_frame,
                text="Email Alerts",
                variable=self.email_alerts_var
            ).pack(anchor='w')

            self.sms_alerts_var = tk.BooleanVar(
                value=self.alert_settings.get('sms_enabled', False)
            )
            ttk.Checkbutton(
                methods_frame,
                text="SMS Alerts",
                variable=self.sms_alerts_var
            ).pack(anchor='w')

            # Alert frequency
            frequency_frame = ttk.LabelFrame(
                alert_config_frame,
                text="Alert Frequency",
                padding=10
            )
            frequency_frame.pack(fill='x', pady=(0, 10))

            ttk.Label(
                frequency_frame,
                text="Minimum time between alerts:"
            ).pack(anchor='w')

            self.alert_frequency_var = tk.StringVar(
                value=self.alert_settings.get('frequency', '24H')
            )
            frequency_combo = ttk.Combobox(
                frequency_frame,
                textvariable=self.alert_frequency_var,
                values=['1H', '4H', '12H', '24H', '48H'],
                state='readonly'
            )
            frequency_combo.pack(fill='x', pady=5)

            # Alert thresholds
            thresholds_frame = ttk.LabelFrame(
                alert_config_frame,
                text="Alert Thresholds",
                padding=10
            )
            thresholds_frame.pack(fill='x')

            # Create threshold entries for each parameter
            self.threshold_vars = {}
            for param in self.tester.ideal_levels.keys():
                param_frame = ttk.Frame(thresholds_frame)
                param_frame.pack(fill='x', pady=2)

                ttk.Label(
                    param_frame,
                    text=f"{param}:"
                ).pack(side='left')

                var = tk.StringVar(
                    value=str(self.alert_settings.get(
                        'thresholds', {}).get(param, '10')
                    )
                )
                self.threshold_vars[param] = var

                ttk.Entry(
                    param_frame,
                    textvariable=var,
                    width=10
                ).pack(side='right')

                ttk.Label(
                    param_frame,
                    text="% deviation"
                ).pack(side='right', padx=5)

            logger.debug("Alert settings created successfully")

        except Exception as e:
            logger.error(f"Error creating alert settings: {str(e)}")
            raise

    def _toggle_password_visibility(self, entry_widget, show_var):
        """Toggle password/token visibility"""
        try:
            if show_var.get():
                entry_widget.config(show='')
            else:
                entry_widget.config(show='*')
        except Exception as e:
            logger.error(f"Error toggling password visibility: {str(e)}")

    def _save_settings(self, dialog):
        """Save all settings"""
        try:
            # Save general settings
            self.settings = {
                'theme': self.theme_var.get(),
                'data_retention': self.retention_var.get(),
                'auto_backup': self.backup_enabled_var.get(),
                'backup_path': self.backup_path_var.get()
            }

            # Save graph settings
            self.graph_settings = {
                'show_grid': self.show_grid_var.get(),
                'show_points': self.show_points_var.get()
            }

            # Save email settings
            self.email_config = {
                'sender_email': self.sender_email_var.get(),
                'app_password': self.app_password_var.get()
            }

            # Save SMS settings
            self.sms_config = {
                'account_sid': self.account_sid_var.get(),
                'auth_token': self.auth_token_var.get(),
                'twilio_number': self.twilio_number_var.get()
            }

            # Save alert settings
            self.alert_settings = {
                'enabled': self.alerts_enabled_var.get(),
                'email_enabled': self.email_alerts_var.get(),
                'sms_enabled': self.sms_alerts_var.get(),
                'frequency': self.alert_frequency_var.get(),
                'thresholds': {
                    param: float(var.get())
                    for param, var in self.threshold_vars.items()
                }
            }

            # Apply theme
            self.style.theme_use(self.settings['theme'])

            # Save to file
            self._save_settings_to_file()

            # Update graph if needed
            if hasattr(self, 'canvas'):
                self._update_combined_graph()

            messagebox.showinfo("Success", "Settings saved successfully!")
            dialog.destroy()

            logger.info("Settings saved successfully")

        except Exception as e:
            logger.error(f"Error saving settings: {str(e)}")
            self._handle_error("Failed to save settings")

    def _save_settings_to_file(self):
        """Save settings to JSON file"""
        try:
            settings_data = {
                'general': self.settings,
                'graph': self.graph_settings,
                'email': {
                    'sender_email': self.email_config.get('sender_email'),
                    # Don't save password in plain text
                    'app_password': self._encrypt_sensitive_data(
                        self.email_config.get('app_password', '')
                    )
                },
                'sms': {
                    'account_sid': self.sms_config.get('account_sid'),
                    'auth_token': self._encrypt_sensitive_data(
                        self.sms_config.get('auth_token', '')
                    ),
                    'twilio_number': self.sms_config.get('twilio_number')
                },
                'alerts': self.alert_settings
            }

            # Ensure settings directory exists
            self.settings_file.parent.mkdir(parents=True, exist_ok=True)

            # Save to file
            with open(self.settings_file, 'w') as f:
                json.dump(settings_data, f, indent=4)

            logger.info("Settings saved to file successfully")

        except Exception as e:
            logger.error(f"Error saving settings to file: {str(e)}")
            raise

    def _load_settings_from_file(self):
        """Load settings from JSON file"""
        try:
            if self.settings_file.exists():
                with open(self.settings_file, 'r') as f:
                    settings_data = json.load(f)

                # Load general settings
                self.settings = settings_data.get('general', {})
                self.graph_settings = settings_data.get('graph', {})

                # Load email settings
                email_data = settings_data.get('email', {})
                self.email_config = {
                    'sender_email': email_data.get('sender_email'),
                    'app_password': self._decrypt_sensitive_data(
                        email_data.get('app_password', '')
                    )
                }

                # Load SMS settings
                sms_data = settings_data.get('sms', {})
                self.sms_config = {
                    'account_sid': sms_data.get('account_sid'),
                    'auth_token': self._decrypt_sensitive_data(
                        sms_data.get('auth_token', '')
                    ),
                    'twilio_number': sms_data.get('twilio_number')
                }

                # Load alert settings
                self.alert_settings = settings_data.get('alerts', {})

                # Apply theme if specified
                if 'theme' in self.settings:
                    self.style.theme_use(self.settings['theme'])

                logger.info("Settings loaded from file successfully")

            else:
                logger.info("No settings file found, using defaults")
                self._initialize_default_settings()

        except Exception as e:
            logger.error(f"Error loading settings: {str(e)}")
            self._initialize_default_settings()

    def _test_email_settings(self):
        """Test email configuration"""
        try:
            # Create test message
            msg = MIMEMultipart()
            msg['From'] = self.sender_email_var.get()
            msg['To'] = self.sender_email_var.get()  # Send to self
            msg['Subject'] = "Test Email - Pool Chemistry Monitor"

            body = "This is a test email from your Pool Chemistry Monitor application."
            msg.attach(MIMEText(body, 'plain'))

            # Connect to Gmail SMTP
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()

            # Try to login
            try:
                server.login(
                    self.sender_email_var.get(),
                    self.app_password_var.get()
                )
                
                # Send email
                server.send_message(msg)
                server.quit()
                
                messagebox.showinfo(
                    "Success",
                    "Test email sent successfully!"
                )
                
                logger.info("Test email sent successfully")
                
            except smtplib.SMTPAuthenticationError:
                messagebox.showerror(
                    "Authentication Error",
                    "Invalid email or app password. Please check your credentials."
                )
                logger.error("Email authentication failed")
                
            except Exception as e:
                raise
                
        except Exception as e:
            logger.error(f"Error testing email settings: {str(e)}")
            self._handle_error("Failed to send test email")

    def _test_sms_settings(self):
        """Test SMS configuration"""
        try:
            from twilio.rest import Client

            # Initialize Twilio client
            client = Client(
                self.account_sid_var.get(),
                self.auth_token_var.get()
            )

            # Send test message
            message = client.messages.create(
                body="Test message from Pool Chemistry Monitor",
                from_=self.twilio_number_var.get(),
                to=self.twilio_number_var.get()  # Send to self
            )

            messagebox.showinfo(
                "Success",
                "Test SMS sent successfully!"
            )
            
            logger.info("Test SMS sent successfully")
            
        except Exception as e:
            logger.error(f"Error testing SMS settings: {str(e)}")
            self._handle_error("Failed to send test SMS")

    def _encrypt_sensitive_data(self, data: str) -> str:
        """
        Encrypt sensitive data using Fernet encryption
        
        Args:
            data (str): Data to encrypt
            
        Returns:
            str: Encrypted data in base64 format
        """
        try:
            if not data:
                return ""

            # Generate key if not exists
            if not hasattr(self, 'encryption_key'):
                self.encryption_key = Fernet.generate_key()
                key_file = Path("data/encryption.key")
                key_file.parent.mkdir(parents=True, exist_ok=True)
                with open(key_file, 'wb') as f:
                    f.write(self.encryption_key)

            # Create Fernet instance
            fernet = Fernet(self.encryption_key)
            
            # Encrypt data
            encrypted_data = fernet.encrypt(data.encode())
            
            # Convert to base64 for storage
            return base64.b64encode(encrypted_data).decode()
            
        except Exception as e:
            logger.error(f"Error encrypting data: {str(e)}")
            return ""

    def _decrypt_sensitive_data(self, encrypted_data: str) -> str:
        """
        Decrypt sensitive data using Fernet encryption
        
        Args:
            encrypted_data (str): Encrypted data in base64 format
            
        Returns:
            str: Decrypted data
        """
        try:
            if not encrypted_data:
                return ""

            # Load encryption key
            if not hasattr(self, 'encryption_key'):
                key_file = Path("data/encryption.key")
                if not key_file.exists():
                    return ""
                with open(key_file, 'rb') as f:
                    self.encryption_key = f.read()

            # Create Fernet instance
            fernet = Fernet(self.encryption_key)
            
            # Decode base64 and decrypt
            encrypted_bytes = base64.b64decode(encrypted_data)
            decrypted_data = fernet.decrypt(encrypted_bytes)
            
            return decrypted_data.decode()
            
        except Exception as e:
            logger.error(f"Error decrypting data: {str(e)}")
            return ""

    def _save_data(self):
        """Save application data to file"""
        try:
            # Create a clean data structure without recursive references
            data = {
                'customer_info': {
                    'name': self.customer_info.get('name', ''),
                    'address': self.customer_info.get('address', ''),
                    'phone': self.customer_info.get('phone', ''),
                    'pool_type': self.customer_info.get('pool_type', 'Select Pool Type')
                },
                'chemistry_data': {
                    'readings': self.chemistry_data.get('readings', []) if hasattr(self, 'chemistry_data') else []
                },
                'last_update': datetime.now().isoformat()
            }

            # Ensure data directory exists
            self.data_file.parent.mkdir(parents=True, exist_ok=True)

            # Save to file
            with open(self.data_file, 'w') as f:
                json.dump(data, f, indent=4, default=str)

            # Create backup if enabled
            if self.settings.get('auto_backup', True):
                self._create_backup('auto')

            logger.info("Data saved successfully")
            
        except Exception as e:
            logger.error(f"Error saving data: {str(e)}")
            self._handle_error("Failed to save data")

    def _load_data(self):
        """Load application data from file"""
        try:
            if self.data_file.exists():
                with open(self.data_file, 'r') as f:
                    data = json.load(f)

                # Load customer info with default values if missing
                self.customer_info = {
                    'name': data.get('customer_info', {}).get('name', ''),
                    'address': data.get('customer_info', {}).get('address', ''),
                    'phone': data.get('customer_info', {}).get('phone', ''),
                    'pool_type': data.get('customer_info', {}).get('pool_type', 'Select Pool Type')
                }

                # Update customer info fields if they exist
                if 'customer_name_entry' in self.__dict__:
                    self.customer_name_entry.delete(0, tk.END)
                    self.customer_name_entry.insert(0, self.customer_info['name'])
                
                if hasattr(self, 'address_entry'):
                    self.address_entry.delete(0, tk.END)
                    self.address_entry.insert(0, self.customer_info['address'])
                
                if hasattr(self, 'phone_entry'):
                    self.phone_entry.delete(0, tk.END)
                    self.phone_entry.insert(0, self.customer_info['phone'])

                # Load chemistry data
                self.chemistry_data = {
                    'readings': data.get('chemistry_data', {}).get('readings', [])
                }

                # Update last update time
                self.last_update = datetime.fromisoformat(
                    data.get('last_update', datetime.now().isoformat())
                )

                logger.info("Data loaded successfully")
                
            else:
                logger.info("No data file found, initializing empty data")
                self._initialize_empty_data()
                
        except Exception as e:
            logger.error(f"Error loading data: {str(e)}")
            self._initialize_empty_data()

    def _create_backup(self):
        """Create backup of current data"""
        try:
            # Get backup directory
            backup_dir = Path(self.settings.get('backup_path', './backup'))
            backup_dir.mkdir(parents=True, exist_ok=True)

            # Create timestamp for backup file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = backup_dir / f"pool_data_backup_{timestamp}.json"

            # Copy current data file
            if self.data_file.exists():
                shutil.copy2(self.data_file, backup_file)

                # Remove old backups if needed
                self._cleanup_old_backups(backup_dir)

                logger.info(f"Backup created: {backup_file}")
                
        except Exception as e:
            logger.error(f"Error creating backup: {str(e)}")
            self._handle_error("Failed to create backup")

    def _cleanup_old_backups(self, backup_dir: Path):
        """Remove old backup files"""
        try:
            # Get all backup files
            backup_files = list(backup_dir.glob("pool_data_backup_*.json"))
            
            # Sort by modification time
            backup_files.sort(key=lambda x: x.stat().st_mtime)
            
            # Keep only last 5 backups
            while len(backup_files) > 5:
                oldest = backup_files.pop(0)
                oldest.unlink()
                logger.debug(f"Removed old backup: {oldest}")
                
        except Exception as e:
            logger.error(f"Error cleaning up backups: {str(e)}")

    def _handle_error(self, message: str, show_dialog: bool = True):
        """Handle application errors"""
        try:
            # Get error details
            exc_type, exc_value, exc_traceback = sys.exc_info()
            
            # Format error message
            error_msg = f"Error: {message}\n"
            if exc_value:
                error_msg += f"Details: {str(exc_value)}"
                
            # Log error
            logger.error(error_msg)
            if exc_traceback:
                logger.error("Traceback:", exc_info=True)
                
            # Show dialog if requested
            if show_dialog:
                self.after_idle(lambda: messagebox.showerror("Error", error_msg))
                
            # Add to error queue for status display
            self._add_to_error_queue(error_msg)
            
        except Exception as e:
            logger.critical(f"Error handler failed: {str(e)}")

    def _validate_email(self, email: str) -> bool:
        """
        Validate email address format
        
        Args:
            email (str): Email address to validate
            
        Returns:
            bool: True if valid, False otherwise
        """
        try:
            # Basic email validation pattern
            pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            return bool(re.match(pattern, email))
        except Exception as e:
            logger.error(f"Error validating email: {str(e)}")
            return False

    def _validate_phone(self, phone: str) -> bool:
        """
        Validate phone number format
        
        Args:
            phone (str): Phone number to validate
            
        Returns:
            bool: True if valid, False otherwise
        """
        try:
            # Remove any non-numeric characters
            cleaned = re.sub(r'\D', '', phone)
            # Check if length is valid (10 or 11 digits)
            return len(cleaned) in (10, 11)
        except Exception as e:
            logger.error(f"Error validating phone: {str(e)}")
            return False

    def _format_value(self, value: float, param: str) -> str:
        """
        Format numeric value based on parameter type
        
        Args:
            value (float): Value to format
            param (str): Parameter name
            
        Returns:
            str: Formatted value
        """
        try:
            if "pH" in param:
                return f"{value:.1f}"
            elif any(x in param for x in ["Chlorine", "Bromine"]):
                return f"{value:.1f} ppm"
            else:
                return f"{int(value)} ppm"
        except Exception as e:
            logger.error(f"Error formatting value: {str(e)}")
            return str(value)

    def _initialize_empty_data(self):
        """Initialize empty data structures"""
        try:
            # Initialize with basic structure
            self.customer_info = {
                'name': '',
                'address': '',
                'phone': '',
                'pool_type': 'Select Pool Type'
            }
            
            self.chemistry_data = {
                'readings': []
            }
            
            self.last_update = datetime.now()
            
            logger.debug("Empty data initialized")
            
        except Exception as e:
            logger.error(f"Error initializing empty data: {str(e)}")
            raise

    def _initialize_default_settings(self):
        """Initialize default application settings"""
        try:
            self.settings = {
                'theme': 'default',
                'data_retention': '6M',
                'auto_backup': True,
                'backup_path': './backup'
            }
            
            self.graph_settings = {
                'show_grid': True,
                'show_points': True
            }
            
            self.email_config = {
                'sender_email': '',
                'app_password': ''
            }
            
            self.sms_config = {
                'account_sid': '',
                'auth_token': '',
                'twilio_number': ''
            }
            
            self.alert_settings = {
                'enabled': True,
                'email_enabled': True,
                'sms_enabled': False,
                'frequency': '24H',
                'thresholds': {
                    param: 10.0 for param in self.tester.ideal_levels.keys()
                }
            }
            
            logger.info("Default settings initialized")
            
        except Exception as e:
            logger.error(f"Error initializing default settings: {str(e)}")
            raise

    def _browse_backup_path(self):
        """Open directory browser for backup location"""
        try:
            path = filedialog.askdirectory(
                title="Select Backup Directory",
                initialdir=self.settings.get('backup_path', './backup')
            )
            
            if path:
                self.backup_path_var.set(path)
                
        except Exception as e:
            logger.error(f"Error browsing backup path: {str(e)}")
            self._handle_error("Failed to select backup directory")

    def _show_weather_error(self):
        """Display weather error message"""
        try:
            if hasattr(self, 'weather_frame'):
                for widget in self.weather_frame.winfo_children():
                    widget.destroy()
                    
                ttk.Label(
                    self.weather_frame,
                    text="Weather data unavailable",
                    foreground="red"
                ).pack(pady=10)
                
        except Exception as e:
            logger.error(f"Error showing weather error: {str(e)}")

    def _cleanup(self):
        """Perform cleanup before application exit"""
        try:
            # Cancel all scheduled tasks
            for attr in dir(self):
                if attr.endswith('_id'):
                    try:
                        self.after_cancel(getattr(self, attr))
                    except:
                        pass

            # Save current state
            self._save_data()
            self._save_settings_to_file()
            self._save_task_history()

            # Clean up matplotlib resources
            if hasattr(self, 'fig'):
                plt.close(self.fig)
            plt.close('all')  # Close any remaining figures

            # Clean up GUI resources
            for widget in self.winfo_children():
                try:
                    widget.destroy()
                except:
                    pass

            # Clean up tooltips
            if hasattr(self, 'tooltips'):
                for tooltip in self.tooltips:
                    try:
                        tooltip.destroy()
                    except:
                        pass

            # Clean up files
            self._cleanup_temp_files()
            self._compress_old_data()

            logger.info("Application cleanup completed")

        except Exception as e:
            logger.error(f"Error during cleanup: {str(e)}")

    def _resize_graph(self):
        """Resize matplotlib graph"""
        try:
            if hasattr(self, 'fig') and hasattr(self, 'canvas'):
                # Get new size
                width = self.winfo_width() - 100  # Padding
                height = self.winfo_height() - 200  # Padding
                
                # Update figure size
                self.fig.set_size_inches(
                    width / self.fig.get_dpi(),
                    height / self.fig.get_dpi()
                )
                
                # Adjust layout
                self.fig.tight_layout()
                
                # Redraw canvas
                self.canvas.draw()
                
        except Exception as e:
            logger.error(f"Error resizing graph: {str(e)}")

    def _confirm_exit(self):
        """Show exit confirmation dialog"""
        try:
            if messagebox.askokcancel(
                "Confirm Exit",
                "Are you sure you want to exit?\nAny unsaved changes will be lost."
            ):
                self._on_closing()
                
        except Exception as e:
            logger.error(f"Error showing exit confirmation: {str(e)}")
            self._on_closing()

    def _show_about(self):
        """Show about dialog"""
        try:
            dialog = tk.Toplevel(self)
            dialog.title("About Pool Chemistry Monitor")
            dialog.geometry("400x500")
            dialog.transient(self)
            dialog.grab_set()

            # Center dialog
            dialog.geometry(f"+{self.winfo_x() + 50}+{self.winfo_y() + 50}")

            # Add content
            ttk.Label(
                dialog,
                text="Pool Chemistry Monitor",
                font=("Helvetica", 16, "bold"),
                foreground=self.primary_color
            ).pack(pady=20)

            ttk.Label(
                dialog,
                text="Version 2.0",
                font=("Helvetica", 12)
            ).pack()

            ttk.Label(
                dialog,
                text="© 2025 Virtual Control LLC",
                font=("Helvetica", 10)
            ).pack(pady=10)

            # Description
            description = ttk.Label(
                dialog,
                text=(
                    "Professional pool chemistry management software\n"
                    "for monitoring, analysis, and recommendations.\n\n"
                    "Features:\n"
                    "• Real-time chemical monitoring\n"
                    "• Weather impact analysis\n"
                    "• Automated recommendations\n"
                    "• Test strip visualization\n"
                    "• Trend analysis and graphs\n"
                    "• PDF/Excel report generation\n"
                    "• Email and SMS notifications"
                ),
                justify="center"
            )
            description.pack(pady=20)

            # Close button
            ttk.Button(
                dialog,
                text="Close",
                command=dialog.destroy
            ).pack(pady=20)
            
        except Exception as e:
            logger.error(f"Error showing about dialog: {str(e)}")
            if 'dialog' in locals():
                dialog.destroy()

    def _update_recommendations(self):
        """Update chemical recommendations based on current readings"""
        try:
            # Clear existing recommendations
            if 'recommendation_text' in self.__dict__:
                self.recommendation_text.configure(state='normal')
                self.recommendation_text.delete(1.0, tk.END)

            # Get current values
            readings = {}
            for param, entry in self.entries.items():
                try:
                    value = float(entry.get()) if entry.get() else None
                    if value is not None:
                        readings[param] = value
                except ValueError:
                    continue

            if not readings:
                self._show_no_readings_message()
                return

            # Get pool type
            pool_type = self.pool_type_var.get()
            if pool_type == "Select Pool Type":
                self.recommendation_text.insert(tk.END, 
                    "Please select a pool type to get recommendations\n", "warning")
                self.recommendation_text.configure(state='disabled')
                return

            # Calculate recommendations
            recommendations = self._calculate_recommendations(readings)

            # Display recommendations
            self._display_recommendations(recommendations, readings)

            # Save readings
            self._save_readings(readings)

            logger.debug("Recommendations updated successfully")

        except Exception as e:
            logger.error(f"Error updating recommendations: {str(e)}")
            self._handle_error("Failed to update recommendations")

    def _calculate_recommendations(self, readings: dict) -> dict:
        """
        Calculate chemical adjustments needed
        
        Args:
            readings (dict): Current chemical readings
            
        Returns:
            dict: Recommended adjustments
        """
        try:
            recommendations = {
                'adjustments': [],
                'warnings': [],
                'status': 'good'
            }

            # Get pool volume and type
            try:
                pool_volume = float(self.volume_entry.get())
            except ValueError:
                recommendations['warnings'].append("Please enter a valid pool volume")
                return recommendations

            pool_type = self.pool_type_var.get()
            if pool_type == 'Select Pool Type':
                recommendations['warnings'].append("Please select a pool type")
                return recommendations

            # Check each parameter
            for param, value in readings.items():
                ideal_min, ideal_max = self.tester.ideal_levels[param]
                
                if value < ideal_min:
                    adjustment = self._calculate_increase(
                        param, value, ideal_min, pool_volume, pool_type
                    )
                    recommendations['adjustments'].append(adjustment)
                    recommendations['status'] = 'attention'
                    
                elif value > ideal_max:
                    adjustment = self._calculate_decrease(
                        param, value, ideal_max, pool_volume, pool_type
                    )
                    recommendations['adjustments'].append(adjustment)
                    recommendations['status'] = 'attention'

            # Check chemical interactions
            if 'pH' in readings and 'Total Chlorine' in readings:
                if readings['pH'] > 7.8 and readings['Total Chlorine'] < 2.0:
                    recommendations['warnings'].append(
                        "High pH may be reducing chlorine effectiveness"
                    )

            # Check weather impact
            if hasattr(self, 'weather_data'):
                weather_recommendations = self._get_weather_recommendations()
                recommendations['weather'] = weather_recommendations

            return recommendations

        except Exception as e:
            logger.error(f"Error calculating recommendations: {str(e)}")
            raise

    def _calculate_increase(self, param: str, current: float, target: float, 
                           volume: float, pool_type: str) -> dict:
        """
        Calculate chemical increase needed
        
        Args:
            param (str): Chemical parameter
            current (float): Current level
            target (float): Target level
            volume (float): Pool volume
            pool_type (str): Type of pool
            
        Returns:
            dict: Adjustment details
        """
        try:
            difference = target - current
            
            # Get chemical factors based on pool type
            if pool_type == 'Select Pool Type':
                # Use default factors if no pool type selected
                factors = {
                    "pH": 1.0,
                    "Free Chlorine": 1.0,
                    "Total Chlorine": 1.0,
                    "Alkalinity": 1.0,
                    "Hardness": 1.0,
                    "Cyanuric Acid": 1.0,
                    "Salt": 1.0,
                    "Bromine": 1.0
                }
            else:
                factors = self.tester.chemical_factors[pool_type]
            
            if param in factors:
                factor = factors[param]
                amount = difference * volume * factor
                
                return {
                    'param': param,
                    'action': 'increase',
                    'amount': round(amount, 2),
                    'unit': 'oz',
                    'chemical': self.tester.increase_chemicals[param],
                    'notes': self.tester.chemical_warnings.get(param, '')
                }
            else:
                return {
                    'param': param,
                    'action': 'increase',
                    'amount': 0,
                    'unit': 'oz',
                    'chemical': 'Unknown',
                    'notes': 'Chemical factor not found'
                }

        except Exception as e:
            logger.error(f"Error calculating increase: {str(e)}")
            raise

    def _calculate_decrease(self, param: str, current: float, target: float, 
                           volume: float, pool_type: str) -> dict:
        """
        Calculate chemical decrease needed
        
        Args:
            param (str): Chemical parameter
            current (float): Current level
            target (float): Target level
            volume (float): Pool volume
            pool_type (str): Type of pool
            
        Returns:
            dict: Adjustment details
        """
        try:
            difference = current - target
            
            # Get chemical factors based on pool type
            factors = self.tester.chemical_factors[pool_type]
            
            if param in factors:
                factor = factors[param]
                amount = difference * volume * factor
                
                return {
                    'param': param,
                    'action': 'decrease',
                    'amount': round(amount, 2),
                    'unit': 'oz',
                    'chemical': self.tester.decrease_chemicals[param],
                    'notes': self.tester.decrease_notes.get(param, '')
                }
            else:
                return {
                    'param': param,
                    'action': 'decrease',
                    'amount': 0,
                    'unit': 'oz',
                    'chemical': 'Unknown',
                    'notes': 'Chemical factor not found'
                }

        except Exception as e:
            logger.error(f"Error calculating decrease: {str(e)}")
            raise

    def _display_recommendations(self, recommendations: dict, readings: dict):
        """Display chemical recommendations in formatted text"""
        try:
            if not hasattr(self, 'recommendation_text'):
                return

            self.recommendation_text.delete(1.0, tk.END)
            
            # Add header
            self.recommendation_text.insert(tk.END, 
                "Pool Chemistry Recommendations\n", "header")
            self.recommendation_text.insert(tk.END, 
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            # Add current readings
            self.recommendation_text.insert(tk.END, 
                "Current Readings:\n", "section")
            for param, value in readings.items():
                ideal_min, ideal_max = self.tester.ideal_levels[param]
                status = "good" if ideal_min <= value <= ideal_max else "warning"
                
                self.recommendation_text.insert(tk.END, 
                    f"{param}: {self._format_value(value, param)} "
                    f"(Ideal: {ideal_min}-{ideal_max})\n", status)
            
            self.recommendation_text.insert(tk.END, "\n")

            # Add adjustments needed
            if recommendations['adjustments']:
                self.recommendation_text.insert(tk.END, 
                    "Recommended Adjustments:\n", "section")
                
                for adj in recommendations['adjustments']:
                    self.recommendation_text.insert(tk.END,
                        f"• {adj['param']}: {adj['action'].title()} by adding "
                        f"{adj['amount']} {adj['unit']} of {adj['chemical']}\n")
                    if adj['notes']:
                        self.recommendation_text.insert(tk.END,
                            f"  Note: {adj['notes']}\n")
                
                self.recommendation_text.insert(tk.END, "\n")

            # Add warnings
            if recommendations['warnings']:
                self.recommendation_text.insert(tk.END, 
                    "Warnings:\n", "section")
                for warning in recommendations['warnings']:
                    self.recommendation_text.insert(tk.END,
                        f"• {warning}\n", "warning")
                
                self.recommendation_text.insert(tk.END, "\n")

            # Add weather impact
            if 'weather' in recommendations:
                self.recommendation_text.insert(tk.END, 
                    "Weather Impact:\n", "section")
                for impact in recommendations['weather']:
                    self.recommendation_text.insert(tk.END,
                        f"• {impact}\n", "normal")
                
                self.recommendation_text.insert(tk.END, "\n")

            # Add maintenance schedule
            self._add_maintenance_schedule()

            logger.debug("Recommendations displayed successfully")

        except Exception as e:
            logger.error(f"Error displaying recommendations: {str(e)}")
            raise

    def _add_maintenance_schedule(self):
        """Add maintenance schedule to recommendations"""
        try:
            self.recommendation_text.insert(tk.END, 
                "Maintenance Schedule:\n", "section")

            schedule = [
                ("Daily:", [
                    "Check chlorine and pH levels",
                    "Run pool pump for recommended hours",
                    "Empty skimmer and pump baskets"
                ]),
                ("Weekly:", [
                    "Test all chemical levels",
                    "Brush pool walls and floor",
                    "Clean waterline",
                    "Backwash filter if needed"
                ]),
                ("Monthly:", [
                    "Deep clean filter",
                    "Check filter pressure",
                    "Inspect equipment for leaks",
                    "Clean deck and coping"
                ]),
                ("Quarterly:", [
                    "Acid wash filter",
                    "Check and balance stabilizer",
                    "Inspect pool surface",
                    "Clean salt cell (if applicable)"
                ])
            ]

            for period, tasks in schedule:
                self.recommendation_text.insert(tk.END, 
                    f"\n{period}\n", "normal")
                for task in tasks:
                    self.recommendation_text.insert(tk.END,
                        f"• {task}\n", "normal")

            logger.debug("Maintenance schedule added successfully")

        except Exception as e:
            logger.error(f"Error adding maintenance schedule: {str(e)}")
            raise

    def _save_readings(self, readings: dict):
        """Save current readings to history"""
        try:
            # Create reading entry
            reading = {
                'timestamp': datetime.now().isoformat(),
                'values': readings,
                'weather': self.__dict__.get('weather_data', None)
            }

            # Add to chemistry data
            if not hasattr(self, 'chemistry_data'):
                self.chemistry_data = {'readings': []}
            
            self.chemistry_data['readings'].append(reading)

            # Trim old readings based on retention setting
            self._trim_old_readings()

            # Save to file
            self._save_data()

            # Update graphs
            self._update_combined_graph()

            logger.info("Readings saved successfully")

        except Exception as e:
            logger.error(f"Error saving readings: {str(e)}")
            raise

    def _trim_old_readings(self):
        """Remove readings older than retention period"""
        try:
            if not self.chemistry_data.get('readings'):
                return

            retention = self.settings.get('data_retention', '6M')
            cutoff_date = self._get_cutoff_date(retention)

            # Filter readings
            self.chemistry_data['readings'] = [
                r for r in self.chemistry_data['readings']
                if datetime.fromisoformat(r['timestamp']) > cutoff_date
            ]

            logger.debug(f"Old readings trimmed based on {retention} retention")

        except Exception as e:
            logger.error(f"Error trimming old readings: {str(e)}")
            raise

    def _show_no_readings_message(self):
        """Display message when no readings are available"""
        try:
            if 'recommendation_text' in self.__dict__:
                self.recommendation_text.delete(1.0, tk.END)
                self.recommendation_text.insert(tk.END,
                    "No readings available.\n"
                    "Please enter chemical readings to get recommendations.",
                    "warning"
                )
        except Exception as e:
            logger.error(f"Error showing no readings message: {str(e)}")

    def _display_recommendations(self, recommendations: dict, readings: dict):
        """Display chemical recommendations in formatted text"""
        try:
            if not hasattr(self, 'recommendation_text'):
                return

            self.recommendation_text.delete(1.0, tk.END)
            
            # Add header
            self.recommendation_text.insert(tk.END, 
                "Pool Chemistry Recommendations\n", "header")
            self.recommendation_text.insert(tk.END,
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            # Add current readings
            self.recommendation_text.insert(tk.END, 
                "Current Readings:\n", "section")
            for param, value in readings.items():
                ideal_min, ideal_max = self.tester.ideal_levels[param]
                status = "success" if ideal_min <= value <= ideal_max else "warning"
                
                self.recommendation_text.insert(tk.END,
                    f"{param}: {self._format_value(value, param)} "
                    f"(Ideal: {ideal_min}-{ideal_max})\n", status)
            
            self.recommendation_text.insert(tk.END, "\n")

            # Add chemical adjustments
            if recommendations['adjustments']:
                self.recommendation_text.insert(tk.END, 
                    "Recommended Adjustments:\n", "section")
                
                for adj in recommendations['adjustments']:
                    self.recommendation_text.insert(tk.END,
                        f"• {adj['param']}: {adj['action'].title()} by "
                        f"{adj['amount']} {adj['unit']} using {adj['chemical']}\n")
                    if adj['notes']:
                        self.recommendation_text.insert(tk.END,
                            f"  Note: {adj['notes']}\n", "alert")
                
                self.recommendation_text.insert(tk.END, "\n")

            # Add warnings
            if recommendations['warnings']:
                self.recommendation_text.insert(tk.END, 
                    "Warnings:\n", "section")
                for warning in recommendations['warnings']:
                    self.recommendation_text.insert(tk.END,
                        f"• {warning}\n", "warning")
                
                self.recommendation_text.insert(tk.END, "\n")

            # Add weather impact
            if 'weather' in recommendations:
                self.recommendation_text.insert(tk.END, 
                    "Weather Impact:\n", "section")
                self._add_weather_impact(recommendations['weather'])

            logger.debug("Recommendations displayed successfully")

        except Exception as e:
            logger.error(f"Error displaying recommendations: {str(e)}")
            raise

    def _get_weather_recommendations(self) -> dict:
        """Generate weather-based recommendations"""
        try:
            if not hasattr(self, 'weather_data') or not self.weather_data:
                return {}

            recommendations = {
                'impacts': [],
                'actions': []
            }

            current = self.weather_data['current']

            # Temperature impact
            if current['temp'] > 90:
                recommendations['impacts'].append({
                    'factor': 'Temperature',
                    'level': 'High',
                    'effect': 'Increased chlorine depletion'
                })
                recommendations['actions'].append(
                    "Monitor chlorine levels more frequently"
                )
                recommendations['actions'].append(
                    "Consider adding chlorine stabilizer"
                )

            # UV impact
            if current['uv_index'] > 8:
                recommendations['impacts'].append({
                    'factor': 'UV',
                    'level': 'High',
                    'effect': 'Rapid chlorine loss'
                })
                recommendations['actions'].append(
                    "Add additional chlorine stabilizer"
                )
                recommendations['actions'].append(
                    "Consider using pool cover during peak hours"
                )

            # Rain impact
            if current['rain_chance'] > 60:
                recommendations['impacts'].append({
                    'factor': 'Rain',
                    'level': 'Likely',
                    'effect': 'Potential chemical dilution'
                })
                recommendations['actions'].append(
                    "Test chemistry after rainfall"
                )
                recommendations['actions'].append(
                    "Be prepared to adjust chemicals"
                )

            return recommendations

        except Exception as e:
            logger.error(f"Error getting weather recommendations: {str(e)}")
            return {}

    def _add_weather_impact(self, weather_recommendations: dict):
        """Add weather impact details to recommendations"""
        try:
            if not weather_recommendations:
                self.recommendation_text.insert(tk.END,
                    "Weather data unavailable\n", "alert")
                return

            # Add impacts
            if weather_recommendations.get('impacts'):
                for impact in weather_recommendations['impacts']:
                    self.recommendation_text.insert(tk.END,
                        f"• {impact['factor']}: {impact['level']} - "
                        f"{impact['effect']}\n", "alert")

            # Add actions
            if weather_recommendations.get('actions'):
                self.recommendation_text.insert(tk.END, 
                    "\nRecommended Weather Actions:\n", "section")
                for action in weather_recommendations['actions']:
                    self.recommendation_text.insert(tk.END,
                        f"• {action}\n", "normal")

            logger.debug("Weather impact added to recommendations")

        except Exception as e:
            logger.error(f"Error adding weather impact: {str(e)}")
            raise

    def _show_no_readings_message(self):
        """Display message when no readings are available"""
        try:
            if 'recommendation_text' in self.__dict__:
                self.recommendation_text.delete(1.0, tk.END)
                self.recommendation_text.insert(tk.END,
                    "No chemical readings available.\n"
                    "Please enter current chemical levels.", "alert")
                
        except Exception as e:
            logger.error(f"Error showing no readings message: {str(e)}")

    def _arduino_connection_loop(self):
        """Maintain Arduino connection and read data"""
        while not self._stop_threads.is_set():
            try:
                if not self.arduino_connected:
                    self._connect_arduino()
                else:
                    self._read_arduino_data()
                    
                time.sleep(1)  # Prevent CPU overload
                
            except serial.SerialException as e:
                logger.error(f"Serial connection error: {str(e)}")
                self.arduino_connected = False
                time.sleep(5)  # Wait before retry
                
            except Exception as e:
                logger.error(f"Arduino connection loop error: {str(e)}")
                time.sleep(5)  # Wait before retry

    def _connect_arduino(self):
        """Establish connection with Arduino"""
        try:
            # Search for Arduino port
            ports = list(serial.tools.list_ports.comports())
            for port in ports:
                if "Arduino" in port.description:
                    self.arduino_port = port.device
                    self.serial_conn = serial.Serial(
                        port=self.arduino_port,
                        baudrate=9600,
                        timeout=1
                    )
                    self.arduino_connected = True
                    logger.info(f"Connected to Arduino on {self.arduino_port}")
                    break
                    
        except Exception as e:
            logger.error(f"Error connecting to Arduino: {str(e)}")
            self.arduino_connected = False

    def _read_arduino_data(self):
        """Read and process data from Arduino"""
        try:
            if self.serial_conn and self.serial_conn.in_waiting:
                # Read line from Arduino
                data = self.serial_conn.readline().decode('utf-8').strip()
                
                # Process JSON data
                if data:
                    try:
                        sensor_data = json.loads(data)
                        self._process_sensor_data(sensor_data)
                    except json.JSONDecodeError:
                        logger.error("Invalid JSON data from Arduino")
                        
        except Exception as e:
            logger.error(f"Error reading Arduino data: {str(e)}")
            self.arduino_connected = False

    def _process_sensor_data(self, sensor_data: dict):
        """Process sensor readings from Arduino"""
        try:
            # Validate sensor data
            if not self._validate_sensor_data(sensor_data):
                logger.error("Invalid sensor data format")
                return

            # Update UI with sensor readings
            for param, value in sensor_data.items():
                if param in self.entries:
                    self.entries[param].delete(0, tk.END)
                    self.entries[param].insert(0, str(value))

            # Update recommendations if auto-update enabled
            if self.settings.get('auto_update', True):
                self._update_recommendations()

            # Log sensor reading
            self._log_sensor_reading(sensor_data)
            
            logger.debug("Sensor data processed successfully")
            
        except Exception as e:
            logger.error(f"Error processing sensor data: {str(e)}")

    def _validate_sensor_data(self, data: dict) -> bool:
        """
        Validate sensor data format and values
        
        Args:
            data (dict): Sensor data to validate
            
        Returns:
            bool: True if valid, False otherwise
        """
        try:
            required_params = {'pH', 'Chlorine', 'Temperature'}
            
            # Check required parameters
            if not all(param in data for param in required_params):
                return False
                
            # Validate ranges
            if not (0 <= data.get('pH', -1) <= 14):
                return False
            if not (0 <= data.get('Chlorine', -1) <= 10):
                return False
            if not (0 <= data.get('Temperature', -1) <= 120):
                return False
                
            return True
            
        except Exception as e:
            logger.error(f"Error validating sensor data: {str(e)}")
            return False

    def _log_sensor_reading(self, data: dict):
        """Log sensor reading to history"""
        try:
            # Create reading entry
            reading = {
                'timestamp': datetime.now().isoformat(),
                'source': 'sensor',
                'values': data
            }

            # Add to chemistry data
            if not hasattr(self, 'chemistry_data'):
                self.chemistry_data = {'readings': []}
                
            self.chemistry_data['readings'].append(reading)

            # Save to file
            self._save_data()
            
            logger.debug("Sensor reading logged successfully")
            
        except Exception as e:
            logger.error(f"Error logging sensor reading: {str(e)}")

    def _calibrate_ph_sensor(self):
        """Calibrate pH sensor"""
        try:
            if not self.arduino_connected:
                messagebox.showerror(
                    "Error",
                    "Arduino not connected. Please check connection."
                )
                return

            # Create calibration dialog
            dialog = tk.Toplevel(self)
            dialog.title("pH Sensor Calibration")
            dialog.geometry("400x300")
            dialog.transient(self)
            dialog.grab_set()

            # Add calibration instructions
            ttk.Label(
                dialog,
                text="pH Sensor Calibration",
                font=("Helvetica", 14, "bold")
            ).pack(pady=10)

            instructions = ttk.Label(
                dialog,
                text=(
                    "1. Clean the pH sensor\n"
                    "2. Place sensor in pH 7.0 solution\n"
                    "3. Wait for reading to stabilize\n"
                    "4. Click 'Calibrate pH 7.0'\n"
                    "5. Repeat with pH 4.0 solution"
                ),
                justify="left"
            )
            instructions.pack(pady=20)

            # Add calibration buttons
            ttk.Button(
                dialog,
                text="Calibrate pH 7.0",
                command=lambda: self._send_calibration_command("pH7")
            ).pack(pady=5)

            ttk.Button(
                dialog,
                text="Calibrate pH 4.0",
                command=lambda: self._send_calibration_command("pH4")
            ).pack(pady=5)

            ttk.Button(
                dialog,
                text="Close",
                command=dialog.destroy
            ).pack(pady=20)
            
            logger.info("pH calibration dialog shown")
            
        except Exception as e:
            logger.error(f"Error showing calibration dialog: {str(e)}")
            self._handle_error("Failed to show calibration dialog")

    def _send_calibration_command(self, command: str):
        """Send calibration command to Arduino"""
        try:
            if not self.arduino_connected or not self.serial_conn:
                messagebox.showerror(
                    "Error",
                    "Arduino not connected. Please check connection."
                )
                return

            # Send command to Arduino
            self.serial_conn.write(f"{command}\n".encode())
            
            # Wait for response
            time.sleep(2)
            response = self.serial_conn.readline().decode('utf-8').strip()

            if "SUCCESS" in response:
                messagebox.showinfo(
                    "Success",
                    f"Calibration for {command} completed successfully!"
                )
            else:
                messagebox.showerror(
                    "Error",
                    f"Calibration failed: {response}"
                )
                
            logger.info(f"Calibration command {command} sent")
            
        except Exception as e:
            logger.error(f"Error sending calibration command: {str(e)}")
            self._handle_error("Failed to send calibration command")

    def _run_tests(self):
        """Run diagnostic tests"""
        try:
            # Create test dialog
            dialog = tk.Toplevel(self)
            dialog.title("Run Diagnostic Tests")
            dialog.geometry("500x600")
            dialog.transient(self)
            dialog.grab_set()

            # Add scrolled text widget for results
            results_text = scrolledtext.ScrolledText(
                dialog,
                width=60,
                height=30,
                font=("Courier", 10)
            )
            results_text.pack(padx=10, pady=10)

            def run_selected_tests():
                """Run selected diagnostic tests"""
                try:
                    results_text.delete(1.0, tk.END)
                    results_text.insert(tk.END, "Running diagnostic tests...\n\n")

                    # Test database connection
                    self._test_database(results_text)

                    # Test Arduino connection
                    self._test_arduino(results_text)

                    # Test sensor readings
                    self._test_sensors(results_text)

                    # Test data saving
                    self._test_data_saving(results_text)

                    # Test weather API
                    self._test_weather_api(results_text)

                    results_text.insert(tk.END, "\nDiagnostic tests completed.")
                    
                    logger.info("Diagnostic tests completed")
                    
                except Exception as e:
                    logger.error(f"Error running diagnostic tests: {str(e)}")
                    results_text.insert(tk.END, f"\nError: {str(e)}")

            # Add buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill='x', padx=10, pady=10)

            ttk.Button(
                button_frame,
                text="Run Tests",
                command=run_selected_tests,
                style="Primary.TButton"
            ).pack(side='left', padx=5)

            ttk.Button(
                button_frame,
                text="Close",
                command=dialog.destroy
            ).pack(side='right', padx=5)

            logger.info("Test dialog shown")

        except Exception as e:
            logger.error(f"Error showing test dialog: {str(e)}")
            self._handle_error("Failed to show test dialog")

    def _test_database(self, results_text: scrolledtext.ScrolledText):
        """Test database connection and operations"""
        try:
            results_text.insert(tk.END, "Testing database connection...\n")
            
            # Test file operations
            test_data = {'test': 'data'}
            test_file = Path("data/test.json")
            
            # Test write
            with open(test_file, 'w') as f:
                json.dump(test_data, f)
            results_text.insert(tk.END, "✓ Write test successful\n")
            
            # Test read
            with open(test_file, 'r') as f:
                read_data = json.load(f)
            if read_data == test_data:
                results_text.insert(tk.END, "✓ Read test successful\n")
            else:
                results_text.insert(tk.END, "✗ Read test failed\n")
                
            # Clean up
            test_file.unlink()
            results_text.insert(tk.END, "✓ Cleanup successful\n\n")
            
            logger.debug("Database tests completed")
            
        except Exception as e:
            logger.error(f"Error testing database: {str(e)}")
            results_text.insert(tk.END, f"✗ Database test failed: {str(e)}\n\n")

    def _test_arduino(self, results_text: scrolledtext.ScrolledText):
        """Test Arduino connection"""
        try:
            results_text.insert(tk.END, "Testing Arduino connection...\n")
            
            if self.arduino_connected and self.serial_conn:
                # Test communication
                self.serial_conn.write(b"TEST\n")
                time.sleep(1)
                response = self.serial_conn.readline().decode('utf-8').strip()
                
                if response == "OK":
                    results_text.insert(tk.END, 
                        f"✓ Connected to Arduino on {self.arduino_port}\n")
                    results_text.insert(tk.END, 
                        "✓ Communication test successful\n\n")
                else:
                    results_text.insert(tk.END, 
                        "✗ Communication test failed\n\n")
            else:
                results_text.insert(tk.END, 
                    "✗ Arduino not connected\n\n")
                
            logger.debug("Arduino tests completed")
            
        except Exception as e:
            logger.error(f"Error testing Arduino: {str(e)}")
            results_text.insert(tk.END, f"✗ Arduino test failed: {str(e)}\n\n")

    def _test_sensors(self, results_text: scrolledtext.ScrolledText):
        """Test sensor readings"""
        try:
            results_text.insert(tk.END, "Testing sensors...\n")
            
            if self.arduino_connected and self.serial_conn:
                # Request sensor readings
                self.serial_conn.write(b"READ\n")
                time.sleep(1)
                response = self.serial_conn.readline().decode('utf-8').strip()
                
                try:
                    sensor_data = json.loads(response)
                    if self._validate_sensor_data(sensor_data):
                        results_text.insert(tk.END, "✓ pH Sensor: OK\n")
                        results_text.insert(tk.END, "✓ Chlorine Sensor: OK\n")
                        results_text.insert(tk.END, "✓ Temperature Sensor: OK\n\n")
                    else:
                        results_text.insert(tk.END, 
                            "✗ Invalid sensor readings\n\n")
                except json.JSONDecodeError:
                    results_text.insert(tk.END, 
                        "✗ Invalid sensor data format\n\n")
            else:
                results_text.insert(tk.END, 
                    "✗ Cannot test sensors - Arduino not connected\n\n")
                
            logger.debug("Sensor tests completed")
            
        except Exception as e:
            logger.error(f"Error testing sensors: {str(e)}")
            results_text.insert(tk.END, f"✗ Sensor test failed: {str(e)}\n\n")

    def _test_weather_api(self, results_text: scrolledtext.ScrolledText):
        """Test weather API connection and data retrieval"""
        try:
            results_text.insert(tk.END, "Testing weather API...\n")
            
            # Test API connection with sample ZIP code
            test_zip = "02888"  # Test ZIP code
            
            try:
                weather_data = self._fetch_weather_data(test_zip)
                
                if weather_data and 'current' in weather_data:
                    results_text.insert(tk.END, "✓ API connection successful\n")
                    
                    # Verify data structure
                    required_fields = ['temp', 'humidity', 'uv_index', 'wind_mph']
                    missing_fields = [
                        field for field in required_fields 
                        if field not in weather_data['current']
                    ]
                    
                    if not missing_fields:
                        results_text.insert(tk.END, "✓ Data structure valid\n")
                    else:
                        results_text.insert(tk.END, 
                            f"✗ Missing fields: {', '.join(missing_fields)}\n")
                    
                    # Verify forecast data
                    if 'forecast' in weather_data:
                        results_text.insert(tk.END, "✓ Forecast data available\n")
                    else:
                        results_text.insert(tk.END, "✗ No forecast data\n")
                        
                else:
                    results_text.insert(tk.END, "✗ Invalid weather data received\n")
                    
            except requests.RequestException as e:
                results_text.insert(tk.END, f"✗ API connection failed: {str(e)}\n")
                
            results_text.insert(tk.END, "\n")
            logger.debug("Weather API tests completed")
            
        except Exception as e:
            logger.error(f"Error testing weather API: {str(e)}")
            results_text.insert(tk.END, f"✗ Weather API test failed: {str(e)}\n\n")

    def _test_data_saving(self, results_text: scrolledtext.ScrolledText):
        """Test data saving and loading operations"""
        try:
            results_text.insert(tk.END, "Testing data operations...\n")
            
            # Test data structure
            test_data = {
                'customer_info': {
                    'name': 'Test Customer',
                    'address': 'Test Address',
                    'phone': '555-0123',
                    'pool_type': 'Chlorine'
                },
                'chemistry_data': {
                    'readings': [{
                        'timestamp': datetime.now().isoformat(),
                        'values': {
                            'pH': 7.2,
                            'Chlorine': 3.0,
                            'Alkalinity': 100
                        }
                    }]
                }
            }
            
            # Test save operation
            test_file = Path("data/test_save.json")
            try:
                with open(test_file, 'w') as f:
                    json.dump(test_data, f)
                results_text.insert(tk.END, "✓ Save operation successful\n")
                
                # Test load operation
                with open(test_file, 'r') as f:
                    loaded_data = json.load(f)
                
                if loaded_data == test_data:
                    results_text.insert(tk.END, "✓ Load operation successful\n")
                else:
                    results_text.insert(tk.END, "✗ Data integrity check failed\n")
                    
                # Test backup operation
                backup_dir = Path("backup")
                backup_dir.mkdir(exist_ok=True)
                backup_file = backup_dir / f"test_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                
                shutil.copy2(test_file, backup_file)
                if backup_file.exists():
                    results_text.insert(tk.END, "✓ Backup operation successful\n")
                else:
                    results_text.insert(tk.END, "✗ Backup operation failed\n")
                    
                # Clean up test files
                test_file.unlink()
                backup_file.unlink()
                results_text.insert(tk.END, "✓ Cleanup successful\n")
                
            except Exception as e:
                results_text.insert(tk.END, f"✗ File operation failed: {str(e)}\n")
                
            results_text.insert(tk.END, "\n")
            logger.debug("Data saving tests completed")
            
        except Exception as e:
            logger.error(f"Error testing data saving: {str(e)}")
            results_text.insert(tk.END, f"✗ Data saving test failed: {str(e)}\n\n")

    def _save_test_results(self, results: dict):
        """Save diagnostic test results"""
        try:
            # Create test results directory
            results_dir = Path("data/test_results")
            results_dir.mkdir(parents=True, exist_ok=True)
            
            # Create results file with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            results_file = results_dir / f"diagnostic_results_{timestamp}.json"
            
            # Add metadata
            results['metadata'] = {
                'timestamp': datetime.now().isoformat(),
                'app_version': '2.0',
                'system_info': {
                    'platform': platform.system(),
                    'python_version': platform.python_version(),
                    'processor': platform.processor()
                }
            }
            
            # Save results
            with open(results_file, 'w') as f:
                json.dump(results, f, indent=4)
                
            logger.info(f"Test results saved to {results_file}")
            
            # Clean up old test results
            self._cleanup_old_test_results(results_dir)
            
        except Exception as e:
            logger.error(f"Error saving test results: {str(e)}")
            raise

    def _cleanup_old_test_results(self, results_dir: Path):
        """Remove old test results"""
        try:
            # Get all test result files
            result_files = list(results_dir.glob("diagnostic_results_*.json"))
            
            # Sort by modification time
            result_files.sort(key=lambda x: x.stat().st_mtime)
            
            # Keep only last 10 test results
            while len(result_files) > 10:
                oldest = result_files.pop(0)
                oldest.unlink()
                logger.debug(f"Removed old test result: {oldest}")
                
        except Exception as e:
            logger.error(f"Error cleaning up test results: {str(e)}")
            raise

    def _show_help(self):
        """Show application help documentation"""
        try:
            # Create help dialog
            dialog = tk.Toplevel(self)
            dialog.title("Pool Chemistry Monitor Help")
            dialog.geometry("800x600")
            dialog.transient(self)
            dialog.grab_set()

            # Create notebook for help sections
            notebook = ttk.Notebook(dialog)
            notebook.pack(fill='both', expand=True, padx=10, pady=10)

            # Getting Started
            getting_started = ttk.Frame(notebook)
            notebook.add(getting_started, text="Getting Started")
            self._create_getting_started_help(getting_started)

            # Chemical Guide
            chemical_guide = ttk.Frame(notebook)
            notebook.add(chemical_guide, text="Chemical Guide")
            self._create_chemical_guide_help(chemical_guide)

            # Troubleshooting
            troubleshooting = ttk.Frame(notebook)
            notebook.add(troubleshooting, text="Troubleshooting")
            self._create_troubleshooting_help(troubleshooting)

            # About
            about = ttk.Frame(notebook)
            notebook.add(about, text="About")
            self._create_about_help(about)

            # Add close button
            ttk.Button(
                dialog,
                text="Close",
                command=dialog.destroy
            ).pack(pady=10)

            logger.info("Help documentation shown")

        except Exception as e:
            logger.error(f"Error showing help: {str(e)}")
            self._handle_error("Failed to show help documentation")

    def _create_getting_started_help(self, parent):
        """Create getting started help section"""
        try:
            # Create scrolled text widget
            text = scrolledtext.ScrolledText(
                parent,
                wrap=tk.WORD,
                width=70,
                height=30,
                font=("Helvetica", 10)
            )
            text.pack(fill='both', expand=True, padx=10, pady=10)

            # Add content
            text.insert(tk.END, "Getting Started Guide\n\n", "header")
            
            sections = [
                ("Initial Setup", [
                    "1. Enter your customer information",
                    "2. Select your pool type",
                    "3. Enter pool volume",
                    "4. Configure email and SMS settings (optional)"
                ]),
                ("Taking Readings", [
                    "1. Test pool water using test strips or kit",
                    "2. Enter readings in the input fields",
                    "3. Click 'Generate Recommendations' for analysis"
                ]),
                ("Using Arduino Sensors", [
                    "1. Connect Arduino device",
                    "2. Calibrate sensors if needed",
                    "3. Readings will update automatically"
                ]),
                ("Understanding Results", [
                    "• Green values are within ideal range",
                    "• Yellow values need attention",
                    "• Red values require immediate action",
                    "• Follow recommended chemical adjustments"
                ])
            ]

            for title, items in sections:
                text.insert(tk.END, f"\n{title}\n", "section")
                for item in items:
                    text.insert(tk.END, f"{item}\n", "normal")

            # Configure tags
            text.tag_configure("header", 
                font=("Helvetica", 12, "bold"),
                foreground=self.primary_color
            )
            text.tag_configure("section", 
                font=("Helvetica", 10, "bold"),
                foreground="#333333"
            )
            text.tag_configure("normal", 
                font=("Helvetica", 10),
                foreground="#333333"
            )

            # Make read-only
            text.configure(state='disabled')

        except Exception as e:
            logger.error(f"Error creating getting started help: {str(e)}")
            raise

    def _check_updates(self):
        """Check for application updates"""
        try:
            # Show checking dialog
            checking_dialog = tk.Toplevel(self)
            checking_dialog.title("Checking for Updates")
            checking_dialog.geometry("300x150")
            checking_dialog.transient(self)
            checking_dialog.grab_set()

            # Center dialog
            checking_dialog.geometry(f"+{self.winfo_x() + 50}+{self.winfo_y() + 50}")

            # Add progress message
            ttk.Label(
                checking_dialog,
                text="Checking for updates...",
                font=("Helvetica", 12)
            ).pack(pady=20)

            # Add progress bar
            progress = ttk.Progressbar(
                checking_dialog,
                mode='indeterminate'
            )
            progress.pack(fill='x', padx=20)
            progress.start()

            def check_update_thread():
                """Run update check in separate thread"""
                try:
                    # Simulate update check (replace with actual update check)
                    time.sleep(2)
                    
                    # Get current version
                    current_version = "2.0.0"
                    
                    # Get latest version (replace with actual version check)
                    latest_version = self._get_latest_version()

                    # Compare versions
                    if latest_version > current_version:
                        self._show_update_available(latest_version)
                    else:
                        messagebox.showinfo(
                            "Up to Date",
                            "You are running the latest version!"
                        )

                except Exception as e:
                    logger.error(f"Error checking updates: {str(e)}")
                    messagebox.showerror(
                        "Error",
                        "Failed to check for updates. Please try again later."
                    )
                finally:
                    checking_dialog.destroy()

            # Start check in separate thread
            threading.Thread(
                target=check_update_thread,
                daemon=True
            ).start()

            logger.info("Update check initiated")

        except Exception as e:
            logger.error(f"Error initiating update check: {str(e)}")
            self._handle_error("Failed to check for updates")

    def _get_latest_version(self) -> str:
        """Get latest version from server"""
        try:
            # Simulate version check (replace with actual server check)
            return "2.0.0"
        except Exception as e:
            logger.error(f"Error getting latest version: {str(e)}")
            raise

    def _show_update_available(self, latest_version: str):
        """Show update available dialog"""
        try:
            response = messagebox.askyesno(
                "Update Available",
                f"Version {latest_version} is available!\n\n"
                "Would you like to download the update?",
                icon='info'
            )
            
            if response:
                # Open download page (replace with actual URL)
                webbrowser.open("https://www.virtualcontrolllc.com/downloads")
                
        except Exception as e:
            logger.error(f"Error showing update dialog: {str(e)}")
            self._handle_error("Failed to show update information")

    def _create_chemical_guide_help(self, parent):
        """Create chemical guide help section"""
        try:
            # Create scrolled text widget
            text = scrolledtext.ScrolledText(
                parent,
                wrap=tk.WORD,
                width=70,
                height=30,
                font=("Helvetica", 10)
            )
            text.pack(fill='both', expand=True, padx=10, pady=10)

            # Add content
            text.insert(tk.END, "Pool Chemistry Guide\n\n", "header")

            chemicals = [
                ("pH Level (7.2 - 7.8)", [
                    "• Description: Measures water acidity/alkalinity",
                    "• Low pH: Corrosive, eye irritation",
                    "• High pH: Scale formation, reduced chlorine effectiveness",
                    "• Adjustment: Use pH Up or pH Down as needed",
                    "• Test Frequency: 2-3 times per week"
                ]),
                ("Free Chlorine (1.0 - 3.0 ppm)", [
                    "• Description: Primary sanitizer",
                    "• Low Chlorine: Algae growth, unsafe water",
                    "• High Chlorine: Eye/skin irritation, equipment damage",
                    "• Adjustment: Add chlorine tablets or shock",
                    "• Test Frequency: Daily"
                ]),
                ("Total Alkalinity (80 - 120 ppm)", [
                    "• Description: pH buffer/stabilizer",
                    "• Low Alkalinity: pH fluctuation, surface damage",
                    "• High Alkalinity: pH resistance, scaling",
                    "• Adjustment: Use alkalinity up/down",
                    "• Test Frequency: Weekly"
                ]),
                ("Calcium Hardness (200 - 400 ppm)", [
                    "• Description: Prevents water corrosiveness",
                    "• Low Hardness: Surface etching, equipment damage",
                    "• High Hardness: Scaling, cloudy water",
                    "• Adjustment: Add calcium chloride or partial drain",
                    "• Test Frequency: Monthly"
                ]),
                ("Cyanuric Acid (30 - 80 ppm)", [
                    "• Description: Chlorine stabilizer",
                    "• Low CYA: Rapid chlorine loss",
                    "• High CYA: Reduced chlorine effectiveness",
                    "• Adjustment: Add stabilizer or partial drain",
                    "• Test Frequency: Monthly"
                ])
            ]

            for title, items in chemicals:
                text.insert(tk.END, f"\n{title}\n", "section")
                for item in items:
                    text.insert(tk.END, f"{item}\n", "normal")

            # Add chemical interactions section
            text.insert(tk.END, "\nChemical Interactions\n", "section")
            interactions = [
                "• High pH reduces chlorine effectiveness",
                "• Low alkalinity causes pH bounce",
                "• High calcium + high pH causes scaling",
                "• High CYA reduces chlorine effectiveness",
                "• Temperature affects chemical reactions"
            ]
            for item in interactions:
                text.insert(tk.END, f"{item}\n", "normal")

            # Configure tags
            text.tag_configure("header", 
                font=("Helvetica", 12, "bold"),
                foreground=self.primary_color
            )
            text.tag_configure("section", 
                font=("Helvetica", 10, "bold"),
                foreground="#333333"
            )
            text.tag_configure("normal", 
                font=("Helvetica", 10),
                foreground="#333333"
            )

            # Make read-only
            text.configure(state='disabled')

        except Exception as e:
            logger.error(f"Error creating chemical guide: {str(e)}")
            raise

    def _create_troubleshooting_help(self, parent):
        """Create troubleshooting help section"""
        try:
            # Create scrolled text widget
            text = scrolledtext.ScrolledText(
                parent,
                wrap=tk.WORD,
                width=70,
                height=30,
                font=("Helvetica", 10)
            )
            text.pack(fill='both', expand=True, padx=10, pady=10)

            # Add content
            text.insert(tk.END, "Troubleshooting Guide\n\n", "header")

            issues = [
                ("Common Water Issues", [
                    ("Cloudy Water", [
                        "• Check filtration system",
                        "• Test pH and alkalinity",
                        "• Verify chlorine levels",
                        "• Check calcium hardness",
                        "• Consider using clarifier"
                    ]),
                    ("Green Water", [
                        "• Test and adjust chlorine",
                        "• Shock the pool",
                        "• Run filter continuously",
                        "• Brush pool walls",
                        "• Test phosphate levels"
                    ]),
                    ("Scale Formation", [
                        "• Check pH and alkalinity",
                        "• Test calcium hardness",
                        "• Verify water temperature",
                        "• Consider scale inhibitor",
                        "• Brush affected areas"
                    ])
                ]),
                ("Equipment Issues", [
                    ("Sensor Problems", [
                        "• Check connections",
                        "• Verify calibration",
                        "• Clean sensor probes",
                        "• Replace if necessary"
                    ]),
                    ("Reading Errors", [
                        "• Verify test kit freshness",
                        "• Follow testing procedure",
                        "• Check for interference",
                        "• Repeat test if unclear"
                    ])
                ]),
                ("Application Issues", [
                    ("Connection Problems", [
                        "• Check Arduino connection",
                        "• Verify COM port",
                        "• Restart application",
                        "• Update drivers"
                    ]),
                    ("Data Issues", [
                        "• Check file permissions",
                        "• Verify data format",
                        "• Backup data regularly",
                        "• Clear temporary files"
                    ])
                ])
            ]

            for section, categories in issues:
                text.insert(tk.END, f"\n{section}\n", "section")
                for category, items in categories:
                    text.insert(tk.END, f"\n{category}\n", "subsection")
                    for item in items:
                        text.insert(tk.END, f"{item}\n", "normal")

            # Configure tags
            text.tag_configure("header", 
                font=("Helvetica", 12, "bold"),
                foreground=self.primary_color
            )
            text.tag_configure("section", 
                font=("Helvetica", 11, "bold"),
                foreground="#333333"
            )
            text.tag_configure("subsection", 
                font=("Helvetica", 10, "bold"),
                foreground="#666666"
            )
            text.tag_configure("normal", 
                font=("Helvetica", 10),
                foreground="#333333"
            )

            # Make read-only
            text.configure(state='disabled')

        except Exception as e:
            logger.error(f"Error creating troubleshooting guide: {str(e)}")
            raise

    def _create_about_help(self, parent):
        """Create about help section"""
        try:
            # Create main frame
            main_frame = ttk.Frame(parent)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)

            # App title
            ttk.Label(
                main_frame,
                text="Deep Blue Pool Chemistry Monitor",
                font=("Helvetica", 16, "bold"),
                foreground=self.primary_color
            ).pack(pady=(0, 10))

            # Version info
            ttk.Label(
                main_frame,
                text="Version 2.0.0",
                font=("Helvetica", 12)
            ).pack()

            # Copyright
            ttk.Label(
                main_frame,
                text="© 2025 Virtual Control LLC\nAll rights reserved.",
                font=("Helvetica", 10)
            ).pack(pady=10)

            # Description
            description = ttk.Label(
                main_frame,
                text=(
                    "Professional pool chemistry management software\n"
                    "for monitoring, analysis, and recommendations.\n\n"
                    "Features:\n"
                    "• Real-time chemical monitoring\n"
                    "• Weather impact analysis\n"
                    "• Automated recommendations\n"
                    "• Test strip visualization\n"
                    "• Trend analysis and graphs\n"
                    "• PDF/Excel report generation\n"
                    "• Email and SMS notifications\n"
                    "• Arduino sensor integration"
                ),
                font=("Helvetica", 10),
                justify="center"
            )
            description.pack(pady=20)

            # Technical details
            tech_frame = ttk.LabelFrame(
                main_frame,
                text="Technical Information",
                padding=10
            )
            tech_frame.pack(fill='x', pady=(0, 20))

            tech_info = [
                f"Python Version: {platform.python_version()}",
                f"Operating System: {platform.system()} {platform.release()}",
                f"Processor: {platform.processor()}",
                f"Machine: {platform.machine()}"
            ]

            for info in tech_info:
                ttk.Label(
                    tech_frame,
                    text=info,
                    font=("Helvetica", 9)
                ).pack(anchor='w')

            # Contact information
            contact_frame = ttk.LabelFrame(
                main_frame,
                text="Contact Information",
                padding=10
            )
            contact_frame.pack(fill='x')

            ttk.Label(
                contact_frame,
                text="Email: mhayes@virtualcontrolllc.com\nPhone: 401-854-6341",
                font=("Helvetica", 9),
                justify="center"
            ).pack()

            # Website link
            website_frame = ttk.Frame(main_frame)
            website_frame.pack(pady=20)

            website_link = ttk.Label(
                website_frame,
                text="www.virtualcontrolllc.com",
                font=("Helvetica", 9, "underline"),
                foreground="blue",
                cursor="hand2"
            )
            website_link.pack()

            # Bind website click
            website_link.bind(
                "<Button-1>",
                lambda e: webbrowser.open_new("http://www.virtualcontrolllc.com")
            )

            logger.debug("About help section created successfully")

        except Exception as e:
            logger.error(f"Error creating about help: {str(e)}")
            raise

    def _handle_error(self, message: str, show_dialog: bool = True):
        """
        Handle application errors with enhanced logging and user feedback
        
        Args:
            message (str): Error message to display
            show_dialog (bool): Whether to show error dialog
        """
        try:
            # Get error details
            exc_type, exc_value, exc_traceback = sys.exc_info()
            if exc_traceback:
                # Format traceback
                tb_lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
                tb_text = ''.join(tb_lines)
                
                # Log full traceback
                logger.error(f"Error details:\n{tb_text}")
                
                # Save error to file
                self._save_error_log(message, tb_text)
            else:
                logger.error(message)

            if show_dialog:
                # Show error dialog
                dialog = tk.Toplevel(self)
                dialog.title("Error")
                dialog.geometry("400x300")
                dialog.transient(self)
                dialog.grab_set()

                # Error message
                ttk.Label(
                    dialog,
                    text="An error has occurred:",
                    font=("Helvetica", 12, "bold")
                ).pack(pady=(20, 10))

                ttk.Label(
                    dialog,
                    text=message,
                    wraplength=350
                ).pack(pady=(0, 20))

                # Error details in scrolled text
                if exc_traceback:
                    details_frame = ttk.LabelFrame(
                        dialog,
                        text="Technical Details",
                        padding=10
                    )
                    details_frame.pack(fill='both', expand=True, padx=10, pady=(0, 10))

                    details_text = scrolledtext.ScrolledText(
                        details_frame,
                        height=8,
                        width=40,
                        font=("Courier", 8)
                    )
                    details_text.pack(fill='both', expand=True)
                    details_text.insert('1.0', tb_text)
                    details_text.configure(state='disabled')

                # Buttons
                button_frame = ttk.Frame(dialog)
                button_frame.pack(fill='x', padx=10, pady=10)

                ttk.Button(
                    button_frame,
                    text="OK",
                    command=dialog.destroy
                ).pack(side='right', padx=5)

                if exc_traceback:
                    ttk.Button(
                        button_frame,
                        text="Copy Details",
                        command=lambda: self._copy_error_details(tb_text)
                    ).pack(side='right', padx=5)

        except Exception as e:
            # If error handler fails, log and show basic message
            logger.critical(f"Error handler failed: {str(e)}")
            messagebox.showerror("Critical Error", 
                "An error occurred while handling another error.")

    def _save_error_log(self, message: str, traceback: str):
        """Save error details to log file"""
        try:
            # Create error logs directory
            log_dir = Path("logs")
            log_dir.mkdir(exist_ok=True)

            # Create error log file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_file = log_dir / f"error_{timestamp}.log"

            # Write error details
            with open(log_file, 'w') as f:
                f.write(f"Error: {message}\n")
                f.write(f"Timestamp: {datetime.now().isoformat()}\n")
                f.write(f"Version: 2.0.0\n")
                f.write(f"System: {platform.system()} {platform.release()}\n")
                f.write("\nTraceback:\n")
                f.write(traceback)

            logger.info(f"Error log saved to {log_file}")

        except Exception as e:
            logger.critical(f"Failed to save error log: {str(e)}")

    def _copy_error_details(self, details: str):
        """Copy error details to clipboard"""
        try:
            self.clipboard_clear()
            self.clipboard_append(details)
            messagebox.showinfo("Success", "Error details copied to clipboard.")
        except Exception as e:
            logger.error(f"Failed to copy error details: {str(e)}")
            messagebox.showerror("Error", "Failed to copy error details.")

    def _validate_chemical_input(self, param: str, value: float) -> Tuple[bool, str]:
        """
        Validate chemical parameter input with enhanced checks
        
        Args:
            param (str): Chemical parameter name
            value (float): Value to validate
            
        Returns:
            Tuple[bool, str]: (is_valid, error_message)
        """
        try:
            # Check for negative values
            if value < 0:
                return False, f"{param} cannot be negative"

            # Get ideal range
            ideal_min, ideal_max = self.tester.ideal_levels[param]
            
            # Define absolute limits (50% below min to 100% above max)
            absolute_min = ideal_min * 0.5
            absolute_max = ideal_max * 2.0

            # Validate based on parameter type
            if "pH" in param:
                if not 0 <= value <= 14:
                    return False, "pH must be between 0 and 14"
            elif "Chlorine" in param:
                if value > 10:
                    return False, "Chlorine reading too high"
            elif "Alkalinity" in param:
                if value > 300:
                    return False, "Alkalinity reading too high"
            elif "Hardness" in param:
                if value > 1000:
                    return False, "Hardness reading too high"

            # Check against absolute limits
            if value < absolute_min:
                return False, f"{param} reading too low"
            if value > absolute_max:
                return False, f"{param} reading too high"

            return True, ""

        except Exception as e:
            logger.error(f"Error validating chemical input: {str(e)}")
            return False, "Validation error occurred"

    def _format_chemical_value(self, value: float, param: str) -> str:
        """
        Format chemical value with appropriate precision and units
        
        Args:
            value (float): Value to format
            param (str): Chemical parameter name
            
        Returns:
            str: Formatted value with units
        """
        try:
            if "pH" in param:
                return f"{value:.1f}"
            elif any(x in param for x in ["Chlorine", "Bromine"]):
                return f"{value:.1f} ppm"
            elif "Temperature" in param:
                return f"{value:.1f}°F"
            else:
                return f"{int(value)} ppm"
        except Exception as e:
            logger.error(f"Error formatting chemical value: {str(e)}")
            return str(value)

    def _get_status_color(self, value: float, param: str) -> str:
        """
        Get status color based on chemical value
        
        Args:
            value (float): Chemical value
            param (str): Chemical parameter name
            
        Returns:
            str: Hex color code
        """
        try:
            ideal_min, ideal_max = self.tester.ideal_levels[param]
            
            if value < ideal_min:
                # Calculate how far below
                percent_below = (ideal_min - value) / ideal_min
                if percent_below > 0.25:
                    return "#F44336"  # Red
                return "#FF9800"  # Orange
            elif value > ideal_max:
                # Calculate how far above
                percent_above = (value - ideal_max) / ideal_max
                if percent_above > 0.25:
                    return "#F44336"  # Red
                return "#FF9800"  # Orange
            else:
                return "#4CAF50"  # Green
        except Exception as e:
            logger.error(f"Error getting status color: {str(e)}")
            return "#333333"  # Default gray

    def _validate_email_settings(self) -> Tuple[bool, str]:
        """
        Validate email configuration settings
        
        Returns:
            Tuple[bool, str]: (is_valid, error_message)
        """
        try:
            if not self.email_config:
                return False, "Email configuration not found"

            # Validate email address
            email = self.email_config.get('sender_email', '')
            if not self._validate_email(email):
                return False, "Invalid sender email address"

            # Validate app password
            password = self.email_config.get('app_password', '')
            if not password:
                return False, "App password is required"

            return True, ""

        except Exception as e:
            logger.error(f"Error validating email settings: {str(e)}")
            return False, "Validation error occurred"

    def _validate_sms_settings(self) -> Tuple[bool, str]:
        """
        Validate SMS configuration settings
        
        Returns:
            Tuple[bool, str]: (is_valid, error_message)
        """
        try:
            if not self.sms_config:
                return False, "SMS configuration not found"

            # Validate Twilio credentials
            if not self.sms_config.get('account_sid'):
                return False, "Twilio Account SID is required"
            if not self.sms_config.get('auth_token'):
                return False, "Twilio Auth Token is required"

            # Validate phone number
            phone = self.sms_config.get('twilio_number', '')
            if not self._validate_phone(phone):
                return False, "Invalid Twilio phone number"

            return True, ""

        except Exception as e:
            logger.error(f"Error validating SMS settings: {str(e)}")
            return False, "Validation error occurred"

    def _validate_data_file(self, file_path: Path) -> Tuple[bool, str]:
        """
        Validate data file structure and content
        
        Args:
            file_path (Path): Path to data file
            
        Returns:
            Tuple[bool, str]: (is_valid, error_message)
        """
        try:
            if not file_path.exists():
                return False, "File does not exist"

            with open(file_path, 'r') as f:
                data = json.load(f)

            # Check required sections
            required_sections = ['customer_info', 'chemistry_data']
            for section in required_sections:
                if section not in data:
                    return False, f"Missing required section: {section}"

            # Validate chemistry data structure
            if 'readings' not in data['chemistry_data']:
                return False, "Invalid chemistry data structure"

            # Validate readings
            for reading in data['chemistry_data']['readings']:
                if 'timestamp' not in reading or 'values' not in reading:
                    return False, "Invalid reading structure"

            return True, ""

        except json.JSONDecodeError:
            return False, "Invalid JSON format"
        except Exception as e:
            logger.error(f"Error validating data file: {str(e)}")
            return False, "Validation error occurred"

    def _manage_data_retention(self):
        """Manage data retention based on settings"""
        try:
            if not hasattr(self, 'chemistry_data'):
                return

            # Get retention period
            retention = self.settings.get('data_retention', '6M')
            cutoff_date = self._get_cutoff_date(retention)

            # Filter readings
            original_count = len(self.chemistry_data['readings'])
            self.chemistry_data['readings'] = [
                reading for reading in self.chemistry_data['readings']
                if datetime.fromisoformat(reading['timestamp']) > cutoff_date
            ]
            removed_count = original_count - len(self.chemistry_data['readings'])

            if removed_count > 0:
                logger.info(f"Removed {removed_count} old readings based on {retention} retention policy")

        except Exception as e:
            logger.error(f"Error managing data retention: {str(e)}")
            raise

    def _create_backup(self, backup_type: str = 'auto'):
        """
        Create backup of application data
        
        Args:
            backup_type (str): Type of backup ('auto' or 'manual')
        """
        try:
            # Get backup directory
            backup_dir = Path(self.settings.get('backup_path', './backup'))
            backup_dir.mkdir(parents=True, exist_ok=True)

            # Create timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Create backup files
            data_backup = backup_dir / f"pool_data_{backup_type}_{timestamp}.json"
            settings_backup = backup_dir / f"settings_{backup_type}_{timestamp}.json"

            # Backup data
            if hasattr(self, 'chemistry_data'):
                with open(data_backup, 'w') as f:
                    json.dump(self.chemistry_data, f, indent=4)

            # Backup settings
            settings_data = {
                'settings': self.settings,
                'graph_settings': self.graph_settings,
                'email_config': {
                    k: v for k, v in self.email_config.items()
                    if k != 'app_password'
                },
                'sms_config': {
                    k: v for k, v in self.sms_config.items()
                    if k != 'auth_token'
                },
                'alert_settings': self.alert_settings
            }
            
            with open(settings_backup, 'w') as f:
                json.dump(settings_data, f, indent=4)

            # Clean up old backups
            self._cleanup_old_backups(backup_dir, backup_type)

            logger.info(f"{backup_type.title()} backup created successfully")

        except Exception as e:
            logger.error(f"Error creating {backup_type} backup: {str(e)}")
            raise

    def _cleanup_old_backups(self, backup_dir: Path, backup_type: str):
        """
        Remove old backup files
        
        Args:
            backup_dir (Path): Backup directory path
            backup_type (str): Type of backup to clean up
        """
        try:
            # Get all backup files of specified type
            data_backups = list(backup_dir.glob(f"pool_data_{backup_type}_*.json"))
            settings_backups = list(backup_dir.glob(f"settings_{backup_type}_*.json"))

            # Sort by modification time
            data_backups.sort(key=lambda x: x.stat().st_mtime)
            settings_backups.sort(key=lambda x: x.stat().st_mtime)

            # Keep only recent backups
            max_backups = 5 if backup_type == 'auto' else 10
            
            while len(data_backups) > max_backups:
                oldest = data_backups.pop(0)
                oldest.unlink()
                logger.debug(f"Removed old data backup: {oldest}")

            while len(settings_backups) > max_backups:
                oldest = settings_backups.pop(0)
                oldest.unlink()
                logger.debug(f"Removed old settings backup: {oldest}")

        except Exception as e:
            logger.error(f"Error cleaning up old backups: {str(e)}")
            raise

    def _restore_backup(self, backup_file: Path) -> bool:
        """
        Restore data from backup file
        
        Args:
            backup_file (Path): Path to backup file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            if not backup_file.exists():
                raise FileNotFoundError("Backup file not found")

            # Validate backup file
            valid, error = self._validate_data_file(backup_file)
            if not valid:
                raise ValueError(f"Invalid backup file: {error}")

            # Create restore point before proceeding
            self._create_backup('restore')

            # Load backup data
            with open(backup_file, 'r') as f:
                backup_data = json.load(f)

            # Restore data
            if 'chemistry_data' in backup_data:
                self.chemistry_data = backup_data['chemistry_data']
            if 'customer_info' in backup_data:
                self.customer_info = backup_data['customer_info']

            # Update UI
            self._update_customer_info()
            self._update_combined_graph()

            logger.info(f"Backup restored successfully from {backup_file}")
            return True

        except Exception as e:
            logger.error(f"Error restoring backup: {str(e)}")
            return False

    def _update_customer_info(self):
        """Update UI with customer information"""
        try:
            if hasattr(self, 'customer_info'):
                # Update customer fields
                if hasattr(self, 'customer_name_entry'):
                    self.customer_name_entry.delete(0, tk.END)
                    self.customer_name_entry.insert(
                        0, 
                        self.customer_info.get('name', '')
                    )
                
                if hasattr(self, 'address_entry'):
                    self.address_entry.delete(0, tk.END)
                    self.address_entry.insert(
                        0, 
                        self.customer_info.get('address', '')
                    )
                
                if hasattr(self, 'phone_entry'):
                    self.phone_entry.delete(0, tk.END)
                    self.phone_entry.insert(
                        0, 
                        self.customer_info.get('phone', '')
                    )

                # Update pool type
                if hasattr(self, 'pool_type_var'):
                    self.pool_type_var.set(
                        self.customer_info.get('pool_type', 'Select Pool Type')
                    )

        except Exception as e:
            logger.error(f"Error updating customer info: {str(e)}")
            raise

    def _import_data(self):
        """Import data from external file"""
        try:
            # Get file path
            file_path = filedialog.askopenfilename(
                title="Import Data",
                filetypes=[
                    ("JSON files", "*.json"),
                    ("CSV files", "*.csv"),
                    ("Excel files", "*.xlsx")
                ]
            )
            
            if not file_path:
                return

            # Process based on file type
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.json':
                success = self._import_json(file_path)
            elif file_ext == '.csv':
                success = self._import_csv(file_path)
            elif file_ext == '.xlsx':
                success = self._import_excel(file_path)
            else:
                raise ValueError("Unsupported file type")

            if success:
                messagebox.showinfo("Success", "Data imported successfully!")
                self._update_combined_graph()
            else:
                messagebox.showerror("Error", "Failed to import data")

            logger.info(f"Data import completed: {file_path}")

        except Exception as e:
            logger.error(f"Error importing data: {str(e)}")
            self._handle_error("Failed to import data")

    def _import_json(self, file_path: str) -> bool:
        """
        Import data from JSON file
        
        Args:
            file_path (str): Path to JSON file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Validate file
            valid, error = self._validate_data_file(Path(file_path))
            if not valid:
                raise ValueError(error)

            # Create backup before import
            self._create_backup('import')

            # Load data
            with open(file_path, 'r') as f:
                imported_data = json.load(f)

            # Merge data
            if 'chemistry_data' in imported_data:
                if not hasattr(self, 'chemistry_data'):
                    self.chemistry_data = {'readings': []}
                
                # Add new readings
                existing_timestamps = {
                    r['timestamp'] for r in self.chemistry_data['readings']
                }
                
                new_readings = [
                    r for r in imported_data['chemistry_data']['readings']
                    if r['timestamp'] not in existing_timestamps
                ]
                
                self.chemistry_data['readings'].extend(new_readings)

            # Sort readings by timestamp
            self.chemistry_data['readings'].sort(
                key=lambda x: x['timestamp']
            )

            return True

        except Exception as e:
            logger.error(f"Error importing JSON: {str(e)}")
            return False

    def _import_csv(self, file_path: str) -> bool:
        """
        Import data from CSV file
        
        Args:
            file_path (str): Path to CSV file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Create backup before import
            self._create_backup('import')

            # Read CSV file
            readings = []
            with open(file_path, 'r') as f:
                reader = csv.DictReader(f)
                
                for row in reader:
                    try:
                        # Parse timestamp
                        timestamp = datetime.strptime(
                            f"{row['Date']} {row['Time']}", 
                            '%Y-%m-%d %H:%M:%S'
                        ).isoformat()

                        # Parse values
                        values = {}
                        for param in self.tester.ideal_levels.keys():
                            if param in row and row[param]:
                                try:
                                    values[param] = float(row[param])
                                except ValueError:
                                    continue

                        if values:
                            readings.append({
                                'timestamp': timestamp,
                                'values': values
                            })

                    except (KeyError, ValueError):
                        continue

            # Add new readings
            if not hasattr(self, 'chemistry_data'):
                self.chemistry_data = {'readings': []}
                
            existing_timestamps = {
                r['timestamp'] for r in self.chemistry_data['readings']
            }
            
            new_readings = [
                r for r in readings
                if r['timestamp'] not in existing_timestamps
            ]
            
            self.chemistry_data['readings'].extend(new_readings)

            # Sort readings by timestamp
            self.chemistry_data['readings'].sort(
                key=lambda x: x['timestamp']
            )

            return True

        except Exception as e:
            logger.error(f"Error importing CSV: {str(e)}")
            return False

    def _import_excel(self, file_path: str) -> bool:
        """
        Import data from Excel file
        
        Args:
            file_path (str): Path to Excel file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Create backup before import
            self._create_backup('import')

            # Load workbook
            wb = load_workbook(file_path)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Process rows
            readings = []
            for row in ws.iter_rows(min_row=2):
                try:
                    # Get date and time
                    date_cell = row[headers.index('Date')].value
                    time_cell = row[headers.index('Time')].value
                    
                    if not date_cell or not time_cell:
                        continue

                    # Parse timestamp
                    timestamp = datetime.combine(
                        date_cell if isinstance(date_cell, date) 
                        else datetime.strptime(str(date_cell), '%Y-%m-%d').date(),
                        time_cell if isinstance(time_cell, time)
                        else datetime.strptime(str(time_cell), '%H:%M:%S').time()
                    ).isoformat()

                    # Parse values
                    values = {}
                    for param in self.tester.ideal_levels.keys():
                        try:
                            idx = headers.index(param)
                            if row[idx].value is not None:
                                values[param] = float(row[idx].value)
                        except (ValueError, IndexError):
                            continue

                    if values:
                        readings.append({
                            'timestamp': timestamp,
                            'values': values
                        })

                except (ValueError, IndexError):
                    continue

            # Add new readings
            if not hasattr(self, 'chemistry_data'):
                self.chemistry_data = {'readings': []}
                
            existing_timestamps = {
                r['timestamp'] for r in self.chemistry_data['readings']
            }
            
            new_readings = [
                r for r in readings
                if r['timestamp'] not in existing_timestamps
            ]
            
            self.chemistry_data['readings'].extend(new_readings)

            # Sort readings by timestamp
            self.chemistry_data['readings'].sort(
                key=lambda x: x['timestamp']
            )

            return True

        except Exception as e:
            logger.error(f"Error importing Excel: {str(e)}")
            return False

    def _export_data(self):
        """Export data to selected format"""
        try:
            # Create export dialog
            dialog = tk.Toplevel(self)
            dialog.title("Export Data")
            dialog.geometry("400x300")
            dialog.transient(self)
            dialog.grab_set()

            # Center dialog
            dialog.geometry(f"+{self.winfo_x() + 50}+{self.winfo_y() + 50}")

            # Export options frame
            options_frame = ttk.LabelFrame(dialog, text="Export Options", padding=10)
            options_frame.pack(fill='x', padx=20, pady=20)

            # File format selection
            format_var = tk.StringVar(value="excel")
            ttk.Radiobutton(
                options_frame,
                text="Excel (.xlsx)",
                variable=format_var,
                value="excel"
            ).pack(anchor='w', pady=5)
            
            ttk.Radiobutton(
                options_frame,
                text="CSV (.csv)",
                variable=format_var,
                value="csv"
            ).pack(anchor='w', pady=5)
            
            ttk.Radiobutton(
                options_frame,
                text="JSON (.json)",
                variable=format_var,
                value="json"
            ).pack(anchor='w', pady=5)

            # Date range frame
            date_frame = ttk.LabelFrame(dialog, text="Date Range", padding=10)
            date_frame.pack(fill='x', padx=20, pady=(0, 20))

            # Date range selection
            range_var = tk.StringVar(value="all")
            ttk.Radiobutton(
                date_frame,
                text="All Data",
                variable=range_var,
                value="all"
            ).pack(anchor='w', pady=5)
            
            ttk.Radiobutton(
                date_frame,
                text="Last Month",
                variable=range_var,
                value="1M"
            ).pack(anchor='w', pady=5)
            
            ttk.Radiobutton(
                date_frame,
                text="Last Week",
                variable=range_var,
                value="1W"
            ).pack(anchor='w', pady=5)

            def export():
                try:
                    format_type = format_var.get()
                    date_range = range_var.get()
                    
                    if format_type == "excel":
                        success = self._export_to_excel(date_range)
                    elif format_type == "csv":
                        success = self._export_to_csv(date_range)
                    else:  # json
                        success = self._export_to_json(date_range)
                        
                    if success:
                        messagebox.showinfo("Success", "Data exported successfully!")
                        dialog.destroy()
                    else:
                        messagebox.showerror("Error", "Failed to export data")
                        
                except Exception as e:
                    logger.error(f"Error in export function: {str(e)}")
                    messagebox.showerror("Error", "Failed to export data")

            # Buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill='x', padx=20, pady=10)
            
            ttk.Button(
                button_frame,
                text="Export",
                command=export,
                style="Primary.TButton"
            ).pack(side='right', padx=5)
            
            ttk.Button(
                button_frame,
                text="Cancel",
                command=dialog.destroy
            ).pack(side='right', padx=5)

        except Exception as e:
            logger.error(f"Error showing export dialog: {str(e)}")
            self._handle_error("Failed to show export dialog")

    def _filter_readings_by_date(self, date_range: str) -> List[dict]:
        """
        Filter readings by date range
        
        Args:
            date_range (str): Date range code ('all', '1M', '1W')
            
        Returns:
            List[dict]: Filtered readings
        """
        try:
            if not hasattr(self, 'chemistry_data') or not self.chemistry_data.get('readings'):
                return []

            if date_range == "all":
                return self.chemistry_data['readings']

            cutoff_date = self._get_cutoff_date(date_range)
            
            return [
                reading for reading in self.chemistry_data['readings']
                if datetime.fromisoformat(reading['timestamp']) > cutoff_date
            ]

        except Exception as e:
            logger.error(f"Error filtering readings: {str(e)}")
            return []

    def _export_to_excel(self, date_range: str) -> bool:
        """
        Export data to Excel file
        
        Args:
            date_range (str): Date range to export
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Get file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Export to Excel"
            )
            
            if not file_path:
                return False

            # Get filtered readings
            readings = self._filter_readings_by_date(date_range)
            if not readings:
                raise ValueError("No data to export")

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Chemical Readings"

            # Write headers
            headers = ["Date", "Time"] + list(self.tester.ideal_levels.keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="E3F2FD",
                    end_color="E3F2FD",
                    fill_type="solid"
                )

            # Write data
            for row, reading in enumerate(readings, 2):
                dt = datetime.fromisoformat(reading['timestamp'])
                
                # Write date and time
                ws.cell(row=row, column=1, value=dt.date())
                ws.cell(row=row, column=2, value=dt.time())
                
                # Write chemical values
                for col, param in enumerate(self.tester.ideal_levels.keys(), 3):
                    value = reading['values'].get(param)
                    if value is not None:
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.number_format = '0.00'

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Save workbook
            wb.save(file_path)
            logger.info(f"Data exported to Excel: {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            return False

    def _export_data(self):
        """Export data to selected format"""
        try:
            # Create export dialog
            dialog = tk.Toplevel(self)
            dialog.title("Export Data")
            dialog.geometry("400x300")
            dialog.transient(self)
            dialog.grab_set()

            # Center dialog
            dialog.geometry(f"+{self.winfo_x() + 50}+{self.winfo_y() + 50}")

            # Export options frame
            options_frame = ttk.LabelFrame(dialog, text="Export Options", padding=10)
            options_frame.pack(fill='x', padx=20, pady=20)

            # File format selection
            format_var = tk.StringVar(value="excel")
            ttk.Label(options_frame, text="File Format:").pack(anchor='w')
            
            ttk.Radiobutton(
                options_frame,
                text="Excel (.xlsx)",
                variable=format_var,
                value="excel"
            ).pack(anchor='w', pady=2)
            
            ttk.Radiobutton(
                options_frame,
                text="CSV (.csv)",
                variable=format_var,
                value="csv"
            ).pack(anchor='w', pady=2)
            
            ttk.Radiobutton(
                options_frame,
                text="PDF (.pdf)",
                variable=format_var,
                value="pdf"
            ).pack(anchor='w', pady=2)

            # Date range selection
            range_frame = ttk.LabelFrame(dialog, text="Date Range", padding=10)
            range_frame.pack(fill='x', padx=20, pady=(0, 20))

            range_var = tk.StringVar(value="all")
            ttk.Radiobutton(
                range_frame,
                text="All Data",
                variable=range_var,
                value="all"
            ).pack(anchor='w', pady=2)
            
            ttk.Radiobutton(
                range_frame,
                text="Last Month",
                variable=range_var,
                value="1M"
            ).pack(anchor='w', pady=2)
            
            ttk.Radiobutton(
                range_frame,
                text="Last Week",
                variable=range_var,
                value="1W"
            ).pack(anchor='w', pady=2)

            def export():
                try:
                    format_type = format_var.get()
                    date_range = range_var.get()
                    
                    if format_type == "excel":
                        success = self._export_to_excel(date_range)
                    elif format_type == "csv":
                        success = self._export_to_csv(date_range)
                    elif format_type == "pdf":
                        success = self._export_to_pdf(date_range)
                    
                    if success:
                        messagebox.showinfo("Success", "Data exported successfully!")
                        dialog.destroy()
                    else:
                        messagebox.showerror("Error", "Failed to export data")
                        
                except Exception as e:
                    logger.error(f"Error in export dialog: {str(e)}")
                    messagebox.showerror("Error", "Export failed")

            # Buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill='x', padx=20, pady=10)
            
            ttk.Button(
                button_frame,
                text="Export",
                command=export,
                style="Primary.TButton"
            ).pack(side='right', padx=5)
            
            ttk.Button(
                button_frame,
                text="Cancel",
                command=dialog.destroy
            ).pack(side='right', padx=5)

        except Exception as e:
            logger.error(f"Error showing export dialog: {str(e)}")
            self._handle_error("Failed to show export dialog")

    def _filter_readings_by_range(self, date_range: str) -> List[dict]:
        """
        Filter readings by date range
        
        Args:
            date_range (str): Date range code ('all', '1M', '1W')
            
        Returns:
            List[dict]: Filtered readings
        """
        try:
            if not hasattr(self, 'chemistry_data') or not self.chemistry_data.get('readings'):
                return []

            if date_range == "all":
                return self.chemistry_data['readings']

            cutoff_date = self._get_cutoff_date(date_range)
            
            return [
                reading for reading in self.chemistry_data['readings']
                if datetime.fromisoformat(reading['timestamp']) > cutoff_date
            ]

        except Exception as e:
            logger.error(f"Error filtering readings: {str(e)}")
            return []

    def _format_export_data(self, readings: List[dict]) -> List[dict]:
        """
        Format readings for export
        
        Args:
            readings (List[dict]): Raw readings
            
        Returns:
            List[dict]: Formatted readings
        """
        try:
            formatted_data = []
            
            for reading in readings:
                # Parse timestamp
                dt = datetime.fromisoformat(reading['timestamp'])
                
                # Create row
                row = {
                    'Date': dt.strftime('%Y-%m-%d'),
                    'Time': dt.strftime('%H:%M:%S')
                }
                
                # Add values
                for param, value in reading['values'].items():
                    row[param] = self._format_chemical_value(value, param)
                
                formatted_data.append(row)

            return formatted_data

        except Exception as e:
            logger.error(f"Error formatting export data: {str(e)}")
            return []

    def _get_export_headers(self) -> List[str]:
        """Get headers for export files"""
        try:
            headers = ['Date', 'Time']
            headers.extend(self.tester.ideal_levels.keys())
            return headers
        except Exception as e:
            logger.error(f"Error getting export headers: {str(e)}")
            return ['Date', 'Time']

    def _export_to_csv(self, date_range: str) -> bool:
        """
        Export data to CSV file
        
        Args:
            date_range (str): Date range to export
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Get file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")],
                title="Export to CSV"
            )
            
            if not file_path:
                return False

            # Get filtered readings
            readings = self._filter_readings_by_range(date_range)
            if not readings:
                raise ValueError("No data to export")

            # Format data
            formatted_data = self._format_export_data(readings)
            headers = self._get_export_headers()

            # Write CSV file
            with open(file_path, 'w', newline='') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                writer.writerows(formatted_data)

            logger.info(f"Data exported to CSV: {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            return False

    def _export_to_pdf(self, date_range: str) -> bool:
        """
        Export data to PDF file
        
        Args:
            date_range (str): Date range to export
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Get file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Export to PDF"
            )
            
            if not file_path:
                return False

            # Get filtered readings
            readings = self._filter_readings_by_range(date_range)
            if not readings:
                raise ValueError("No data to export")

            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )

            # Create story
            story = []
            styles = getSampleStyleSheet()
            
            # Add title
            title = Paragraph(
                "Pool Chemistry Report",
                styles['Title']
            )
            story.append(title)
            story.append(Spacer(1, 12))

            # Add date range
            if date_range == "all":
                range_text = "All Data"
            elif date_range == "1M":
                range_text = "Last Month"
            else:
                range_text = "Last Week"
                
            date_text = Paragraph(
                f"Date Range: {range_text}",
                styles['Normal']
            )
            story.append(date_text)
            story.append(Spacer(1, 12))

            # Add customer info if available
            if hasattr(self, 'customer_info'):
                customer_text = Paragraph(
                    f"Customer: {self.customer_info.get('name', 'N/A')}<br/>"
                    f"Address: {self.customer_info.get('address', 'N/A')}<br/>"
                    f"Phone: {self.customer_info.get('phone', 'N/A')}<br/>"
                    f"Pool Type: {self.customer_info.get('pool_type', 'N/A')}",
                    styles['Normal']
                )
                story.append(customer_text)
                story.append(Spacer(1, 12))

            # Create data table
            headers = self._get_export_headers()
            formatted_data = self._format_export_data(readings)
            
            # Convert data to table format
            table_data = [headers]
            for row in formatted_data:
                table_data.append([row.get(h, '') for h in headers])

            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOX', (0, 0), (-1, -1), 2, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            story.append(table)

            # Add ideal ranges
            story.append(Spacer(1, 20))
            story.append(Paragraph("Ideal Ranges:", styles['Heading2']))
            
            for param, (min_val, max_val) in self.tester.ideal_levels.items():
                story.append(Paragraph(
                    f"{param}: {min_val} - {max_val}",
                    styles['Normal']
                ))

            # Build PDF
            doc.build(story)

            logger.info(f"Data exported to PDF: {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to PDF: {str(e)}")
            return False

    def _export_graph(self, format_type: str = 'png') -> bool:
        """
        Export current graph to image file
        
        Args:
            format_type (str): Image format ('png' or 'pdf')
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Get file path
            file_types = [
                ("PNG files", "*.png") if format_type == 'png' 
                else ("PDF files", "*.pdf")
            ]
            file_path = filedialog.asksaveasfilename(
                defaultextension=f".{format_type}",
                filetypes=file_types,
                title="Export Graph"
            )
            
            if not file_path:
                return False

            # Save current figure
            if hasattr(self, 'fig'):
                self.fig.savefig(
                    file_path,
                    format=format_type,
                    dpi=300,
                    bbox_inches='tight'
                )
                logger.info(f"Graph exported to {format_type}: {file_path}")
                return True
            else:
                raise ValueError("No graph available to export")

        except Exception as e:
            logger.error(f"Error exporting graph: {str(e)}")
            return False

    def _generate_report(self) -> bool:
        """
        Generate comprehensive PDF report
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Get file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Generate Report"
            )
            
            if not file_path:
                return False

            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )

            # Create story
            story = []
            styles = getSampleStyleSheet()

            # Add title
            title = Paragraph(
                "Pool Chemistry Analysis Report",
                styles['Title']
            )
            story.append(title)
            story.append(Spacer(1, 12))

            # Add date
            date_text = Paragraph(
                f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles['Normal']
            )
            story.append(date_text)
            story.append(Spacer(1, 20))

            # Add customer information
            if hasattr(self, 'customer_info'):
                story.append(Paragraph("Customer Information", styles['Heading1']))
                customer_text = Paragraph(
                    f"Name: {self.customer_info.get('name', 'N/A')}<br/>"
                    f"Address: {self.customer_info.get('address', 'N/A')}<br/>"
                    f"Phone: {self.customer_info.get('phone', 'N/A')}<br/>"
                    f"Pool Type: {self.customer_info.get('pool_type', 'N/A')}",
                    styles['Normal']
                )
                story.append(customer_text)
                story.append(Spacer(1, 20))

            # Add current readings
            story.append(Paragraph("Current Readings", styles['Heading1']))
            readings_data = []
            headers = ['Parameter', 'Current Value', 'Ideal Range', 'Status']
            readings_data.append(headers)

            for param, entry in self.entries.items():
                try:
                    value = float(entry.get()) if entry.get() else None
                    if value is not None:
                        ideal_min, ideal_max = self.tester.ideal_levels[param]
                        status = "GOOD" if ideal_min <= value <= ideal_max else "NEEDS ATTENTION"
                        readings_data.append([
                            param,
                            self._format_chemical_value(value, param),
                            f"{ideal_min} - {ideal_max}",
                            status
                        ])
                except ValueError:
                    continue

            if len(readings_data) > 1:
                table = Table(readings_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BOX', (0, 0), (-1, -1), 2, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(table)
                story.append(Spacer(1, 20))

            # Add recommendations
            if 'recommendation_text' in self.__dict__:
                story.append(Paragraph("Recommendations", styles['Heading1']))
                recommendations = self.recommendation_text.get(1.0, tk.END).strip()
                story.append(Paragraph(recommendations, styles['Normal']))
                story.append(Spacer(1, 20))

            # Add graph if available
            if hasattr(self, 'fig'):
                story.append(Paragraph("Trend Analysis", styles['Heading1']))
                # Save graph to temporary file
                temp_graph = Path("temp_graph.png")
                self.fig.savefig(temp_graph, format='png', dpi=300, bbox_inches='tight')
                img = Image(str(temp_graph))
                img.drawHeight = 300
                img.drawWidth = 500
                story.append(img)
                temp_graph.unlink()  # Clean up temp file
                story.append(Spacer(1, 20))

            # Build PDF
            doc.build(story)
            logger.info(f"Report generated: {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            return False

    def _check_alerts(self):
        """Check chemical levels and send alerts if needed"""
        try:
            if not self.alert_settings.get('enabled', True):
                return

            # Check last alert time
            current_time = datetime.now()
            if hasattr(self, 'last_alert_time'):
                frequency = self.alert_settings.get('frequency', '24H')
                wait_hours = int(frequency.replace('H', ''))
                if (current_time - self.last_alert_time).total_seconds() < wait_hours * 3600:
                    return

            # Check chemical levels
            alerts = []
            for param, entry in self.entries.items():
                try:
                    value = float(entry.get()) if entry.get() else None
                    if value is not None:
                        ideal_min, ideal_max = self.tester.ideal_levels[param]
                        threshold = self.alert_settings.get('thresholds', {}).get(param, 10)
                        
                        # Calculate percentage deviation
                        if value < ideal_min:
                            deviation = ((ideal_min - value) / ideal_min) * 100
                            if deviation > threshold:
                                alerts.append(f"{param} is {deviation:.1f}% below ideal range")
                        elif value > ideal_max:
                            deviation = ((value - ideal_max) / ideal_max) * 100
                            if deviation > threshold:
                                alerts.append(f"{param} is {deviation:.1f}% above ideal range")
                except ValueError:
                    continue

            if alerts:
                self._send_alerts(alerts)
                self.last_alert_time = current_time

        except Exception as e:
            logger.error(f"Error checking alerts: {str(e)}")

    def _send_alerts(self, alerts: List[str]):
        """
        Send alerts via configured methods
        
        Args:
            alerts (List[str]): List of alert messages
        """
        try:
            # Prepare alert message
            message = "Pool Chemistry Alert:\n\n"
            message += "\n".join(f"• {alert}" for alert in alerts)
            message += "\n\nPlease check pool chemistry and take appropriate action."

            # Send email alert
            if self.alert_settings.get('email_enabled'):
                self._send_alert_email(message)

            # Send SMS alert
            if self.alert_settings.get('sms_enabled'):
                self._send_alert_sms(message)

            logger.info(f"Alerts sent: {', '.join(alerts)}")

        except Exception as e:
            logger.error(f"Error sending alerts: {str(e)}")

    def _send_alert_email(self, message: str):
        """
        Send alert via email
        
        Args:
            message (str): Alert message
        """
        try:
            # Validate email settings
            valid, error = self._validate_email_settings()
            if not valid:
                raise ValueError(error)

            # Create email message
            msg = MIMEMultipart()
            msg['From'] = self.email_config['sender_email']
            msg['To'] = self.customer_info.get('email', self.email_config['sender_email'])
            msg['Subject'] = "Pool Chemistry Alert"

            # Add HTML body
            message_formatted = message.replace('\n', '<br>')
            html = (
                f"""
                <html>
                    <head>
                        <style>
                            body {{ font-family: Arial, sans-serif; }}
                            .alert {{ color: red; font-weight: bold; }}
                            .message {{ margin: 20px 0; }}
                        </style>
                    </head>
                    <body>
                        <h2 class="alert">Pool Chemistry Alert</h2>
                        <div class="message">
                            {message_formatted}
                        </div>
                    </body>
                </html>
                    """
            )
            msg.attach(MIMEText(html, 'html'))

            # Send email
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(
                    self.email_config['sender_email'],
                    self.email_config['app_password']
                )
                server.send_message(msg)

            logger.info("Alert email sent successfully")

        except Exception as e:
            logger.error(f"Error sending alert email: {str(e)}")
            raise

    def _send_alert_sms(self, message: str):
        """
        Send alert via SMS
        
        Args:
            message (str): Alert message
        """
        try:
            # Validate SMS settings
            valid, error = self._validate_sms_settings()
            if not valid:
                raise ValueError(error)

            # Initialize Twilio client
            from twilio.rest import Client
            client = Client(
                self.sms_config['account_sid'],
                self.sms_config['auth_token']
            )

            # Send message
            message = client.messages.create(
                body=message[:160],  # SMS length limit
                from_=self.sms_config['twilio_number'],
                to=self.customer_info.get('phone')
            )

            logger.info(f"Alert SMS sent successfully: {message.sid}")

        except Exception as e:
            logger.error(f"Error sending alert SMS: {str(e)}")
            raise

    def _schedule_alert_check(self):
        """Schedule periodic alert checks"""
        try:
            self._check_alerts()
            # Schedule next check in 15 minutes
            self.after(900000, self._schedule_alert_check)
            
        except Exception as e:
            logger.error(f"Error scheduling alert check: {str(e)}")

    def _monitor_weather(self):
        """Monitor weather conditions and analyze impact"""
        try:
            if not hasattr(self, 'weather_data') or not self.weather_data:
                return

            current = self.weather_data['current']
            impacts = []
            severity = 0

            # Temperature impact
            if current['temp'] > 90:
                impacts.append({
                    'factor': 'Temperature',
                    'level': 'SEVERE',
                    'impact': 'Very high chlorine depletion rate',
                    'actions': [
                        'Test chemistry twice daily',
                        'Add extra chlorine as needed',
                        'Consider using chlorine stabilizer',
                        'Monitor pH more frequently'
                    ]
                })
                severity += 3
            elif current['temp'] > 85:
                impacts.append({
                    'factor': 'Temperature',
                    'level': 'HIGH',
                    'impact': 'Increased chlorine depletion',
                    'actions': [
                        'Test chemistry daily',
                        'Monitor chlorine levels closely'
                    ]
                })
                severity += 2

            # UV impact
            if current['uv_index'] > 10:
                impacts.append({
                    'factor': 'UV',
                    'level': 'SEVERE',
                    'impact': 'Extreme chlorine loss rate',
                    'actions': [
                        'Add extra chlorine stabilizer',
                        'Test frequently throughout day',
                        'Consider pool cover during peak hours'
                    ]
                })
                severity += 3
            elif current['uv_index'] > 8:
                impacts.append({
                    'factor': 'UV',
                    'level': 'HIGH',
                    'impact': 'Rapid chlorine loss',
                    'actions': [
                        'Add chlorine stabilizer',
                        'Test more frequently'
                    ]
                })
                severity += 2

            # Rain impact
            if current['rain_chance'] > 80:
                impacts.append({
                    'factor': 'Rain',
                    'level': 'HIGH',
                    'impact': 'Potential chemical dilution',
                    'actions': [
                        'Check chemistry after rain',
                        'Monitor water level',
                        'Be prepared to adjust chemicals'
                    ]
                })
                severity += 2
            elif current['rain_chance'] > 60:
                impacts.append({
                    'factor': 'Rain',
                    'level': 'MODERATE',
                    'impact': 'Possible chemical dilution',
                    'actions': [
                        'Monitor weather forecast',
                        'Be prepared to test after rain'
                    ]
                })
                severity += 1

            # Update weather impact display
            self._update_weather_impact(impacts, severity)

            # Schedule next check in 30 minutes
            self.after(1800000, self._monitor_weather)

            logger.debug("Weather monitoring completed")

        except Exception as e:
            logger.error(f"Error monitoring weather: {str(e)}")

    def _update_weather(self):
        """Update weather data from API"""
        try:
            # Validate widget existence
            if not hasattr(self, 'zip_entry'):
                logger.error("ZIP entry widget not initialized")
                return

            # Get and validate ZIP code
            zip_code = self.zip_entry.get().strip()
            if not zip_code:
                logger.debug("No ZIP code provided")
                return

            # Validate ZIP code format
            if not zip_code.isdigit() or len(zip_code) != 5:
                logger.warning(f"Invalid ZIP code format: {zip_code}")
                messagebox.showwarning(
                    "Invalid ZIP Code",
                    "Please enter a valid 5-digit ZIP code."
                )
                return

            # Fetch weather data
            self.weather_data = self._fetch_weather_data(zip_code)
            
            # Debug print weather data
            logger.debug(f"Weather Data: {self.weather_data}")
            
            # Update display if weather data was fetched successfully
            if self.weather_data:
                # Update current conditions
                current = self.weather_data['current']
                for key, value in current.items():
                    if key in self.current_labels:
                        if key == 'temp':
                            self.current_labels[key].config(
                                text=f"{value}°F"
                            )
                        elif key in ['humidity', 'rain_chance']:
                            self.current_labels[key].config(
                                text=f"{value}%"
                            )
                        else:
                            self.current_labels[key].config(
                                text=str(value)
                            )
                            
                # Update forecast display
                self._update_forecast_display()
                    
                # Update impact analysis
                self._update_impact_details(current)

                # Update last update time
                self.last_update = datetime.now()
                
                logger.info("Weather data updated successfully")
                
            else:
                logger.warning("No weather data received")
                self._show_weather_error()

        except requests.RequestException as e:
            logger.error(f"Weather API request failed: {str(e)}")
            self._handle_error(
                "Failed to connect to weather service. Please check your internet connection."
            )
            self._show_weather_error()
            
        except Exception as e:
            logger.error(f"Error updating weather: {str(e)}")
            self._handle_error("Failed to update weather data")
            self._show_weather_error()
            
        finally:
            # Schedule next update regardless of success/failure
            try:
                if hasattr(self, 'weather_update_id'):
                    self.after_cancel(self.weather_update_id)
                self.weather_update_id = self.after(1800000, self._update_weather)  # 30 minutes
            except Exception as e:
                logger.error(f"Error scheduling next weather update: {str(e)}")

    def _show_weather_error(self):
        """Display weather error message"""
        try:
            if hasattr(self, 'weather_frame'):
                # Clear existing content
                for widget in self.weather_frame.winfo_children():
                    widget.destroy()
                    
                # Add error message
                ttk.Label(
                    self.weather_frame,
                    text="Weather data unavailable",
                    foreground="red",
                    font=("Helvetica", 10, "bold")
                ).pack(pady=10)
                
                # Add retry button
                ttk.Button(
                    self.weather_frame,
                    text="Retry",
                    command=self._update_weather,
                    style="Small.TButton"
                ).pack(pady=5)
                
        except Exception as e:
            logger.error(f"Error showing weather error: {str(e)}")

    def _fetch_weather_data(self, zip_code: str) -> dict:
        """Fetch weather data from API"""
        try:
            # API endpoint (using weatherapi.com)
            base_url = "http://api.weatherapi.com/v1"
            api_key = self.weather_api_key
            
            # Get forecast data for 5 days
            forecast_url = (
                f"{base_url}/forecast.json"
                f"?key={api_key}"
                f"&q={zip_code}"
                f"&days=5"
                f"&aqi=no"
            )
            
            # Make request with timeout
            response = requests.get(forecast_url, timeout=10)
            response.raise_for_status()
            
            data = response.json()
            
            # Format weather data
            weather_data = {
                'current': {
                    'temp': data['current']['temp_f'],
                    'humidity': data['current']['humidity'],
                    'uv_index': data['current']['uv'],
                    'wind_mph': data['current']['wind_mph'],
                    'rain_chance': data['current']['precip_in'] * 100
                },
                'forecast': []
            }
            
            # Process forecast data
            for day in data['forecast']['forecastday']:
                forecast_day = {
                    'date': day['date'],
                    'temp': day['day']['avgtemp_f'],
                    'condition': day['day']['condition']['text'],
                    'rain_chance': day['day']['daily_chance_of_rain'],
                    'uv_index': day['day']['uv'],
                    'wind_mph': day['day']['maxwind_mph']
                }
                weather_data['forecast'].append(forecast_day)
                
            logger.info(f"Weather data fetched for ZIP: {zip_code}")
            return weather_data
            
        except Exception as e:
            logger.error(f"Error fetching weather data: {str(e)}")
            raise

    def _update_forecast_display(self):
        """Update the 5-day forecast display"""
        try:
            if not hasattr(self, 'weather_data') or not self.weather_data:
                return

            # Initialize forecast frames if they don't exist
            if not hasattr(self, 'forecast_frames'):
                self.forecast_frames = []
                forecast_container = ttk.Frame(self.weather_frame)
                forecast_container.pack(fill='x', pady=10)

                for _ in range(5):
                    frame = ttk.Frame(forecast_container)
                    frame.pack(side='left', expand=True, fill='x', padx=5)
                    
                    frames = {
                        'date': ttk.Label(frame),
                        'temp': ttk.Label(frame),
                        'condition': ttk.Label(frame),
                        'rain_chance': ttk.Label(frame),
                        'uv_index': ttk.Label(frame),
                        'impact': ttk.Label(frame)
                    }
                    
                    for label in frames.values():
                        label.pack(pady=2)
                    
                    self.forecast_frames.append(frames)

            # Update forecast data
            for i, day in enumerate(self.weather_data.get('forecast', [])[:5]):
                try:
                    # Parse date
                    date_obj = datetime.strptime(day['date'], '%Y-%m-%d')
                    frames = self.forecast_frames[i]

                    # Update labels
                    frames['date'].config(
                        text=date_obj.strftime('%A, %b %d'),
                        foreground=self.primary_color
                    )
                    frames['temp'].config(
                        text=f"Temp: {day['temp']}°F"
                    )
                    frames['condition'].config(
                        text=f"Condition: {day['condition']}"
                    )
                    frames['rain_chance'].config(
                        text=f"Rain: {day['rain_chance']}%"
                    )
                    frames['uv_index'].config(
                        text=f"UV: {day['uv_index']}"
                    )

                    # Calculate and display impact
                    impact_level = self._calculate_forecast_impact(day)
                    impact_color = self._get_impact_color(impact_level)
                    frames['impact'].config(
                        text=f"Impact: {impact_level}",
                        foreground=impact_color
                    )

                except Exception as e:
                    logger.error(f"Error updating forecast day {i}: {str(e)}")

            logger.debug("Forecast display updated successfully")

        except Exception as e:
            logger.error(f"Error updating forecast display: {str(e)}")
            self._show_weather_error()

    def _create_forecast_card(self, forecast_data: dict):
        """
        Create forecast card widget
        
        Args:
            forecast_data (dict): Forecast data for one day
        """
        try:
            # Create card frame
            card = ttk.Frame(self.forecast_frame, style="Card.TFrame")
            card.pack(fill='x', pady=(0, 5))

            # Add date
            date_obj = datetime.strptime(forecast_data['date'], '%Y-%m-%d')
            ttk.Label(
                card,
                text=date_obj.strftime('%A, %b %d'),
                font=("Helvetica", 10, "bold")
            ).pack(anchor='w')

            # Create conditions frame
            conditions = ttk.Frame(card)
            conditions.pack(fill='x', pady=5)

            # Add weather conditions
            ttk.Label(
                conditions,
                text=f"Temp: {forecast_data['temp']}°F",
                font=("Helvetica", 9)
            ).pack(side='left', padx=5)

            ttk.Label(
                conditions,
                text=f"UV: {forecast_data['uv_index']}",
                font=("Helvetica", 9)
            ).pack(side='left', padx=5)

            ttk.Label(
                conditions,
                text=f"Rain: {forecast_data['rain_chance']}%",
                font=("Helvetica", 9)
            ).pack(side='left', padx=5)

            # Add impact analysis
            impact_level = self._calculate_forecast_impact(forecast_data)
            impact_color = self._get_impact_color(impact_level)
            
            ttk.Label(
                card,
                text=f"Pool Impact: {impact_level}",
                font=("Helvetica", 9, "bold"),
                foreground=impact_color
            ).pack(anchor='w')

        except Exception as e:
            logger.error(f"Error creating forecast card: {str(e)}")

    def _calculate_forecast_impact(self, forecast_data: dict) -> str:
        """
        Calculate pool impact level from forecast data
        
        Args:
            forecast_data (dict): Forecast data for one day
            
        Returns:
            str: Impact level description
        """
        try:
            severity = 0

            # Temperature impact
            if forecast_data['temp'] > 90:
                severity += 3
            elif forecast_data['temp'] > 85:
                severity += 2
            elif forecast_data['temp'] > 80:
                severity += 1

            # UV impact
            if forecast_data['uv_index'] > 10:
                severity += 3
            elif forecast_data['uv_index'] > 8:
                severity += 2
            elif forecast_data['uv_index'] > 6:
                severity += 1

            # Rain impact
            if forecast_data['rain_chance'] > 80:
                severity += 3
            elif forecast_data['rain_chance'] > 60:
                severity += 2
            elif forecast_data['rain_chance'] > 40:
                severity += 1

            # Determine impact level
            if severity >= 7:
                return "SEVERE"
            elif severity >= 5:
                return "HIGH"
            elif severity >= 3:
                return "MODERATE"
            else:
                return "LOW"

        except Exception as e:
            logger.error(f"Error calculating forecast impact: {str(e)}")
            return "UNKNOWN"

    def _get_impact_color(self, impact_level: str) -> str:
        """
        Get color code for impact level
        
        Args:
            impact_level (str): Impact level description
            
        Returns:
            str: Hex color code
        """
        try:
            colors = {
                "SEVERE": "#F44336",  # Red
                "HIGH": "#FF9800",    # Orange
                "MODERATE": "#FFC107", # Yellow
                "LOW": "#4CAF50",     # Green
                "UNKNOWN": "#9E9E9E"  # Gray
            }
            return colors.get(impact_level, "#9E9E9E")

        except Exception as e:
            logger.error(f"Error getting impact color: {str(e)}")
            return "#9E9E9E"

    def _create_weather_visualization(self):
        """Create weather impact visualization"""
        try:
            if not hasattr(self, 'weather_canvas'):
                return

            # Clear canvas
            self.weather_canvas.delete("all")

            if not self.weather_data:
                self.weather_canvas.create_text(
                    200, 100,
                    text="Weather data unavailable",
                    font=("Helvetica", 10),
                    fill="gray"
                )
                return

            # Draw temperature gauge
            self._draw_temperature_gauge()

            # Draw UV index indicator
            self._draw_uv_indicator()

            # Draw rain chance meter
            self._draw_rain_meter()

            logger.debug("Weather visualization created")

        except Exception as e:
            logger.error(f"Error creating weather visualization: {str(e)}")

    def _draw_temperature_gauge(self):
        """Draw temperature gauge visualization"""
        try:
            # Gauge parameters
            center_x = 100
            center_y = 100
            radius = 60
            start_angle = 150
            end_angle = 30

            # Draw gauge background
            self._draw_arc(
                center_x, center_y, radius,
                start_angle, end_angle,
                width=20,
                fill="#E0E0E0"
            )

            # Get temperature and calculate angle
            temp = self.weather_data['current']['temp']
            temp_angle = self._map_value(
                temp, 60, 100,  # Temperature range
                start_angle, end_angle
            )

            # Draw temperature arc
            color = self._get_temperature_color(temp)
            self._draw_arc(
                center_x, center_y, radius,
                start_angle, temp_angle,
                width=20,
                fill=color
            )

            # Add temperature text
            self.weather_canvas.create_text(
                center_x, center_y - 20,
                text=f"{temp}°F",
                font=("Helvetica", 16, "bold"),
                fill=color
            )

            self.weather_canvas.create_text(
                center_x, center_y + 10,
                text="Temperature",
                font=("Helvetica", 10)
            )

        except Exception as e:
            logger.error(f"Error drawing temperature gauge: {str(e)}")

    def _draw_uv_indicator(self):
        """Draw UV index indicator visualization"""
        try:
            # Indicator parameters
            x = 250
            y = 100
            width = 120
            height = 20

            # Draw background
            self._draw_gradient_rect(
                x - width/2, y - height/2,
                x + width/2, y + height/2,
                ["#4CAF50", "#FFC107", "#FF9800", "#F44336"]
            )

            # Get UV index and calculate position
            uv = self.weather_data['current']['uv_index']
            uv_x = self._map_value(
                uv, 0, 12,  # UV index range
                x - width/2, x + width/2
            )

            # Draw indicator triangle
            self.weather_canvas.create_polygon(
                uv_x, y - height/2 - 10,
                uv_x - 5, y - height/2 - 5,
                uv_x + 5, y - height/2 - 5,
                fill="black"
            )

            # Add UV text
            self.weather_canvas.create_text(
                x, y + height,
                text=f"UV Index: {uv}",
                font=("Helvetica", 10)
            )

        except Exception as e:
            logger.error(f"Error drawing UV indicator: {str(e)}")

    def _draw_rain_meter(self):
        """Draw rain chance meter visualization"""
        try:
            # Meter parameters
            x = 400
            y = 100
            width = 30
            height = 120

            # Draw background
            self.weather_canvas.create_rectangle(
                x - width/2, y - height/2,
                x + width/2, y + height/2,
                fill="#E0E0E0",
                outline="#BDBDBD"
            )

            # Get rain chance and calculate height
            chance = self.weather_data['current']['rain_chance']
            fill_height = (chance / 100) * height

            # Draw fill level
            color = self._get_rain_color(chance)
            self.weather_canvas.create_rectangle(
                x - width/2, y + height/2 - fill_height,
                x + width/2, y + height/2,
                fill=color,
                outline=color
            )

            # Add percentage text
            self.weather_canvas.create_text(
                x, y - height/2 - 10,
                text=f"{chance}%",
                font=("Helvetica", 12, "bold"),
                fill=color
            )

            self.weather_canvas.create_text(
                x, y + height/2 + 15,
                text="Rain Chance",
                font=("Helvetica", 10)
            )

        except Exception as e:
            logger.error(f"Error drawing rain meter: {str(e)}")

    def _draw_arc(self, x: float, y: float, r: float, 
                  start: float, end: float, **kwargs):
        """
        Draw an arc on the canvas
        
        Args:
            x (float): Center X coordinate
            y (float): Center Y coordinate
            r (float): Radius
            start (float): Start angle in degrees
            end (float): End angle in degrees
            **kwargs: Additional canvas create_arc arguments
        """
        try:
            return self.weather_canvas.create_arc(
                x-r, y-r, x+r, y+r,
                start=start,
                extent=end-start,
                style=tk.ARC,
                **kwargs
            )
        except Exception as e:
            logger.error(f"Error drawing arc: {str(e)}")

    def _draw_gradient_rect(self, x1: float, y1: float, 
                           x2: float, y2: float, colors: List[str]):
        """
        Draw rectangle with gradient fill
        
        Args:
            x1 (float): Start X coordinate
            y1 (float): Start Y coordinate
            x2 (float): End X coordinate
            y2 (float): End Y coordinate
            colors (List[str]): List of gradient colors
        """
        try:
            width = x2 - x1
            segment_width = width / (len(colors) - 1)
            
            for i in range(len(colors) - 1):
                self.weather_canvas.create_rectangle(
                    x1 + i * segment_width,
                    y1,
                    x1 + (i + 1) * segment_width,
                    y2,
                    fill=colors[i],
                    outline=colors[i]
                )
        except Exception as e:
            logger.error(f"Error drawing gradient rectangle: {str(e)}")

    def _get_temperature_color(self, temp: float) -> str:
        """
        Get color based on temperature value
        
        Args:
            temp (float): Temperature in Fahrenheit
            
        Returns:
            str: Hex color code
        """
        try:
            if temp >= 95:
                return "#F44336"  # Red
            elif temp >= 90:
                return "#FF9800"  # Orange
            elif temp >= 85:
                return "#FFC107"  # Yellow
            elif temp >= 80:
                return "#4CAF50"  # Green
            else:
                return "#2196F3"  # Blue
        except Exception as e:
            logger.error(f"Error getting temperature color: {str(e)}")
            return "#9E9E9E"  # Gray

    def _get_rain_color(self, chance: float) -> str:
        """
        Get color based on rain chance
        
        Args:
            chance (float): Rain chance percentage
            
        Returns:
            str: Hex color code
        """
        try:
            if chance >= 80:
                return "#0D47A1"  # Dark Blue
            elif chance >= 60:
                return "#1976D2"  # Medium Blue
            elif chance >= 40:
                return "#42A5F5"  # Light Blue
            elif chance >= 20:
                return "#90CAF9"  # Very Light Blue
            else:
                return "#BBDEFB"  # Pale Blue
        except Exception as e:
            logger.error(f"Error getting rain color: {str(e)}")
            return "#9E9E9E"  # Gray

    def _map_value(self, value: float, in_min: float, in_max: float, 
                   out_min: float, out_max: float) -> float:
        """
        Map a value from one range to another
        
        Args:
            value (float): Value to map
            in_min (float): Input range minimum
            in_max (float): Input range maximum
            out_min (float): Output range minimum
            out_max (float): Output range maximum
            
        Returns:
            float: Mapped value
        """
        try:
            # Ensure value is within input range
            value = max(min(value, in_max), in_min)
            
            # Calculate mapping
            return (value - in_min) * (out_max - out_min) / (in_max - in_min) + out_min
        except Exception as e:
            logger.error(f"Error mapping value: {str(e)}")
            return out_min

    def _interpolate_color(self, color1: str, color2: str, factor: float) -> str:
        """
        Interpolate between two colors
        
        Args:
            color1 (str): First color in hex format
            color2 (str): Second color in hex format
            factor (float): Interpolation factor (0-1)
            
        Returns:
            str: Interpolated color in hex format
        """
        try:
            # Convert hex to RGB
            r1 = int(color1[1:3], 16)
            g1 = int(color1[3:5], 16)
            b1 = int(color1[5:7], 16)
            
            r2 = int(color2[1:3], 16)
            g2 = int(color2[3:5], 16)
            b2 = int(color2[5:7], 16)
            
            # Interpolate
            r = int(r1 + (r2 - r1) * factor)
            g = int(g1 + (g2 - g1) * factor)
            b = int(b1 + (b2 - b1) * factor)
            
            # Convert back to hex
            return f"#{r:02x}{g:02x}{b:02x}"
        except Exception as e:
            logger.error(f"Error interpolating color: {str(e)}")
            return "#000000"

    def _create_gradient_colors(self, start_color: str, end_color: str, 
                              steps: int) -> List[str]:
        """
        Create a list of gradient colors
        
        Args:
            start_color (str): Starting color in hex format
            end_color (str): Ending color in hex format
            steps (int): Number of gradient steps
            
        Returns:
            List[str]: List of gradient colors
        """
        try:
            colors = []
            for i in range(steps):
                factor = i / (steps - 1)
                color = self._interpolate_color(start_color, end_color, factor)
                colors.append(color)
            return colors
        except Exception as e:
            logger.error(f"Error creating gradient colors: {str(e)}")
            return [start_color, end_color]

    def _format_weather_value(self, value: float, type_: str) -> str:
        """
        Format weather value with appropriate units
        
        Args:
            value (float): Value to format
            type_ (str): Type of weather value
            
        Returns:
            str: Formatted value with units
        """
        try:
            if type_ == 'temp':
                return f"{value:.1f}°F"
            elif type_ == 'uv':
                return f"{value:.1f}"
            elif type_ == 'rain':
                return f"{value:.0f}%"
            elif type_ == 'wind':
                return f"{value:.1f} mph"
            else:
                return f"{value}"
        except Exception as e:
            logger.error(f"Error formatting weather value: {str(e)}")
            return str(value)

    def _create_maintenance_schedule(self):
        """Create and manage maintenance schedule"""
        try:
            # Create maintenance window
            schedule_window = tk.Toplevel(self)
            schedule_window.title("Maintenance Schedule")
            schedule_window.geometry("600x800")
            schedule_window.transient(self)
            schedule_window.grab_set()

            # Create notebook for different schedules
            notebook = ttk.Notebook(schedule_window)
            notebook.pack(fill='both', expand=True, padx=10, pady=10)

            # Daily tasks
            daily_frame = ttk.Frame(notebook, padding=10)
            notebook.add(daily_frame, text="Daily")
            self._create_task_list(daily_frame, self._get_daily_tasks())

            # Weekly tasks
            weekly_frame = ttk.Frame(notebook, padding=10)
            notebook.add(weekly_frame, text="Weekly")
            self._create_task_list(weekly_frame, self._get_weekly_tasks())

            # Monthly tasks
            monthly_frame = ttk.Frame(notebook, padding=10)
            notebook.add(monthly_frame, text="Monthly")
            self._create_task_list(monthly_frame, self._get_monthly_tasks())

            # Seasonal tasks
            seasonal_frame = ttk.Frame(notebook, padding=10)
            notebook.add(seasonal_frame, text="Seasonal")
            self._create_task_list(seasonal_frame, self._get_seasonal_tasks())

            # Add reminder settings
            self._create_reminder_settings(schedule_window)

            logger.debug("Maintenance schedule window created")

        except Exception as e:
            logger.error(f"Error creating maintenance schedule: {str(e)}")
            self._handle_error("Failed to create maintenance schedule")

    def _create_task_list(self, parent: ttk.Frame, tasks: List[dict]):
        """
        Create task list with checkboxes and reminders
        
        Args:
            parent (ttk.Frame): Parent frame
            tasks (List[dict]): List of task dictionaries
        """
        try:
            # Create scrolled frame
            canvas = tk.Canvas(parent)
            scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            # Add tasks
            for task in tasks:
                task_frame = ttk.Frame(scrollable_frame)
                task_frame.pack(fill='x', pady=2)

                var = tk.BooleanVar(value=False)
                cb = ttk.Checkbutton(
                    task_frame,
                    text=task['name'],
                    variable=var,
                    command=lambda t=task: self._on_task_toggle(t)
                )
                cb.pack(side='left')

                if task.get('critical', False):
                    ttk.Label(
                        task_frame,
                        text="CRITICAL",
                        foreground="red",
                        font=("Helvetica", 8, "bold")
                    ).pack(side='left', padx=5)

                # Add reminder button
                ttk.Button(
                    task_frame,
                    text="Set Reminder",
                    command=lambda t=task: self._set_task_reminder(t),
                    style="Small.TButton"
                ).pack(side='right')

            # Pack canvas and scrollbar
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

        except Exception as e:
            logger.error(f"Error creating task list: {str(e)}")
            raise

    def _get_daily_tasks(self) -> List[dict]:
        """Get daily maintenance tasks"""
        return [
            {
                'name': "Check chlorine levels",
                'critical': True,
                'reminder_default': "09:00"
            },
            {
                'name': "Check pH levels",
                'critical': True,
                'reminder_default': "09:00"
            },
            {
                'name': "Run pool pump",
                'critical': True,
                'reminder_default': "08:00"
            },
            {
                'name': "Empty skimmer baskets",
                'critical': False,
                'reminder_default': "17:00"
            },
            {
                'name': "Check water level",
                'critical': False,
                'reminder_default': "17:00"
            }
        ]

    def _get_weekly_tasks(self) -> List[dict]:
        """Get weekly maintenance tasks"""
        return [
            {
                'name': "Test all chemical levels",
                'critical': True,
                'reminder_default': "MON 09:00"
            },
            {
                'name': "Brush pool walls and floor",
                'critical': False,
                'reminder_default': "WED 17:00"
            },
            {
                'name': "Clean waterline",
                'critical': False,
                'reminder_default': "SAT 10:00"
            },
            {
                'name': "Backwash filter",
                'critical': True,
                'reminder_default': "SUN 16:00"
            },
            {
                'name': "Check equipment for leaks",
                'critical': True,
                'reminder_default': "SUN 16:00"
            }
        ]

    def _get_monthly_tasks(self) -> List[dict]:
        """Get monthly maintenance tasks"""
        return [
            {
                'name': "Deep clean filter",
                'critical': True,
                'reminder_default': "1ST 09:00"
            },
            {
                'name': "Check stabilizer levels",
                'critical': True,
                'reminder_default': "1ST 09:00"
            },
            {
                'name': "Clean deck and coping",
                'critical': False,
                'reminder_default': "15TH 10:00"
            },
            {
                'name': "Inspect pool surface",
                'critical': False,
                'reminder_default': "LAST 15:00"
            }
        ]

    def _get_seasonal_tasks(self) -> List[dict]:
        """Get seasonal maintenance tasks"""
        return [
            {
                'name': "Open pool (Spring)",
                'critical': True,
                'reminder_default': "MAR 15"
            },
            {
                'name': "Close pool (Fall)",
                'critical': True,
                'reminder_default': "SEP 15"
            },
            {
                'name': "Replace pump seals",
                'critical': False,
                'reminder_default': "APR 1"
            },
            {
                'name': "Replace filter media",
                'critical': False,
                'reminder_default': "APR 1"
            }
        ]

    def _create_reminder_settings(self, parent: tk.Toplevel):
        """
        Create reminder settings panel
        
        Args:
            parent (tk.Toplevel): Parent window
        """
        try:
            # Create settings frame
            settings_frame = ttk.LabelFrame(
                parent,
                text="Reminder Settings",
                padding=10
            )
            settings_frame.pack(fill='x', padx=10, pady=10)

            # Notification methods
            methods_frame = ttk.Frame(settings_frame)
            methods_frame.pack(fill='x', pady=5)

            ttk.Label(
                methods_frame,
                text="Notification Methods:"
            ).pack(side='left')

            self.email_reminder_var = tk.BooleanVar(
                value=self.settings.get('email_reminders', True)
            )
            ttk.Checkbutton(
                methods_frame,
                text="Email",
                variable=self.email_reminder_var
            ).pack(side='left', padx=10)

            self.sms_reminder_var = tk.BooleanVar(
                value=self.settings.get('sms_reminders', False)
            )
            ttk.Checkbutton(
                methods_frame,
                text="SMS",
                variable=self.sms_reminder_var
            ).pack(side='left', padx=10)

            # Reminder advance time
            advance_frame = ttk.Frame(settings_frame)
            advance_frame.pack(fill='x', pady=5)

            ttk.Label(
                advance_frame,
                text="Send reminders"
            ).pack(side='left')

            self.reminder_advance = tk.StringVar(
                value=self.settings.get('reminder_advance', '1 hour')
            )
            advance_combo = ttk.Combobox(
                advance_frame,
                textvariable=self.reminder_advance,
                values=['30 minutes', '1 hour', '2 hours', '1 day'],
                width=15,
                state='readonly'
            )
            advance_combo.pack(side='left', padx=5)

            ttk.Label(
                advance_frame,
                text="before task"
            ).pack(side='left', padx=5)

        except Exception as e:
            logger.error(f"Error creating reminder settings: {str(e)}")
            raise

    def _set_task_reminder(self, task: dict):
        """
        Set reminder for a maintenance task
        
        Args:
            task (dict): Task dictionary
        """
        try:
            # Create reminder dialog
            dialog = tk.Toplevel(self)
            dialog.title(f"Set Reminder - {task['name']}")
            dialog.geometry("400x300")
            dialog.transient(self)
            dialog.grab_set()

            # Reminder time
            time_frame = ttk.LabelFrame(
                dialog,
                text="Reminder Time",
                padding=10
            )
            time_frame.pack(fill='x', padx=10, pady=10)

            # Use default time
            default_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(
                time_frame,
                text=f"Use default ({task['reminder_default']})",
                variable=default_var,
                command=lambda: self._toggle_custom_time(
                    default_var,
                    custom_frame
                )
            ).pack(anchor='w')

            # Custom time frame
            custom_frame = ttk.Frame(time_frame)
            custom_frame.pack(fill='x', pady=5)
            custom_frame.pack_forget()  # Hidden by default

            hour_var = tk.StringVar(value="09")
            minute_var = tk.StringVar(value="00")
            
            ttk.Spinbox(
                custom_frame,
                from_=0,
                to=23,
                width=5,
                textvariable=hour_var,
                format="%02.0f"
            ).pack(side='left', padx=5)

            ttk.Label(
                custom_frame,
                text=":"
            ).pack(side='left')

            ttk.Spinbox(
                custom_frame,
                from_=0,
                to=59,
                width=5,
                textvariable=minute_var,
                format="%02.0f"
            ).pack(side='left', padx=5)

            # Repeat options
            repeat_frame = ttk.LabelFrame(
                dialog,
                text="Repeat",
                padding=10
            )
            repeat_frame.pack(fill='x', padx=10, pady=10)

            repeat_var = tk.StringVar(value="never")
            ttk.Radiobutton(
                repeat_frame,
                text="Never",
                value="never",
                variable=repeat_var
            ).pack(anchor='w')

            ttk.Radiobutton(
                repeat_frame,
                text="Daily",
                value="daily",
                variable=repeat_var
            ).pack(anchor='w')

            ttk.Radiobutton(
                repeat_frame,
                text="Weekly",
                value="weekly",
                variable=repeat_var
            ).pack(anchor='w')

            ttk.Radiobutton(
                repeat_frame,
                text="Monthly",
                value="monthly",
                variable=repeat_var
            ).pack(anchor='w')

            # Save button
            def save_reminder():
                try:
                    reminder_time = (
                        task['reminder_default'] if default_var.get()
                        else f"{hour_var.get()}:{minute_var.get()}"
                    )
                    self._add_reminder(
                        task,
                        reminder_time,
                        repeat_var.get()
                    )
                    dialog.destroy()
                    messagebox.showinfo(
                        "Success",
                        "Reminder set successfully!"
                    )
                except Exception as e:
                    logger.error(f"Error saving reminder: {str(e)}")
                    messagebox.showerror(
                        "Error",
                        "Failed to set reminder"
                    )

            ttk.Button(
                dialog,
                text="Save",
                command=save_reminder,
                style="Primary.TButton"
            ).pack(pady=10)

        except Exception as e:
            logger.error(f"Error setting task reminder: {str(e)}")
            self._handle_error("Failed to set reminder")

    def _toggle_custom_time(self, default_var: tk.BooleanVar, 
                           custom_frame: ttk.Frame):
        """
        Toggle custom time input visibility
        
        Args:
            default_var (tk.BooleanVar): Default time checkbox variable
            custom_frame (ttk.Frame): Custom time input frame
        """
        try:
            if default_var.get():
                custom_frame.pack_forget()
            else:
                custom_frame.pack(fill='x', pady=5)
        except Exception as e:
            logger.error(f"Error toggling custom time: {str(e)}")

    def _add_reminder(self, task: dict, time: str, repeat: str):
        """
        Add new reminder to schedule
        
        Args:
            task (dict): Task dictionary
            time (str): Reminder time
            repeat (str): Repeat frequency
        """
        try:
            # Create reminder object
            reminder = {
                'task': task['name'],
                'time': time,
                'repeat': repeat,
                'last_sent': None,
                'enabled': True
            }

            # Add to reminders list
            if not hasattr(self, 'reminders'):
                self.reminders = []
            self.reminders.append(reminder)

            # Save reminders
            self._save_reminders()

            # Schedule next check
            self._schedule_reminder_check()

            logger.info(f"Reminder added for {task['name']}")

        except Exception as e:
            logger.error(f"Error adding reminder: {str(e)}")
            raise

    def _save_reminders(self):
        """Save reminders to file"""
        try:
            # Create reminders directory
            reminders_dir = Path("data/reminders")
            reminders_dir.mkdir(parents=True, exist_ok=True)

            # Save to file
            with open(reminders_dir / "reminders.json", 'w') as f:
                json.dump(self.reminders, f, indent=4)

            logger.debug("Reminders saved successfully")

        except Exception as e:
            logger.error(f"Error saving reminders: {str(e)}")
            raise

    def _load_reminders(self):
        """Load reminders from file"""
        try:
            reminders_file = Path("data/reminders/reminders.json")
            if reminders_file.exists():
                with open(reminders_file, 'r') as f:
                    self.reminders = json.load(f)
                logger.debug("Reminders loaded successfully")
            else:
                self.reminders = []
                logger.debug("No reminders file found")

        except Exception as e:
            logger.error(f"Error loading reminders: {str(e)}")
            self.reminders = []

    def _schedule_reminder_check(self):
        """Schedule next reminder check"""
        try:
            # Cancel existing check if any
            if hasattr(self, '_reminder_check_id'):
                self.after_cancel(self._reminder_check_id)

            # Schedule new check in 1 minute
            self._reminder_check_id = self.after(60000, self._check_reminders)

            logger.debug("Reminder check scheduled")

        except Exception as e:
            logger.error(f"Error scheduling reminder check: {str(e)}")

    def _check_reminders(self):
        """Check and send due reminders"""
        try:
            current_time = datetime.now()

            for reminder in self.reminders:
                if not reminder['enabled']:
                    continue

                # Check if reminder is due
                if self._is_reminder_due(reminder, current_time):
                    # Send notification
                    self._send_reminder_notification(reminder)
                    
                    # Update last sent time
                    reminder['last_sent'] = current_time.isoformat()

            # Save updated reminders
            self._save_reminders()

            # Schedule next check
            self._schedule_reminder_check()

            logger.debug("Reminders checked successfully")

        except Exception as e:
            logger.error(f"Error checking reminders: {str(e)}")
            self._schedule_reminder_check()  # Ensure next check is scheduled

    def _is_reminder_due(self, reminder: dict, current_time: datetime) -> bool:
        """
        Check if reminder is due
        
        Args:
            reminder (dict): Reminder dictionary
            current_time (datetime): Current time
            
        Returns:
            bool: True if reminder is due
        """
        try:
            # Get reminder time
            time_parts = reminder['time'].split(':')
            reminder_time = current_time.replace(
                hour=int(time_parts[0]),
                minute=int(time_parts[1]),
                second=0,
                microsecond=0
            )

            # Check last sent time
            if reminder['last_sent']:
                last_sent = datetime.fromisoformat(reminder['last_sent'])
                
                # Check repeat frequency
                if reminder['repeat'] == 'never':
                    return False
                elif reminder['repeat'] == 'daily':
                    if current_time.date() <= last_sent.date():
                        return False
                elif reminder['repeat'] == 'weekly':
                    if (current_time - last_sent).days < 7:
                        return False
                elif reminder['repeat'] == 'monthly':
                    if (current_time.year, current_time.month) <= (
                        last_sent.year, last_sent.month):
                        return False

            # Check if current time matches reminder time
            time_diff = abs((current_time - reminder_time).total_seconds())
            return time_diff < 60  # Within 1 minute

        except Exception as e:
            logger.error(f"Error checking if reminder is due: {str(e)}")
            return False

    def _send_reminder_notification(self, reminder: dict):
        """
        Send reminder notification
        
        Args:
            reminder (dict): Reminder dictionary
        """
        try:
            message = f"Maintenance Reminder: {reminder['task']}"

            # Send email notification
            if self.settings.get('email_reminders'):
                self._send_reminder_email(message)

            # Send SMS notification
            if self.settings.get('sms_reminders'):
                self._send_reminder_sms(message)

            logger.info(f"Reminder notification sent for {reminder['task']}")

        except Exception as e:
            logger.error(f"Error sending reminder notification: {str(e)}")

    def _send_reminder_email(self, message: str):
        """
        Send reminder via email
        
        Args:
            message (str): Reminder message
        """
        try:
            # Validate email settings
            valid, error = self._validate_email_settings()
            if not valid:
                raise ValueError(error)

            # Create email message
            msg = MIMEMultipart()
            msg['From'] = self.email_config['sender_email']
            msg['To'] = self.customer_info.get('email', self.email_config['sender_email'])
            msg['Subject'] = "Pool Maintenance Reminder"

            # Create HTML body
            html = f"""
            <html>
                <head>
                    <style>
                        body {{ font-family: Arial, sans-serif; }}
                        .reminder {{ 
                            background-color: #E3F2FD;
                            padding: 20px;
                            border-radius: 5px;
                            margin: 20px 0;
                        }}
                        .message {{ font-size: 16px; }}
                        .footer {{ color: #666; font-size: 12px; }}
                    </style>
                </head>
                <body>
                    <div class="reminder">
                        <h2>Pool Maintenance Reminder</h2>
                        <p class="message">{message}</p>
                    </div>
                    <p class="footer">
                        This is an automated reminder from your Pool Chemistry Monitor.
                    </p>
                </body>
            </html>
            """
            msg.attach(MIMEText(html, 'html'))

            # Send email
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(
                    self.email_config['sender_email'],
                    self.email_config['app_password']
                )
                server.send_message(msg)

            logger.info("Reminder email sent successfully")

        except Exception as e:
            logger.error(f"Error sending reminder email: {str(e)}")
            raise

    def _send_reminder_sms(self, message: str):
        """
        Send reminder via SMS
        
        Args:
            message (str): Reminder message
        """
        try:
            # Validate SMS settings
            valid, error = self._validate_sms_settings()
            if not valid:
                raise ValueError(error)

            # Initialize Twilio client
            from twilio.rest import Client
            client = Client(
                self.sms_config['account_sid'],
                self.sms_config['auth_token']
            )

            # Send message
            message = client.messages.create(
                body=message,
                from_=self.sms_config['twilio_number'],
                to=self.customer_info.get('phone')
            )

            logger.info(f"Reminder SMS sent successfully: {message.sid}")

        except Exception as e:
            logger.error(f"Error sending reminder SMS: {str(e)}")
            raise

    def _on_task_toggle(self, task: dict):
        """
        Handle task completion toggle
        
        Args:
            task (dict): Task dictionary
        """
        try:
            # Record task completion
            completion = {
                'task': task['name'],
                'timestamp': datetime.now().isoformat(),
                'completed': True
            }

            # Add to task history
            if not hasattr(self, 'task_history'):
                self.task_history = []
            self.task_history.append(completion)

            # Save task history
            self._save_task_history()

            # Update maintenance log
            self._update_maintenance_log(completion)

            logger.info(f"Task completed: {task['name']}")

        except Exception as e:
            logger.error(f"Error handling task toggle: {str(e)}")

    def _save_task_history(self):
        """Save task completion history"""
        try:
            # Create history directory
            history_dir = Path("data/history")
            history_dir.mkdir(parents=True, exist_ok=True)

            # Save to file
            with open(history_dir / "task_history.json", 'w') as f:
                json.dump(self.task_history, f, indent=4)

            logger.debug("Task history saved successfully")

        except Exception as e:
            logger.error(f"Error saving task history: {str(e)}")
            raise

    def _update_maintenance_log(self, completion: dict):
        """
        Update maintenance log with completed task
        
        Args:
            completion (dict): Task completion record
        """
        try:
            if not hasattr(self, 'maintenance_text'):
                return

            # Format entry
            timestamp = datetime.fromisoformat(completion['timestamp'])
            entry = (
                f"{timestamp.strftime('%Y-%m-%d %H:%M')} - "
                f"Completed: {completion['task']}\n"
            )

            # Add to log
            self.maintenance_text.configure(state='normal')
            self.maintenance_text.insert('1.0', entry)
            self.maintenance_text.configure(state='disabled')

            logger.debug("Maintenance log updated")

        except Exception as e:
            logger.error(f"Error updating maintenance log: {str(e)}")

    def _load_task_history(self):
        """Load task completion history"""
        try:
            history_file = Path("data/history/task_history.json")
            if history_file.exists():
                with open(history_file, 'r') as f:
                    self.task_history = json.load(f)
                logger.debug("Task history loaded successfully")
            else:
                self.task_history = []
                logger.debug("No task history file found")

        except Exception as e:
            logger.error(f"Error loading task history: {str(e)}")
            self.task_history = []

    def _create_maintenance_report(self):
        """Generate maintenance activity report"""
        try:
            # Get file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Save Maintenance Report"
            )
            
            if not file_path:
                return

            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )

            # Create story
            story = []
            styles = getSampleStyleSheet()

            # Add title
            title = Paragraph(
                "Pool Maintenance Report",
                styles['Title']
            )
            story.append(title)
            story.append(Spacer(1, 12))

            # Add date range
            date_text = Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles['Normal']
            )
            story.append(date_text)
            story.append(Spacer(1, 20))

            # Add task completion summary
            story.append(Paragraph("Task Completion Summary", styles['Heading1']))
            story.append(Spacer(1, 12))

            # Calculate completion statistics
            stats = self._calculate_completion_stats()
            
            # Create statistics table
            stat_data = [
                ['Category', 'Completed', 'Pending', 'Completion Rate'],
                ['Daily Tasks', stats['daily']['completed'], 
                 stats['daily']['pending'], f"{stats['daily']['rate']}%"],
                ['Weekly Tasks', stats['weekly']['completed'], 
                 stats['weekly']['pending'], f"{stats['weekly']['rate']}%"],
                ['Monthly Tasks', stats['monthly']['completed'], 
                 stats['monthly']['pending'], f"{stats['monthly']['rate']}%"]
            ]

            table = Table(stat_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 20))

            # Add recent activity
            story.append(Paragraph("Recent Activity", styles['Heading1']))
            story.append(Spacer(1, 12))

            recent_tasks = self._get_recent_tasks()
            if recent_tasks:
                for task in recent_tasks:
                    dt = datetime.fromisoformat(task['timestamp'])
                    task_text = Paragraph(
                        f"{dt.strftime('%Y-%m-%d %H:%M')} - {task['task']}",
                        styles['Normal']
                    )
                    story.append(task_text)
                    story.append(Spacer(1, 6))
            else:
                story.append(Paragraph("No recent activity", styles['Normal']))

            # Add maintenance trends
            story.append(Spacer(1, 20))
            story.append(Paragraph("Maintenance Trends", styles['Heading1']))
            story.append(Spacer(1, 12))

            # Create trends graph
            trends_path = self._create_trends_graph()
            if trends_path:
                img = Image(trends_path)
                img.drawHeight = 300
                img.drawWidth = 500
                story.append(img)
                trends_path.unlink()  # Clean up temp file

            # Build PDF
            doc.build(story)
            logger.info(f"Maintenance report generated: {file_path}")

            return True

        except Exception as e:
            logger.error(f"Error creating maintenance report: {str(e)}")
            return False

    def _calculate_completion_stats(self) -> dict:
        """
        Calculate task completion statistics
        
        Returns:
            dict: Statistics by task category
        """
        try:
            stats = {
                'daily': {'completed': 0, 'pending': 0, 'rate': 0},
                'weekly': {'completed': 0, 'pending': 0, 'rate': 0},
                'monthly': {'completed': 0, 'pending': 0, 'rate': 0}
            }

            # Get date ranges
            now = datetime.now()
            today = now.date()
            week_start = today - timedelta(days=today.weekday())
            month_start = today.replace(day=1)

            # Calculate statistics
            for task in self.task_history:
                task_date = datetime.fromisoformat(task['timestamp']).date()
                
                if task_date == today:
                    stats['daily']['completed'] += 1
                if week_start <= task_date <= today:
                    stats['weekly']['completed'] += 1
                if month_start <= task_date <= today:
                    stats['monthly']['completed'] += 1

            # Calculate pending and rates
            total_daily = len(self._get_daily_tasks())
            total_weekly = len(self._get_weekly_tasks())
            total_monthly = len(self._get_monthly_tasks())

            stats['daily']['pending'] = total_daily - stats['daily']['completed']
            stats['weekly']['pending'] = total_weekly - stats['weekly']['completed']
            stats['monthly']['pending'] = total_monthly - stats['monthly']['completed']

            stats['daily']['rate'] = round(
                (stats['daily']['completed'] / total_daily) * 100
            ) if total_daily > 0 else 0
            
            stats['weekly']['rate'] = round(
                (stats['weekly']['completed'] / total_weekly) * 100
            ) if total_weekly > 0 else 0
            
            stats['monthly']['rate'] = round(
                (stats['monthly']['completed'] / total_monthly) * 100
            ) if total_monthly > 0 else 0

            return stats

        except Exception as e:
            logger.error(f"Error calculating completion stats: {str(e)}")
            return {
                'daily': {'completed': 0, 'pending': 0, 'rate': 0},
                'weekly': {'completed': 0, 'pending': 0, 'rate': 0},
                'monthly': {'completed': 0, 'pending': 0, 'rate': 0}
            }

    def _get_recent_tasks(self, limit: int = 10) -> List[dict]:
        """
        Get most recent completed tasks
        
        Args:
            limit (int): Maximum number of tasks to return
            
        Returns:
            List[dict]: Recent task completions
        """
        try:
            if not hasattr(self, 'task_history'):
                return []

            # Sort by timestamp descending
            sorted_tasks = sorted(
                self.task_history,
                key=lambda x: x['timestamp'],
                reverse=True
            )

            return sorted_tasks[:limit]

        except Exception as e:
            logger.error(f"Error getting recent tasks: {str(e)}")
            return []

    def _create_trends_graph(self) -> Optional[Path]:
        """
        Create maintenance trends graph
        
        Returns:
            Optional[Path]: Path to generated graph image
        """
        try:
            if not hasattr(self, 'task_history') or not self.task_history:
                return None

            # Create figure
            plt.figure(figsize=(10, 6))

            # Get date range
            dates = [
                datetime.fromisoformat(task['timestamp']).date()
                for task in self.task_history
            ]
            min_date = min(dates)
            max_date = max(dates)
            date_range = (max_date - min_date).days + 1

            # Create date bins
            bins = {}
            current_date = min_date
            while current_date <= max_date:
                bins[current_date] = 0
                current_date += timedelta(days=1)

            # Count tasks per day
            for task in self.task_history:
                task_date = datetime.fromisoformat(task['timestamp']).date()
                bins[task_date] += 1

            # Create plot data
            x = list(bins.keys())
            y = list(bins.values())

            # Plot trend line
            plt.plot(x, y, 'b-', label='Daily Tasks')

            # Add moving average
            window = min(7, date_range)
            if window > 1:
                moving_avg = [
                    sum(y[max(0, i-window):i])/min(i, window)
                    for i in range(1, len(y)+1)
                ]
                plt.plot(x, moving_avg, 'r--', label='7-Day Average')

            # Customize plot
            plt.title('Maintenance Activity Trends')
            plt.xlabel('Date')
            plt.ylabel('Tasks Completed')
            plt.grid(True, linestyle='--', alpha=0.7)
            plt.legend()

            # Rotate x-axis labels
            plt.xticks(rotation=45)

            # Adjust layout
            plt.tight_layout()

            # Save to temporary file
            temp_path = Path("temp_trends.png")
            plt.savefig(temp_path)
            plt.close()

            return temp_path

        except Exception as e:
            logger.error(f"Error creating trends graph: {str(e)}")
            return None

    def _analyze_maintenance_patterns(self) -> dict:
        """
        Analyze maintenance patterns and identify trends
        
        Returns:
            dict: Analysis results
        """
        try:
            if not hasattr(self, 'task_history') or not self.task_history:
                return {
                    'status': 'No data available',
                    'patterns': [],
                    'recommendations': []
                }

            analysis = {
                'status': 'Analysis completed',
                'patterns': [],
                'recommendations': []
            }

            # Analyze daily patterns
            daily_completion_rates = self._analyze_daily_patterns()
            if daily_completion_rates:
                # Identify low completion days
                low_days = [
                    day for day, rate in daily_completion_rates.items()
                    if rate < 0.7  # 70% threshold
                ]
                if low_days:
                    analysis['patterns'].append({
                        'type': 'daily',
                        'issue': f"Low completion rates on: {', '.join(low_days)}"
                    })
                    analysis['recommendations'].append(
                        f"Schedule more time for maintenance on {', '.join(low_days)}"
                    )

            # Analyze weekly patterns
            weekly_stats = self._analyze_weekly_patterns()
            if weekly_stats['missed_weeks'] > 0:
                analysis['patterns'].append({
                    'type': 'weekly',
                    'issue': f"Missed {weekly_stats['missed_weeks']} weeks of maintenance"
                })
                analysis['recommendations'].append(
                    "Set up weekly maintenance schedule reminders"
                )

            # Analyze monthly patterns
            monthly_stats = self._analyze_monthly_patterns()
            if monthly_stats['missed_months'] > 0:
                analysis['patterns'].append({
                    'type': 'monthly',
                    'issue': f"Missed {monthly_stats['missed_months']} months of maintenance"
                })
                analysis['recommendations'].append(
                    "Schedule monthly maintenance reviews"
                )

            # Add general recommendations
            if not analysis['patterns']:
                analysis['status'] = "Maintenance schedule well maintained"
                analysis['recommendations'].append(
                    "Continue current maintenance practices"
                )

            return analysis

        except Exception as e:
            logger.error(f"Error analyzing maintenance patterns: {str(e)}")
            return {
                'status': 'Analysis failed',
                'patterns': [],
                'recommendations': ['Unable to analyze maintenance patterns']
            }

    def _analyze_daily_patterns(self) -> Dict[str, float]:
        """
        Analyze daily maintenance patterns
        
        Returns:
            Dict[str, float]: Completion rates by day of week
        """
        try:
            day_counts = {
                'Monday': {'total': 0, 'completed': 0},
                'Tuesday': {'total': 0, 'completed': 0},
                'Wednesday': {'total': 0, 'completed': 0},
                'Thursday': {'total': 0, 'completed': 0},
                'Friday': {'total': 0, 'completed': 0},
                'Saturday': {'total': 0, 'completed': 0},
                'Sunday': {'total': 0, 'completed': 0}
            }

            for task in self.task_history:
                task_date = datetime.fromisoformat(task['timestamp'])
                day_name = task_date.strftime('%A')
                day_counts[day_name]['completed'] += 1
                day_counts[day_name]['total'] += 1

            # Calculate completion rates
            completion_rates = {}
            for day, counts in day_counts.items():
                if counts['total'] > 0:
                    completion_rates[day] = counts['completed'] / counts['total']
                else:
                    completion_rates[day] = 0.0

            return completion_rates

        except Exception as e:
            logger.error(f"Error analyzing daily patterns: {str(e)}")
            return {}

    def _analyze_weekly_patterns(self) -> dict:
        """
        Analyze weekly maintenance patterns
        
        Returns:
            dict: Weekly maintenance statistics
        """
        try:
            stats = {
                'total_weeks': 0,
                'completed_weeks': 0,
                'missed_weeks': 0,
                'completion_rate': 0.0,
                'best_week': None,
                'worst_week': None
            }

            if not hasattr(self, 'task_history') or not self.task_history:
                return stats

            # Group tasks by week
            weekly_tasks = {}
            for task in self.task_history:
                task_date = datetime.fromisoformat(task['timestamp'])
                week_start = task_date.date() - timedelta(days=task_date.weekday())
                
                if week_start not in weekly_tasks:
                    weekly_tasks[week_start] = []
                weekly_tasks[week_start].append(task)

            # Calculate statistics
            total_weekly_tasks = len(self._get_weekly_tasks())
            stats['total_weeks'] = len(weekly_tasks)

            completion_rates = {}
            for week_start, tasks in weekly_tasks.items():
                completion_rate = len(tasks) / total_weekly_tasks
                completion_rates[week_start] = completion_rate
                
                if completion_rate >= 0.7:  # 70% threshold
                    stats['completed_weeks'] += 1
                else:
                    stats['missed_weeks'] += 1

            # Calculate overall completion rate
            if stats['total_weeks'] > 0:
                stats['completion_rate'] = (
                    stats['completed_weeks'] / stats['total_weeks']
                ) * 100

            # Find best and worst weeks
            if completion_rates:
                best_week = max(completion_rates.items(), key=lambda x: x[1])
                worst_week = min(completion_rates.items(), key=lambda x: x[1])
                
                stats['best_week'] = {
                    'date': best_week[0].isoformat(),
                    'rate': best_week[1] * 100
                }
                stats['worst_week'] = {
                    'date': worst_week[0].isoformat(),
                    'rate': worst_week[1] * 100
                }

            return stats

        except Exception as e:
            logger.error(f"Error analyzing weekly patterns: {str(e)}")
            return {
                'total_weeks': 0,
                'completed_weeks': 0,
                'missed_weeks': 0,
                'completion_rate': 0.0,
                'best_week': None,
                'worst_week': None
            }

    def _analyze_monthly_patterns(self) -> dict:
        """
        Analyze monthly maintenance patterns
        
        Returns:
            dict: Monthly maintenance statistics
        """
        try:
            stats = {
                'total_months': 0,
                'completed_months': 0,
                'missed_months': 0,
                'completion_rate': 0.0,
                'best_month': None,
                'worst_month': None
            }

            if not hasattr(self, 'task_history') or not self.task_history:
                return stats

            # Group tasks by month
            monthly_tasks = {}
            for task in self.task_history:
                task_date = datetime.fromisoformat(task['timestamp'])
                month_key = task_date.replace(day=1).date()
                
                if month_key not in monthly_tasks:
                    monthly_tasks[month_key] = []
                monthly_tasks[month_key].append(task)

            # Calculate statistics
            total_monthly_tasks = len(self._get_monthly_tasks())
            stats['total_months'] = len(monthly_tasks)

            completion_rates = {}
            for month_start, tasks in monthly_tasks.items():
                completion_rate = len(tasks) / total_monthly_tasks
                completion_rates[month_start] = completion_rate
                
                if completion_rate >= 0.7:  # 70% threshold
                    stats['completed_months'] += 1
                else:
                    stats['missed_months'] += 1

            # Calculate overall completion rate
            if stats['total_months'] > 0:
                stats['completion_rate'] = (
                    stats['completed_months'] / stats['total_months']
                ) * 100

            # Find best and worst months
            if completion_rates:
                best_month = max(completion_rates.items(), key=lambda x: x[1])
                worst_month = min(completion_rates.items(), key=lambda x: x[1])
                
                stats['best_month'] = {
                    'date': best_month[0].isoformat(),
                    'rate': best_month[1] * 100
                }
                stats['worst_month'] = {
                    'date': worst_month[0].isoformat(),
                    'rate': worst_month[1] * 100
                }

            return stats

        except Exception as e:
            logger.error(f"Error analyzing monthly patterns: {str(e)}")
            return {
                'total_months': 0,
                'completed_months': 0,
                'missed_months': 0,
                'completion_rate': 0.0,
                'best_month': None,
                'worst_month': None
            }

    def _generate_maintenance_summary(self) -> str:
        """
        Generate maintenance summary text
        
        Returns:
            str: Formatted maintenance summary
        """
        try:
            weekly_stats = self._analyze_weekly_patterns()
            monthly_stats = self._analyze_monthly_patterns()
            
            summary = [
                "Maintenance Summary",
                "===================",
                "",
                f"Weekly Completion Rate: {weekly_stats['completion_rate']:.1f}%",
                f"Monthly Completion Rate: {monthly_stats['completion_rate']:.1f}%",
                "",
                "Weekly Statistics:",
                f"- Completed Weeks: {weekly_stats['completed_weeks']}",
                f"- Missed Weeks: {weekly_stats['missed_weeks']}",
                "",
                "Monthly Statistics:",
                f"- Completed Months: {monthly_stats['completed_months']}",
                f"- Missed Months: {monthly_stats['missed_months']}"
            ]

            if weekly_stats['best_week']:
                summary.extend([
                    "",
                    "Best Performance:",
                    f"- Best Week: {weekly_stats['best_week']['date']} "
                    f"({weekly_stats['best_week']['rate']:.1f}%)",
                    f"- Best Month: {monthly_stats['best_month']['date']} "
                    f"({monthly_stats['best_month']['rate']:.1f}%)"
                ])

            return "\n".join(summary)

        except Exception as e:
            logger.error(f"Error generating maintenance summary: {str(e)}")
            return "Error generating maintenance summary"

    def _create_maintenance_visualization(self):
        """Create maintenance visualization dashboard"""
        try:
            # Create visualization window
            viz_window = tk.Toplevel(self)
            viz_window.title("Maintenance Visualization")
            viz_window.geometry("800x600")
            viz_window.transient(self)

            # Create notebook for different views
            notebook = ttk.Notebook(viz_window)
            notebook.pack(fill='both', expand=True, padx=10, pady=10)

            # Completion rates tab
            completion_frame = ttk.Frame(notebook)
            notebook.add(completion_frame, text="Completion Rates")
            self._create_completion_chart(completion_frame)

            # Trends tab
            trends_frame = ttk.Frame(notebook)
            notebook.add(trends_frame, text="Trends")
            self._create_trends_chart(trends_frame)

            # Calendar view tab
            calendar_frame = ttk.Frame(notebook)
            notebook.add(calendar_frame, text="Calendar")
            self._create_maintenance_calendar(calendar_frame)

            logger.debug("Maintenance visualization created")

        except Exception as e:
            logger.error(f"Error creating maintenance visualization: {str(e)}")
            self._handle_error("Failed to create visualization")

    def _create_completion_chart(self, parent: ttk.Frame):
        """
        Create completion rate chart
        
        Args:
            parent (ttk.Frame): Parent frame
        """
        try:
            # Create figure
            fig = Figure(figsize=(8, 6))
            ax = fig.add_subplot(111)

            # Get completion data
            weekly_stats = self._analyze_weekly_patterns()
            monthly_stats = self._analyze_monthly_patterns()

            # Create bar chart
            categories = ['Weekly', 'Monthly']
            rates = [
                weekly_stats['completion_rate'],
                monthly_stats['completion_rate']
            ]
            
            bars = ax.bar(categories, rates)

            # Customize bars
            for bar in bars:
                height = bar.get_height()
                ax.text(
                    bar.get_x() + bar.get_width()/2.,
                    height,
                    f'{height:.1f}%',
                    ha='center',
                    va='bottom'
                )

            # Customize chart
            ax.set_title('Maintenance Completion Rates')
            ax.set_ylim(0, 100)
            ax.grid(True, linestyle='--', alpha=0.7)

            # Add canvas
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

            # Add toolbar
            toolbar = NavigationToolbar2Tk(canvas, parent)
            toolbar.update()

        except Exception as e:
            logger.error(f"Error creating completion chart: {str(e)}")
            raise

    def _create_trends_chart(self, parent: ttk.Frame):
        """
        Create maintenance trends chart
        
        Args:
            parent (ttk.Frame): Parent frame
        """
        try:
            # Create figure
            fig = Figure(figsize=(8, 6))
            ax = fig.add_subplot(111)

            # Get task history
            if not hasattr(self, 'task_history') or not self.task_history:
                ax.text(
                    0.5, 0.5,
                    'No maintenance data available',
                    ha='center',
                    va='center'
                )
            else:
                # Group tasks by date
                task_counts = {}
                for task in self.task_history:
                    date = datetime.fromisoformat(task['timestamp']).date()
                    task_counts[date] = task_counts.get(date, 0) + 1

                # Create plot data
                dates = sorted(task_counts.keys())
                counts = [task_counts[date] for date in dates]

                # Plot trend line
                ax.plot(dates, counts, 'b-', label='Daily Tasks')

                # Add moving average
                window = min(7, len(counts))
                if window > 1:
                    moving_avg = [
                        sum(counts[max(0, i-window):i])/min(i, window)
                        for i in range(1, len(counts)+1)
                    ]
                    ax.plot(dates, moving_avg, 'r--', label='7-Day Average')

                # Customize chart
                ax.set_title('Maintenance Activity Trends')
                ax.set_xlabel('Date')
                ax.set_ylabel('Tasks Completed')
                ax.grid(True, linestyle='--', alpha=0.7)
                ax.legend()

                # Rotate x-axis labels
                plt.setp(ax.get_xticklabels(), rotation=45)

            # Adjust layout
            fig.tight_layout()

            # Add canvas
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

            # Add toolbar
            toolbar = NavigationToolbar2Tk(canvas, parent)
            toolbar.update()

        except Exception as e:
            logger.error(f"Error creating trends chart: {str(e)}")
            raise

    def _create_maintenance_calendar(self, parent: ttk.Frame):
        """
        Create maintenance calendar view
        
        Args:
            parent (ttk.Frame): Parent frame
        """
        try:
            # Create calendar frame
            cal_frame = ttk.Frame(parent)
            cal_frame.pack(fill='both', expand=True)

            # Get current date
            today = datetime.now()
            current_month = today.replace(day=1)

            # Create month navigation
            nav_frame = ttk.Frame(cal_frame)
            nav_frame.pack(fill='x', pady=5)

            def prev_month():
                nonlocal current_month
                current_month = current_month.replace(day=1) - timedelta(days=1)
                update_calendar()

            def next_month():
                nonlocal current_month
                current_month = current_month.replace(day=28) + timedelta(days=5)
                current_month = current_month.replace(day=1)
                update_calendar()

            ttk.Button(
                nav_frame,
                text="<",
                command=prev_month
            ).pack(side='left', padx=5)

            month_label = ttk.Label(
                nav_frame,
                text=current_month.strftime("%B %Y"),
                font=("Helvetica", 12, "bold")
            )
            month_label.pack(side='left', padx=20)

            ttk.Button(
                nav_frame,
                text=">",
                command=next_month
            ).pack(side='left', padx=5)

            # Create calendar grid
            cal_grid = ttk.Frame(cal_frame)
            cal_grid.pack(fill='both', expand=True, pady=5)

            # Configure grid
            for i in range(7):
                cal_grid.grid_columnconfigure(i, weight=1)
            for i in range(6):
                cal_grid.grid_rowconfigure(i+1, weight=1)

            # Add day headers
            days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            for i, day in enumerate(days):
                ttk.Label(
                    cal_grid,
                    text=day,
                    font=("Helvetica", 10, "bold")
                ).grid(row=0, column=i, pady=5)

            def update_calendar():
                # Update month label
                month_label.config(text=current_month.strftime("%B %Y"))

                # Clear existing calendar
                for widget in cal_grid.grid_slaves():
                    if int(widget.grid_info()['row']) > 0:
                        widget.destroy()

                # Get calendar dates
                cal = calendar.monthcalendar(
                    current_month.year,
                    current_month.month
                )

                # Add calendar days
                for week_num, week in enumerate(cal):
                    for day_num, day in enumerate(week):
                        if day != 0:
                            date = current_month.replace(day=day)
                            self._create_calendar_day(
                                cal_grid,
                                week_num + 1,
                                day_num,
                                date
                            )

            # Initial calendar update
            update_calendar()

        except Exception as e:
            logger.error(f"Error creating maintenance calendar: {str(e)}")
            raise

    def _create_maintenance_visualization(self):
        """Create maintenance visualization dashboard"""
        try:
            # Create visualization window
            viz_window = tk.Toplevel(self)
            viz_window.title("Maintenance Visualization")
            viz_window.geometry("800x600")
            viz_window.transient(self)

            # Create notebook for different views
            notebook = ttk.Notebook(viz_window)
            notebook.pack(fill='both', expand=True, padx=10, pady=10)

            # Completion rates tab
            completion_frame = ttk.Frame(notebook)
            notebook.add(completion_frame, text="Completion Rates")
            self._create_completion_chart(completion_frame)

            # Trends tab
            trends_frame = ttk.Frame(notebook)
            notebook.add(trends_frame, text="Trends")
            self._create_trends_chart(trends_frame)

            # Calendar view tab
            calendar_frame = ttk.Frame(notebook)
            notebook.add(calendar_frame, text="Calendar")
            self._create_calendar_view(calendar_frame)

            logger.debug("Maintenance visualization created")

        except Exception as e:
            logger.error(f"Error creating maintenance visualization: {str(e)}")
            self._handle_error("Failed to create visualization")

    def _create_completion_chart(self, parent: ttk.Frame):
        """
        Create completion rate chart
        
        Args:
            parent (ttk.Frame): Parent frame
        """
        try:
            # Create figure
            fig = Figure(figsize=(8, 6))
            ax = fig.add_subplot(111)

            # Get completion data
            weekly_stats = self._analyze_weekly_patterns()
            monthly_stats = self._analyze_monthly_patterns()

            # Create bar chart
            categories = ['Weekly', 'Monthly']
            rates = [
                weekly_stats['completion_rate'],
                monthly_stats['completion_rate']
            ]
            
            bars = ax.bar(categories, rates)

            # Customize chart
            ax.set_ylim(0, 100)
            ax.set_ylabel('Completion Rate (%)')
            ax.set_title('Maintenance Completion Rates')

            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax.text(
                    bar.get_x() + bar.get_width()/2.,
                    height,
                    f'{height:.1f}%',
                    ha='center',
                    va='bottom'
                )

            # Add color based on completion rate
            for bar, rate in zip(bars, rates):
                if rate >= 90:
                    bar.set_color('#4CAF50')  # Green
                elif rate >= 70:
                    bar.set_color('#FFC107')  # Yellow
                else:
                    bar.set_color('#F44336')  # Red

            # Create canvas
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

        except Exception as e:
            logger.error(f"Error creating completion chart: {str(e)}")
            raise

    def _create_trends_chart(self, parent: ttk.Frame):
        """
        Create trends chart
        
        Args:
            parent (ttk.Frame): Parent frame
        """
        try:
            # Create figure
            fig = Figure(figsize=(8, 6))
            ax = fig.add_subplot(111)

            # Get task history
            if not hasattr(self, 'task_history') or not self.task_history:
                ax.text(
                    0.5, 0.5,
                    'No maintenance data available',
                    ha='center',
                    va='center'
                )
            else:
                # Group tasks by date
                task_counts = {}
                for task in self.task_history:
                    date = datetime.fromisoformat(task['timestamp']).date()
                    task_counts[date] = task_counts.get(date, 0) + 1

                # Create line plot
                dates = sorted(task_counts.keys())
                counts = [task_counts[date] for date in dates]
                
                ax.plot(dates, counts, 'b-', label='Daily Tasks')

                # Add moving average
                window = min(7, len(counts))
                if window > 1:
                    moving_avg = [
                        sum(counts[max(0, i-window):i])/min(i, window)
                        for i in range(1, len(counts)+1)
                    ]
                    ax.plot(dates, moving_avg, 'r--', label='7-Day Average')

                # Customize chart
                ax.set_xlabel('Date')
                ax.set_ylabel('Tasks Completed')
                ax.set_title('Maintenance Activity Trends')
                ax.legend()

                # Rotate x-axis labels
                plt.setp(ax.get_xticklabels(), rotation=45)

            # Adjust layout
            fig.tight_layout()

            # Create canvas
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

        except Exception as e:
            logger.error(f"Error creating trends chart: {str(e)}")
            raise

    def _create_calendar_view(self, parent: ttk.Frame):
        """
        Create calendar view of maintenance activities
        
        Args:
            parent (ttk.Frame): Parent frame
        """
        try:
            # Create calendar frame
            cal_frame = ttk.Frame(parent)
            cal_frame.pack(fill='both', expand=True)

            # Get current date
            today = datetime.now()
            current_month = today.replace(day=1)

            # Create month navigation
            nav_frame = ttk.Frame(cal_frame)
            nav_frame.pack(fill='x', pady=5)

            def prev_month():
                nonlocal current_month
                current_month = current_month.replace(day=1) - timedelta(days=1)
                update_calendar()

            def next_month():
                nonlocal current_month
                current_month = current_month.replace(day=28) + timedelta(days=5)
                current_month = current_month.replace(day=1)
                update_calendar()

            ttk.Button(
                nav_frame,
                text="<",
                command=prev_month
            ).pack(side='left', padx=5)

            month_label = ttk.Label(
                nav_frame,
                text=current_month.strftime("%B %Y"),
                font=("Helvetica", 12, "bold")
            )
            month_label.pack(side='left', padx=20)

            ttk.Button(
                nav_frame,
                text=">",
                command=next_month
            ).pack(side='left', padx=5)

            # Create calendar grid
            cal_grid = ttk.Frame(cal_frame)
            cal_grid.pack(fill='both', expand=True, pady=10)

            # Configure grid
            for i in range(7):
                cal_grid.grid_columnconfigure(i, weight=1)
            for i in range(6):
                cal_grid.grid_rowconfigure(i+1, weight=1)

            # Add day headers
            days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            for i, day in enumerate(days):
                ttk.Label(
                    cal_grid,
                    text=day,
                    font=("Helvetica", 10, "bold")
                ).grid(row=0, column=i, pady=5)

            def update_calendar():
                # Update month label
                month_label.config(text=current_month.strftime("%B %Y"))

                # Clear existing calendar
                for widget in cal_grid.grid_slaves():
                    if int(widget.grid_info()['row']) > 0:
                        widget.destroy()

                # Get calendar dates
                cal = calendar.monthcalendar(
                    current_month.year,
                    current_month.month
                )

                # Add dates to grid
                for week_num, week in enumerate(cal):
                    for day_num, day in enumerate(week):
                        if day != 0:
                            date = current_month.replace(day=day)
                            
                            # Create date frame
                            date_frame = ttk.Frame(cal_grid)
                            date_frame.grid(
                                row=week_num+1,
                                column=day_num,
                                sticky='nsew',
                                padx=1,
                                pady=1
                            )

                            # Add date number
                            ttk.Label(
                                date_frame,
                                text=str(day),
                                font=("Helvetica", 9)
                            ).pack(anchor='nw')

                            # Add task indicators
                            task_count = self._get_tasks_for_date(date)
                            if task_count > 0:
                                ttk.Label(
                                    date_frame,
                                    text=f"{task_count} tasks",
                                    font=("Helvetica", 8),
                                    foreground="green"
                                ).pack()

            # Initial calendar update
            update_calendar()

        except Exception as e:
            logger.error(f"Error creating calendar view: {str(e)}")
            raise

    def _get_tasks_for_date(self, date: datetime) -> int:
        """
        Get number of tasks completed on a specific date
        
        Args:
            date (datetime): Date to check
            
        Returns:
            int: Number of tasks completed
        """
        try:
            if not hasattr(self, 'task_history'):
                return 0

            count = 0
            for task in self.task_history:
                task_date = datetime.fromisoformat(task['timestamp']).date()
                if task_date == date.date():
                    count += 1
            return count

        except Exception as e:
            logger.error(f"Error getting tasks for date: {str(e)}")
            return 0

    def _create_calendar_day(self, parent: ttk.Frame, row: int, col: int, date: datetime):
        """
        Create calendar day widget
        
        Args:
            parent (ttk.Frame): Parent frame
            row (int): Grid row
            col (int): Grid column
            date (datetime): Date to display
        """
        try:
            # Create day frame
            day_frame = ttk.Frame(parent, style="Calendar.TFrame")
            day_frame.grid(row=row, column=col, sticky='nsew', padx=1, pady=1)

            # Configure frame weight
            day_frame.grid_columnconfigure(0, weight=1)
            day_frame.grid_rowconfigure(1, weight=1)

            # Add date number
            date_label = ttk.Label(
                day_frame,
                text=str(date.day),
                style="CalendarDate.TLabel"
            )
            date_label.grid(row=0, column=0, sticky='nw', padx=2, pady=2)

            # Get tasks for this date
            tasks = self._get_date_tasks(date)
            
            if tasks:
                # Create task list
                task_frame = ttk.Frame(day_frame)
                task_frame.grid(row=1, column=0, sticky='nsew')
                
                for task in tasks[:3]:  # Show up to 3 tasks
                    task_label = ttk.Label(
                        task_frame,
                        text=f"• {task['name']}",
                        style="CalendarTask.TLabel"
                    )
                    task_label.pack(anchor='w', padx=2)

                # Show more indicator if needed
                if len(tasks) > 3:
                    ttk.Label(
                        task_frame,
                        text=f"+ {len(tasks) - 3} more",
                        style="CalendarMore.TLabel"
                    ).pack(anchor='w', padx=2)

            # Add hover binding for task details
            day_frame.bind(
                "<Enter>",
                lambda e: self._show_day_details(e, date, tasks)
            )
            day_frame.bind(
                "<Leave>",
                lambda e: self._hide_day_details(e)
            )

        except Exception as e:
            logger.error(f"Error creating calendar day: {str(e)}")
            raise

    def _get_date_tasks(self, date: datetime) -> List[dict]:
        """
        Get tasks for specific date
        
        Args:
            date (datetime): Date to check
            
        Returns:
            List[dict]: List of tasks
        """
        try:
            tasks = []
            if hasattr(self, 'task_history'):
                for task in self.task_history:
                    task_date = datetime.fromisoformat(task['timestamp']).date()
                    if task_date == date.date():
                        tasks.append(task)
            return tasks

        except Exception as e:
            logger.error(f"Error getting date tasks: {str(e)}")
            return []

    def _show_day_details(self, event: tk.Event, date: datetime, tasks: List[dict]):
        """
        Show day details tooltip
        
        Args:
            event (tk.Event): Mouse event
            date (datetime): Date to show
            tasks (List[dict]): Tasks to display
        """
        try:
            # Create tooltip window
            self.day_tooltip = tk.Toplevel(self)
            self.day_tooltip.wm_overrideredirect(True)
            
            # Position near mouse
            x = event.widget.winfo_rootx() + event.x + 15
            y = event.widget.winfo_rooty() + event.y + 10
            self.day_tooltip.wm_geometry(f"+{x}+{y}")

            # Create content
            frame = ttk.Frame(self.day_tooltip, style="Tooltip.TFrame")
            frame.pack(fill='both', expand=True)

            # Add date header
            ttk.Label(
                frame,
                text=date.strftime("%B %d, %Y"),
                style="TooltipHeader.TLabel"
            ).pack(padx=10, pady=(5, 0))

            if tasks:
                # Add task list
                for task in tasks:
                    task_frame = ttk.Frame(frame)
                    task_frame.pack(fill='x', padx=10, pady=2)

                    ttk.Label(
                        task_frame,
                        text=f"• {task['name']}",
                        style="TooltipTask.TLabel"
                    ).pack(side='left')

                    time = datetime.fromisoformat(task['timestamp']).strftime("%H:%M")
                    ttk.Label(
                        task_frame,
                        text=time,
                        style="TooltipTime.TLabel"
                    ).pack(side='right')
            else:
                ttk.Label(
                    frame,
                    text="No tasks scheduled",
                    style="TooltipEmpty.TLabel"
                ).pack(padx=10, pady=5)

        except Exception as e:
            logger.error(f"Error showing day details: {str(e)}")

    def _hide_day_details(self, event: tk.Event):
        """
        Hide day details tooltip
        
        Args:
            event (tk.Event): Mouse event
        """
        try:
            if hasattr(self, 'day_tooltip'):
                self.day_tooltip.destroy()
                del self.day_tooltip

        except Exception as e:
            logger.error(f"Error hiding day details: {str(e)}")

    def _schedule_tasks(self):
        """Schedule recurring maintenance tasks"""
        try:
            # Create scheduling dialog
            dialog = tk.Toplevel(self)
            dialog.title("Schedule Maintenance Tasks")
            dialog.geometry("600x800")
            dialog.transient(self)
            dialog.grab_set()

            # Create notebook for task categories
            notebook = ttk.Notebook(dialog)
            notebook.pack(fill='both', expand=True, padx=10, pady=10)

            # Daily tasks tab
            daily_frame = ttk.Frame(notebook)
            notebook.add(daily_frame, text="Daily Tasks")
            self._create_task_scheduler(
                daily_frame,
                self._get_daily_tasks(),
                "daily"
            )

            # Weekly tasks tab
            weekly_frame = ttk.Frame(notebook)
            notebook.add(weekly_frame, text="Weekly Tasks")
            self._create_task_scheduler(
                weekly_frame,
                self._get_weekly_tasks(),
                "weekly"
            )

            # Monthly tasks tab
            monthly_frame = ttk.Frame(notebook)
            notebook.add(monthly_frame, text="Monthly Tasks")
            self._create_task_scheduler(
                monthly_frame,
                self._get_monthly_tasks(),
                "monthly"
            )

            # Add save button
            ttk.Button(
                dialog,
                text="Save Schedule",
                command=lambda: self._save_task_schedule(dialog),
                style="Primary.TButton"
            ).pack(pady=10)

            logger.debug("Task scheduler created")

        except Exception as e:
            logger.error(f"Error creating task scheduler: {str(e)}")
            self._handle_error("Failed to create task scheduler")

    def _create_task_scheduler(self, parent: ttk.Frame, tasks: List[dict], frequency: str):
        """
        Create task scheduling interface
        
        Args:
            parent (ttk.Frame): Parent frame
            tasks (List[dict]): List of tasks
            frequency (str): Task frequency
        """
        try:
            # Create scrolled frame
            canvas = tk.Canvas(parent)
            scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            # Pack canvas and scrollbar
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            # Add tasks
            for task in tasks:
                task_frame = ttk.Frame(scrollable_frame)
                task_frame.pack(fill='x', pady=5)

                # Task name
                ttk.Label(
                    task_frame,
                    text=task['name'],
                    font=("Helvetica", 10, "bold")
                ).pack(anchor='w')

                # Time selection
                time_frame = ttk.Frame(task_frame)
                time_frame.pack(fill='x', pady=5)

                if frequency == "daily":
                    self._add_daily_scheduler(time_frame, task)
                elif frequency == "weekly":
                    self._add_weekly_scheduler(time_frame, task)
                else:  # monthly
                    self._add_monthly_scheduler(time_frame, task)

                # Notification options
                notify_frame = ttk.Frame(task_frame)
                notify_frame.pack(fill='x')

                task['notify_email'] = tk.BooleanVar(value=True)
                ttk.Checkbutton(
                    notify_frame,
                    text="Email",
                    variable=task['notify_email']
                ).pack(side='left', padx=5)

                task['notify_sms'] = tk.BooleanVar(value=False)
                ttk.Checkbutton(
                    notify_frame,
                    text="SMS",
                    variable=task['notify_sms']
                ).pack(side='left', padx=5)

                # Add separator
                ttk.Separator(scrollable_frame).pack(fill='x', pady=10)

        except Exception as e:
            logger.error(f"Error creating task scheduler: {str(e)}")
            raise

    def _add_daily_scheduler(self, parent: ttk.Frame, task: dict):
        """
        Add daily scheduling options
        
        Args:
            parent (ttk.Frame): Parent frame
            task (dict): Task dictionary
        """
        try:
            ttk.Label(
                parent,
                text="Time:"
            ).pack(side='left', padx=5)

            # Hour selection
            task['hour'] = tk.StringVar(value="09")
            hour_spin = ttk.Spinbox(
                parent,
                from_=0,
                to=23,
                width=3,
                format="%02.0f",
                textvariable=task['hour']
            )
            hour_spin.pack(side='left')

            ttk.Label(
                parent,
                text=":"
            ).pack(side='left')

            # Minute selection
            task['minute'] = tk.StringVar(value="00")
            minute_spin = ttk.Spinbox(
                parent,
                from_=0,
                to=59,
                width=3,
                format="%02.0f",
                textvariable=task['minute']
            )
            minute_spin.pack(side='left')

        except Exception as e:
            logger.error(f"Error adding daily scheduler: {str(e)}")
            raise

    def _add_weekly_scheduler(self, parent: ttk.Frame, task: dict):
        """
        Add weekly scheduling options
        
        Args:
            parent (ttk.Frame): Parent frame
            task (dict): Task dictionary
        """
        try:
            # Day selection
            ttk.Label(
                parent,
                text="Day:"
            ).pack(side='left', padx=5)

            task['day'] = tk.StringVar(value="Monday")
            day_combo = ttk.Combobox(
                parent,
                values=[
                    "Monday", "Tuesday", "Wednesday",
                    "Thursday", "Friday", "Saturday", "Sunday"
                ],
                textvariable=task['day'],
                state='readonly',
                width=10
            )
            day_combo.pack(side='left', padx=5)

            # Time selection
            ttk.Label(
                parent,
                text="Time:"
            ).pack(side='left', padx=5)

            task['hour'] = tk.StringVar(value="09")
            hour_spin = ttk.Spinbox(
                parent,
                from_=0,
                to=23,
                width=3,
                format="%02.0f",
                textvariable=task['hour']
            )
            hour_spin.pack(side='left')

            ttk.Label(
                parent,
                text=":"
            ).pack(side='left')

            task['minute'] = tk.StringVar(value="00")
            minute_spin = ttk.Spinbox(
                parent,
                from_=0,
                to=59,
                width=3,
                format="%02.0f",
                textvariable=task['minute']
            )
            minute_spin.pack(side='left')

        except Exception as e:
            logger.error(f"Error adding weekly scheduler: {str(e)}")
            raise

    def _add_monthly_scheduler(self, parent: ttk.Frame, task: dict):
        """
        Add monthly scheduling options
        
        Args:
            parent (ttk.Frame): Parent frame
            task (dict): Task dictionary
        """
        try:
            # Day selection
            ttk.Label(
                parent,
                text="Day:"
            ).pack(side='left', padx=5)

            task['day'] = tk.StringVar(value="1")
            days = ["1st", "15th", "Last"] + [f"{i}th" for i in range(2, 29)]
            day_combo = ttk.Combobox(
                parent,
                values=days,
                textvariable=task['day'],
                state='readonly',
                width=10
            )
            day_combo.pack(side='left', padx=5)

            # Time selection
            ttk.Label(
                parent,
                text="Time:"
            ).pack(side='left', padx=5)

            task['hour'] = tk.StringVar(value="09")
            hour_spin = ttk.Spinbox(
                parent,
                from_=0,
                to=23,
                width=3,
                format="%02.0f",
                textvariable=task['hour']
            )
            hour_spin.pack(side='left')

            ttk.Label(
                parent,
                text=":"
            ).pack(side='left')

            task['minute'] = tk.StringVar(value="00")
            minute_spin = ttk.Spinbox(
                parent,
                from_=0,
                to=59,
                width=3,
                format="%02.0f",
                textvariable=task['minute']
            )
            minute_spin.pack(side='left')

        except Exception as e:
            logger.error(f"Error adding monthly scheduler: {str(e)}")
            raise

    def _save_task_schedule(self, dialog: tk.Toplevel):
        """
        Save task schedule
        
        Args:
            dialog (tk.Toplevel): Schedule dialog window
        """
        try:
            schedule = {
                'daily': [],
                'weekly': [],
                'monthly': []
            }

            # Collect daily tasks
            for task in self._get_daily_tasks():
                if 'hour' in task and 'minute' in task:
                    schedule['daily'].append({
                        'name': task['name'],
                        'hour': int(task['hour'].get()),
                        'minute': int(task['minute'].get()),
                        'notify_email': task['notify_email'].get(),
                        'notify_sms': task['notify_sms'].get()
                    })

            # Collect weekly tasks
            for task in self._get_weekly_tasks():
                if 'day' in task:
                    schedule['weekly'].append({
                        'name': task['name'],
                        'day': task['day'].get(),
                        'hour': int(task['hour'].get()),
                        'minute': int(task['minute'].get()),
                        'notify_email': task['notify_email'].get(),
                        'notify_sms': task['notify_sms'].get()
                    })

            # Collect monthly tasks
            for task in self._get_monthly_tasks():
                if hasattr(task, 'day'):
                    schedule['monthly'].append({
                        'name': task['name'],
                        'day': task['day'].get(),
                        'hour': int(task['hour'].get()),
                        'minute': int(task['minute'].get()),
                        'notify_email': task['notify_email'].get(),
                        'notify_sms': task['notify_sms'].get()
                    })

            # Save schedule
            self._save_schedule_to_file(schedule)

            # Update task checking
            self._schedule_task_check()

            # Close dialog
            dialog.destroy()

            # Show confirmation
            messagebox.showinfo(
                "Success",
                "Task schedule saved successfully!"
            )

            logger.info("Task schedule saved")

        except Exception as e:
            logger.error(f"Error saving task schedule: {str(e)}")
            self._handle_error("Failed to save task schedule")

    def check_scheduled_tasks(self):
        """Check and execute scheduled tasks"""
        try:
            current_time = datetime.now()
            schedule = self._load_schedule_from_file()

            # Check daily tasks
            for task in schedule['daily']:
                if self._is_task_due(task, current_time, 'daily'):
                    self._execute_scheduled_task(task)

            # Check weekly tasks
            for task in schedule['weekly']:
                if self._is_task_due(task, current_time, 'weekly'):
                    self._execute_scheduled_task(task)

            # Check monthly tasks
            for task in schedule['monthly']:
                if self._is_task_due(task, current_time, 'monthly'):
                    self._execute_scheduled_task(task)

            # Schedule next check
            self.task_check_id = self.after(60000, self.check_scheduled_tasks)

            logger.debug("Scheduled tasks checked")

        except Exception as e:
            logger.error(f"Error checking scheduled tasks: {str(e)}")
            # Ensure next check is scheduled even if error occurs
            self.task_check_id = self.after(60000, self.check_scheduled_tasks)

    def _is_task_due(self, task: dict, current_time: datetime, frequency: str) -> bool:
        """
        Check if task is due for execution
        
        Args:
            task (dict): Task dictionary
            current_time (datetime): Current time
            frequency (str): Task frequency
            
        Returns:
            bool: True if task is due
        """
        try:
            # Check hour and minute
            if current_time.hour != task['hour'] or current_time.minute != task['minute']:
                return False

            if frequency == 'daily':
                return True
            elif frequency == 'weekly':
                # Convert day name to weekday number (0 = Monday)
                day_map = {
                    'Monday': 0, 'Tuesday': 1, 'Wednesday': 2,
                    'Thursday': 3, 'Friday': 4, 'Saturday': 5, 'Sunday': 6
                }
                return current_time.weekday() == day_map[task['day']]
            else:  # monthly
                if task['day'] == 'Last':
                    # Check if it's the last day of the month
                    next_day = current_time + timedelta(days=1)
                    return next_day.day == 1
                else:
                    # Extract day number from format like "1st", "2nd", etc.
                    day = int(''.join(filter(str.isdigit, task['day'])))
                    return current_time.day == day

        except Exception as e:
            logger.error(f"Error checking if task is due: {str(e)}")
            return False

    def _execute_scheduled_task(self, task: dict):
        """
        Execute scheduled task
        
        Args:
            task (dict): Task dictionary
        """
        try:
            # Record task execution
            execution = {
                'task': task['name'],
                'timestamp': datetime.now().isoformat(),
                'scheduled': True
            }

            # Add to task history
            if not hasattr(self, 'task_history'):
                self.task_history = []
            self.task_history.append(execution)

            # Save task history
            self._save_task_history()

            # Send notifications
            if task.get('notify_email', False):
                self._send_task_notification_email(task)
            if task.get('notify_sms', False):
                self._send_task_notification_sms(task)

            # Update maintenance log
            self._update_maintenance_log(execution)

            logger.info(f"Scheduled task executed: {task['name']}")

        except Exception as e:
            logger.error(f"Error executing scheduled task: {str(e)}")

    def _send_task_notification_email(self, task: dict):
        """
        Send task notification via email
        
        Args:
            task (dict): Task dictionary
        """
        try:
            # Validate email settings
            valid, error = self._validate_email_settings()
            if not valid:
                raise ValueError(error)

            # Create email message
            msg = MIMEMultipart()
            msg['From'] = self.email_config['sender_email']
            msg['To'] = self.customer_info.get('email', self.email_config['sender_email'])
            msg['Subject'] = "Pool Maintenance Task Reminder"

            # Create HTML body
            html = f"""
            <html>
                <head>
                    <style>
                        body {{ font-family: Arial, sans-serif; }}
                        .task {{ 
                            background-color: #E3F2FD;
                            padding: 20px;
                            border-radius: 5px;
                            margin: 20px 0;
                        }}
                        .message {{ font-size: 16px; }}
                        .footer {{ color: #666; font-size: 12px; }}
                    </style>
                </head>
                <body>
                    <div class="task">
                        <h2>Maintenance Task Reminder</h2>
                        <p class="message">
                            The following maintenance task is due:
                            <br><br>
                            <strong>{task['name']}</strong>
                        </p>
                    </div>
                    <p class="footer">
                        This is an automated reminder from your Pool Chemistry Monitor.
                    </p>
                </body>
            </html>
            """
            msg.attach(MIMEText(html, 'html'))

            # Send email
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(
                    self.email_config['sender_email'],
                    self.email_config['app_password']
                )
                server.send_message(msg)

            logger.info(f"Task notification email sent for: {task['name']}")

        except Exception as e:
            logger.error(f"Error sending task notification email: {str(e)}")

    def _send_task_notification_sms(self, task: dict):
        """
        Send task notification via SMS
        
        Args:
            task (dict): Task dictionary
        """
        try:
            # Validate SMS settings
            valid, error = self._validate_sms_settings()
            if not valid:
                raise ValueError(error)

            # Initialize Twilio client
            from twilio.rest import Client
            client = Client(
                self.sms_config['account_sid'],
                self.sms_config['auth_token']
            )

            # Create message
            message = (
                f"Pool Maintenance Reminder:\n"
                f"Task due: {task['name']}\n"
                f"Please complete this maintenance task."
            )

            # Send message
            message = client.messages.create(
                body=message,
                from_=self.sms_config['twilio_number'],
                to=self.customer_info.get('phone')
            )

            logger.info(f"Task notification SMS sent for: {task['name']}")

        except Exception as e:
            logger.error(f"Error sending task notification SMS: {str(e)}")

    def _manage_task_history(self):
        """Manage and clean up task history"""
        try:
            if not hasattr(self, 'task_history'):
                return

            # Get retention period
            retention_days = self.settings.get('task_retention', 90)
            cutoff_date = datetime.now() - timedelta(days=retention_days)

            # Filter old entries
            self.task_history = [
                task for task in self.task_history
                if datetime.fromisoformat(task['timestamp']) > cutoff_date
            ]

            # Save updated history
            self._save_task_history()

            logger.debug(f"Task history cleaned up (retention: {retention_days} days)")

        except Exception as e:
            logger.error(f"Error managing task history: {str(e)}")

    def _export_task_history(self):
        """Export task history to file"""
        try:
            # Get file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[
                    ("Excel files", "*.xlsx"),
                    ("CSV files", "*.csv"),
                    ("JSON files", "*.json")
                ],
                title="Export Task History"
            )
            
            if not file_path:
                return

            # Get file extension
            ext = Path(file_path).suffix.lower()

            if ext == '.xlsx':
                self._export_task_history_excel(file_path)
            elif ext == '.csv':
                self._export_task_history_csv(file_path)
            else:  # .json
                self._export_task_history_json(file_path)

            messagebox.showinfo(
                "Success",
                "Task history exported successfully!"
            )

            logger.info(f"Task history exported to: {file_path}")

        except Exception as e:
            logger.error(f"Error exporting task history: {str(e)}")
            self._handle_error("Failed to export task history")

    def _export_task_history_excel(self, file_path: str):
        """
        Export task history to Excel
        
        Args:
            file_path (str): Export file path
        """
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Task History"

            # Add headers
            headers = ["Date", "Time", "Task", "Scheduled"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="E3F2FD",
                    end_color="E3F2FD",
                    fill_type="solid"
                )

            # Add data
            for row, task in enumerate(self.task_history, 2):
                dt = datetime.fromisoformat(task['timestamp'])
                
                ws.cell(row=row, column=1, value=dt.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=dt.strftime('%H:%M:%S'))
                ws.cell(row=row, column=3, value=task['task'])
                ws.cell(row=row, column=4, value=task.get('scheduled', False))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Save workbook
            wb.save(file_path)

        except Exception as e:
            logger.error(f"Error exporting task history to Excel: {str(e)}")
            raise

    def _export_task_history_csv(self, file_path: str):
        """
        Export task history to CSV
        
        Args:
            file_path (str): Export file path
        """
        try:
            # Write CSV file
            with open(file_path, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers
                writer.writerow(["Date", "Time", "Task", "Scheduled"])
                
                # Write data
                for task in self.task_history:
                    dt = datetime.fromisoformat(task['timestamp'])
                    writer.writerow([
                        dt.strftime('%Y-%m-%d'),
                        dt.strftime('%H:%M:%S'),
                        task['task'],
                        task.get('scheduled', False)
                    ])

        except Exception as e:
            logger.error(f"Error exporting task history to CSV: {str(e)}")
            raise

    def _export_task_history_json(self, file_path: str):
        """
        Export task history to JSON
        
        Args:
            file_path (str): Export file path
        """
        try:
            # Format data for export
            export_data = {
                'metadata': {
                    'exported_at': datetime.now().isoformat(),
                    'total_tasks': len(self.task_history)
                },
                'tasks': self.task_history
            }

            # Write JSON file
            with open(file_path, 'w') as f:
                json.dump(export_data, f, indent=4)

        except Exception as e:
            logger.error(f"Error exporting task history to JSON: {str(e)}")
            raise

    def _cleanup_old_data(self):
        """Clean up old data files"""
        try:
            # Clean up task history
            self._manage_task_history()

            # Clean up backups
            backup_dir = Path(self.settings.get('backup_path', './backup'))
            if backup_dir.exists():
                self._cleanup_old_backups(backup_dir, 'auto')
                self._cleanup_old_backups(backup_dir, 'manual')

            # Clean up logs
            log_dir = Path("logs")
            if log_dir.exists():
                self._cleanup_old_logs(log_dir)

            # Clean up temporary files
            temp_dir = Path("temp")
            if temp_dir.exists():
                shutil.rmtree(temp_dir)
                temp_dir.mkdir(exist_ok=True)

            logger.info("Data cleanup completed")

        except Exception as e:
            logger.error(f"Error cleaning up old data: {str(e)}")

    def _cleanup_old_logs(self, log_dir: Path):
        """
        Clean up old log files
        
        Args:
            log_dir (Path): Log directory path
        """
        try:
            # Get all log files
            log_files = list(log_dir.glob("*.log"))
            
            # Sort by modification time
            log_files.sort(key=lambda x: x.stat().st_mtime)
            
            # Keep only last 30 days of logs
            cutoff_date = datetime.now() - timedelta(days=30)
            
            for log_file in log_files:
                if datetime.fromtimestamp(log_file.stat().st_mtime) < cutoff_date:
                    log_file.unlink()
                    logger.debug(f"Removed old log file: {log_file}")

        except Exception as e:
            logger.error(f"Error cleaning up logs: {str(e)}")
            raise

    def _compress_old_data(self):
        """Compress old data files"""
        try:
            # Get data directory
            data_dir = Path("data")
            if not data_dir.exists():
                return

            # Create archive directory
            archive_dir = Path("archive")
            archive_dir.mkdir(exist_ok=True)

            # Create archive name with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            archive_name = f"data_archive_{timestamp}.zip"
            archive_path = archive_dir / archive_name

            # Create ZIP file
            with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Add all files from data directory
                for file_path in data_dir.rglob('*'):
                    if file_path.is_file():
                        arcname = file_path.relative_to(data_dir)
                        zipf.write(file_path, arcname)

            # Remove old archives (keep last 5)
            archives = list(archive_dir.glob("data_archive_*.zip"))
            archives.sort(key=lambda x: x.stat().st_mtime)
            
            while len(archives) > 5:
                oldest = archives.pop(0)
                oldest.unlink()
                logger.debug(f"Removed old archive: {oldest}")

            logger.info(f"Data compressed to archive: {archive_path}")

        except Exception as e:
            logger.error(f"Error compressing old data: {str(e)}")

    def _validate_data_integrity(self) -> Tuple[bool, List[str]]:
        """
        Validate data file integrity
        
        Returns:
            Tuple[bool, List[str]]: (is_valid, error_messages)
        """
        try:
            errors = []

            # Check task history
            if hasattr(self, 'task_history'):
                for task in self.task_history:
                    if 'timestamp' not in task:
                        errors.append("Task missing timestamp")
                    if 'task' not in task:
                        errors.append("Task missing name")
                    try:
                        datetime.fromisoformat(task['timestamp'])
                    except:
                        errors.append(f"Invalid timestamp format: {task['timestamp']}")

            # Check schedule
            schedule = self._load_schedule_from_file()
            for frequency in ['daily', 'weekly', 'monthly']:
                for task in schedule.get(frequency, []):
                    if 'name' not in task:
                        errors.append(f"Scheduled {frequency} task missing name")
                    if 'hour' not in task or 'minute' not in task:
                        errors.append(f"Scheduled {frequency} task missing time")

            return len(errors) == 0, errors

        except Exception as e:
            logger.error(f"Error validating data integrity: {str(e)}")
            return False, ["Validation error occurred"]

    def _repair_data_files(self, errors: List[str]) -> bool:
        """
        Attempt to repair corrupted data files
        
        Args:
            errors (List[str]): List of validation errors
            
        Returns:
            bool: True if repair was successful
        """
        try:
            # Create backup before repair
            self._create_backup('repair')

            # Fix task history
            if hasattr(self, 'task_history'):
                # Remove invalid entries
                valid_tasks = []
                for task in self.task_history:
                    if ('timestamp' in task and 
                        'task' in task and 
                        self._is_valid_timestamp(task['timestamp'])):
                        valid_tasks.append(task)
                
                self.task_history = valid_tasks
                self._save_task_history()

            # Fix schedule
            schedule = self._load_schedule_from_file()
            fixed_schedule = {
                'daily': [],
                'weekly': [],
                'monthly': []
            }

            for frequency in ['daily', 'weekly', 'monthly']:
                for task in schedule.get(frequency, []):
                    if self._is_valid_scheduled_task(task, frequency):
                        fixed_schedule[frequency].append(task)

            self._save_schedule_to_file(fixed_schedule)

            # Validate after repair
            is_valid, remaining_errors = self._validate_data_integrity()
            
            if remaining_errors:
                logger.warning("Some errors could not be fixed automatically")
                return False
                
            logger.info("Data files repaired successfully")
            return True

        except Exception as e:
            logger.error(f"Error repairing data files: {str(e)}")
            return False

    def _is_valid_timestamp(self, timestamp: str) -> bool:
        """
        Validate timestamp format
        
        Args:
            timestamp (str): Timestamp to validate
            
        Returns:
            bool: True if valid
        """
        try:
            datetime.fromisoformat(timestamp)
            return True
        except:
            return False

    def _is_valid_scheduled_task(self, task: dict, frequency: str) -> bool:
        """
        Validate scheduled task
        
        Args:
            task (dict): Task to validate
            frequency (str): Task frequency
            
        Returns:
            bool: True if valid
        """
        try:
            # Check required fields
            if not all(key in task for key in ['name', 'hour', 'minute']):
                return False

            # Validate time
            if not (0 <= task['hour'] <= 23 and 0 <= task['minute'] <= 59):
                return False

            # Validate frequency-specific fields
            if frequency == 'weekly':
                if 'day' not in task:
                    return False
                valid_days = [
                    'Monday', 'Tuesday', 'Wednesday',
                    'Thursday', 'Friday', 'Saturday', 'Sunday'
                ]
                if task['day'] not in valid_days:
                    return False
            elif frequency == 'monthly':
                if 'day' not in task:
                    return False
                if task['day'] != 'Last':
                    try:
                        day = int(''.join(filter(str.isdigit, task['day'])))
                        if not 1 <= day <= 28:
                            return False
                    except:
                        return False

            return True

        except Exception as e:
            logger.error(f"Error validating scheduled task: {str(e)}")
            return False

    def _cleanup(self):
        """Perform cleanup before application exit"""
        try:
            # Save current data
            self._save_data()
            self._save_settings_to_file()
            self._save_task_history()

            # Clean up resources
            if hasattr(self, 'weather_update_id'):
                self.after_cancel(self.weather_update_id)
            
            if hasattr(self, '_task_check_id'):
                self.after_cancel(self._task_check_id)
            
            if hasattr(self, 'fig'):
                plt.close(self.fig)

            # Close any open tooltips
            if hasattr(self, 'tooltips'):
                for tooltip in self.tooltips:
                    try:
                        tooltip.destroy()
                    except:
                        pass

            # Clean up temporary files
            temp_dir = Path("temp")
            if temp_dir.exists():
                shutil.rmtree(temp_dir)

            # Compress old data
            self._compress_old_data()

            logger.info("Application cleanup completed")

        except Exception as e:
            logger.error(f"Error during cleanup: {str(e)}")

    def __del__(self):
        """Class destructor"""
        try:
            self._cleanup()
        except:
            pass

# Class initialization complete
logger.info("PoolApp class definition completed")

def main():
    """Main application entry point"""
    try:
        # Configure root logger
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_dir / 'pool_app.log'),
                logging.StreamHandler()
            ]
        )
        
        logger.info("Starting Deep Blue Chemistry Pool Management System")
        logger.info(f"Python version: {platform.python_version()}")
        logger.info(f"Operating system: {platform.system()} {platform.release()}")
        
        # Create data directories
        Path("data").mkdir(exist_ok=True)
        Path("backup").mkdir(exist_ok=True)
        Path("temp").mkdir(exist_ok=True)
        
        # Initialize water tester
        logger.debug("Initializing WaterTester")
        tester = WaterTester()
        
        # Create and run main application
        logger.debug("Creating main application window")
        app = PoolApp(tester)
        
        # Configure main window
        app.title("Deep Blue Pool Chemistry Monitor")
        app.state('zoomed')  # Start maximized
        app.minsize(1200, 800)
        
        # Center window on screen
        screen_width = app.winfo_screenwidth()
        screen_height = app.winfo_screenheight()
        window_width = 1400
        window_height = 1000
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        app.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Set application icon if available
        try:
            app.iconbitmap("assets/icon.ico")
        except:
            logger.warning("Application icon not found")
        
        logger.info("Starting main application loop")
        app.mainloop()
        
    except Exception as e:
        logger.critical(f"Critical error in main application: {str(e)}")
        messagebox.showerror(
            "Critical Error",
            "Application failed to start. Please check the logs for details."
        )
        raise
        
    finally:
        try:
            # Perform cleanup
            logger.info("Shutting down application")
            if 'app' in locals():
                app._cleanup()
            
            # Clean up temporary files
            temp_dir = Path("temp")
            if temp_dir.exists():
                shutil.rmtree(temp_dir)
                
            # Shutdown logging
            logging.shutdown()
            
        except Exception as e:
            print(f"Error during cleanup: {str(e)}")

if __name__ == "__main__":
    main()
