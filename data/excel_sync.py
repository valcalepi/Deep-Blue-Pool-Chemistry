import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import requests
import os
from dotenv import load_dotenv

load_dotenv()
API_KEY = os.getenv("WEATHER_API_KEY")
LOCATION = "St. Cloud"

EXCEL_FILE = 'pool_log.xlsx'
SHEET_NAME = 'PoolData'
THRESHOLDS = {
    'pH': (7.2, 7.8),
    'chlorine': (1.0, 3.0),
    'temperature': (70, 90)
}

def get_weather_adjustments():
    try:
        url = f"http://api.weatherapi.com/v1/forecast.json?key={API_KEY}&q={LOCATION}&days=3"
        response = requests.get(url)
        data = response.json()
        forecast = data['forecast']['forecastday']

        rain_expected = any(day['day']['daily_chance_of_rain'] > 50 for day in forecast)
        uv_index = forecast[0]['day']['uv']

        adjustments = []
        if rain_expected:
            adjustments.append("Rain expected: increase stabilizer or shock dose.")
        if uv_index > 7:
            adjustments.append("High UV: consider adding stabilizer to reduce chlorine loss.")
        return " | ".join(adjustments) if adjustments else "No weather-based adjustments needed."
    except Exception:
        return "Weather data unavailable."

def append_to_excel(data):
    timestamp = datetime.now().isoformat()
    weather_note = get_weather_adjustments()
    row = {'Timestamp': timestamp, **data, 'Notes': weather_note}

    try:
        book = load_workbook(EXCEL_FILE)
        sheet = book[SHEET_NAME]
    except FileNotFoundError:
        df = pd.DataFrame([row])
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        return

    next_row = sheet.max_row + 1
    for col_index, key in enumerate(row.keys(), start=1):
        sheet.cell(row=next_row, column=col_index).value = row[key]
        if key in THRESHOLDS:
            val = row[key]
            low, high = THRESHOLDS[key]
            if val < low or val > high:
                sheet.cell(row=next_row, column=col_index).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    book.save(EXCEL_FILE)

